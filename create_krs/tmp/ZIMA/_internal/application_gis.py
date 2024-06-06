import openpyxl
import re
import well_data
import sys
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QIntValidator, QDoubleValidator
from PyQt5.QtWidgets import QWidget, QLabel, QComboBox, QLineEdit, QGridLayout, QTabWidget, QMainWindow, QPushButton, \
    QMessageBox, QApplication, QTableWidget, QTableWidgetItem, QVBoxLayout, QHeaderView




class TabPage_SO_pvr(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        self.validator_int = QIntValidator(0, 600)
        self.validator_float = QDoubleValidator(0.87, 1.65, 2)

        self.number_brigada_label = QLabel('Номер бригады', self)
        self.number_brigada_combo = QComboBox(self)
        brigada_list = list(well_data.dict_telephon.keys())
        self.number_brigada_combo.addItems(brigada_list)


        self.number_telephone_label = QLabel('номер телефона, self')
        self.number_telephone_edit = QLineEdit(self)


        self.date_new_label = QLabel('Дата заявки', self)
        self.date_new_edit = QLineEdit(self)
        self.date_new_edit.setText(f'{well_data.current_date}')

        self.time_new_label = QLabel('Время заявки', self)
        self.time_new_edit = QLineEdit(self)
        self.time_new_edit.setText(f'12:00')

        self.work_label = QLabel("Ранее проведенные работы:", self)
        self.work_edit = QLineEdit(self)

        self.nkt_label = QLabel("Внешний диаметр НКТ:", self)
        self.nkt_edit = QLineEdit(self)
        if well_data.column_additional is False and well_data.column_diametr._value > 110 or \
                (well_data.column_additional and well_data.column_additional_diametr._value > 110):
            self.nkt_edit.setText('73')
        else:
            self.nkt_edit.setText('60')


        self.nkt_shoe_label = QLabel("Башмак НКТ:", self)
        self.nkt_shoe_edit = QLineEdit(self)

        self.nkt_com_label = QLabel("Компоновка НКТ (низ НКТ):", self)
        self.nkt_com_edit = QLineEdit(self)
        self.nkt_com_edit.setText('воронка')

        self.paker_type_label = QLabel("Тип пакера", self)
        self.paker_type = QLineEdit(self)

        self.pakerLabel = QLabel("глубина пакера", self)
        self.paker_depth = QLineEdit(self)

        self.fluid_label = QLabel("уд.вес жидкости глушения", self)
        self.fluid_edit = QLineEdit(self)

        self.labelType = QLabel("Кровля записи", self)
        self.lineEditType = QLineEdit(self)
        self.lineEditType.setClearButtonEnabled(True)

        self.labelType2 = QLabel("Подошва записи", self)
        self.lineEditType2 = QLineEdit(self)
        self.lineEditType2.setClearButtonEnabled(True)

        self.labelGeores = QLabel("вид исследования", self)
        self.ComboBoxGeophygist = QComboBox(self)
        self.ComboBoxGeophygist.addItems(
            ['Гироскоп', 'АКЦ', 'АКЦ + СГДТ', 'СГДТ', 'ИНГК', 'ЭМДС', 'ПТС', 'РК', 'ГК и ЛМ',
             'отбивка забоя', 'привязка', 'ВП', 'РГД по колонне', 'РГД по НКТ', 'Цем желонки'])
        self.ComboBoxGeophygist.currentTextChanged.connect(self.geophygist_data)

        self.grid = QGridLayout(self)

        self.grid.addWidget(self.number_brigada_label, 7, 2)
        self.grid.addWidget(self.number_brigada_combo, 8, 2)

        self.grid.addWidget(self.number_telephone_label, 7, 3)
        self.grid.addWidget(self.number_telephone_edit, 8, 3)

        self.grid.addWidget(self.date_new_label, 9, 3)
        self.grid.addWidget(self.date_new_edit, 10, 3)

        self.grid.addWidget(self.time_new_label, 9, 4)
        self.grid.addWidget(self.time_new_edit, 10, 4)

        self.grid.addWidget(self.work_label, 11, 3, 1, 2)
        self.grid.addWidget(self.work_edit, 12, 3, 1, 4)

        self.grid.addWidget(self.nkt_label, 13, 2)
        self.grid.addWidget(self.nkt_edit, 14, 2)

        self.grid.addWidget(self.nkt_shoe_label, 13, 3)
        self.grid.addWidget(self.nkt_shoe_edit, 14, 3)

        self.grid.addWidget(self.nkt_com_label, 13, 4)
        self.grid.addWidget(self.nkt_com_edit, 14, 4)


        self.grid.addWidget(self.paker_type_label, 15, 2)
        self.grid.addWidget(self.paker_type, 16, 2)

        self.grid.addWidget(self.pakerLabel, 15, 3)
        self.grid.addWidget(self.paker_depth, 16, 3)

        self.grid.addWidget(self.fluid_label, 17, 3)
        self.grid.addWidget(self.fluid_edit, 18, 3)

        self.grid.addWidget(self.labelGeores, 22, 2)
        self.grid.addWidget(self.ComboBoxGeophygist, 23, 2)

        self.grid.addWidget(self.labelType, 22, 3)
        self.grid.addWidget(self.lineEditType, 23, 3)

        self.grid.addWidget(self.labelType2, 22, 4)
        self.grid.addWidget(self.lineEditType2, 23, 4)

        self.number_brigada_combo.currentTextChanged.connect(self.update_brigade)

    def update_brigade(self, index):
        self.number_telephone_edit.setText(str(well_data.dict_telephon[self.number_brigada_combo.currentText()]))
    def geophygist_data(self):

        if self.ComboBoxGeophygist.currentText() in ['Гироскоп', 'АКЦ', 'ЭМДС', 'ПТС', 'РК', 'ГК и ЛМ']:
            self.lineEditType.setText('0')
            self.lineEditType2.setText(f'{well_data.current_bottom}')

class TabWidget(QTabWidget):
    def __init__(self):
        super().__init__()
        self.addTab(TabPage_SO_pvr(self), 'Заявка на ГИС')

class GisApplication(QMainWindow):

    def __init__(self, table_pvr, parent=None):
        super(QMainWindow, self).__init__()

        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        self.table_pvr = table_pvr
        self.tabWidget = TabWidget()

        self.tableWidget = QTableWidget(0, 3)
        self.tableWidget.setHorizontalHeaderLabels(
            ["Геофизические исследования", "Кровля записи", "Подошва записи"])
        for i in range(3):
            self.tableWidget.horizontalHeader().setSectionResizeMode(i, QHeaderView.Stretch)


        self.tableWidget.setSortingEnabled(True)
        self.tableWidget.setAlternatingRowColors(True)

        self.buttonAdd = QPushButton('Добавить исследования в таблицу')
        self.buttonAdd.clicked.connect(self.addRowTable)
        self.buttonDel = QPushButton('Удалить интервалы перфорации в таблице')
        self.buttonDel.clicked.connect(self.del_row_table)
        self.buttonadd_work = QPushButton('Создать заявку')
        self.buttonadd_work.clicked.connect(self.add_work, Qt.QueuedConnection)
        self.buttonAddProject = QPushButton('Добавить исследования из плана')
        self.buttonAddProject.clicked.connect(self.addPerfProject)


        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tabWidget, 0, 0, 1, 2)
        vbox.addWidget(self.tableWidget, 1, 0, 1, 2)
        vbox.addWidget(self.buttonAdd, 2, 0)
        vbox.addWidget(self.buttonDel, 2, 1)
        vbox.addWidget(self.buttonadd_work, 3, 0)
        vbox.addWidget(self.buttonAddProject, 3, 1)


    def addPerfProject(self):

        if len(well_data.gis_list) == 0:
            mes = QMessageBox.warning(self, 'Ошибка', 'Исследования в плане работ не найдены')
            return

        for pvr in well_data.gis_list:
            rows = self.tableWidget.rowCount()
            try:
                type_gis = self.geophysic_sel(pvr)
                self.tableWidget.insertRow(rows)
                self.tableWidget.setItem(rows, 0, QTableWidgetItem(type_gis[0]))
                self.tableWidget.setItem(rows, 1, QTableWidgetItem(type_gis[1]))
                self.tableWidget.setItem(rows, 2, QTableWidgetItem(type_gis[2]))
            except:
                mes = QMessageBox.information(self, 'Ошибка', f'Ошибка добавления исследования {pvr}')



    def geophysicalSelect(self, geophysic):
        return geophysic

    def addRowTable(self):

        editType = self.tabWidget.currentWidget().lineEditType.text().replace(',', '.')
        editType2 = self.tabWidget.currentWidget().lineEditType2.text().replace(',', '.')
        researchGis = self.tabWidget.currentWidget().ComboBoxGeophygist.currentText()


        if not editType or not editType2 or not researchGis:
            msg = QMessageBox.information(self, 'Внимание', 'Заполните все поля!')
            return
        if well_data.current_bottom < float(editType2):
            msg = QMessageBox.information(self, 'Внимание', 'глубина исследований ниже текущего забоя')
            return

        self.tableWidget.setSortingEnabled(False)
        rows = self.tableWidget.rowCount()
        self.tableWidget.insertRow(rows)

        self.tableWidget.setItem(rows, 0, QTableWidgetItem(researchGis))
        self.tableWidget.setItem(rows, 1, QTableWidgetItem(editType))
        self.tableWidget.setItem(rows, 2, QTableWidgetItem(editType2))

        self.tableWidget.setSortingEnabled(True)

    def geophysic_sel(self, row):
        print(row)
        if 'АКЦ' in row:
            type = 'АКЦ'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]
        elif 'СГДТ' in row:
            type = 'СГДТ'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]

        elif 'Определение текущей нефтенасыщенности' in row:
            type = 'ИНГК'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]

        elif 'гироскоп' in row:
            type = 'Гироскоп'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]
        elif '2.4.1' in row:
            type = 'РК'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]
        elif '2.6.11' in row:
            type = 'ЭМДС'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]
        elif '2.6.10' in row:
            type = 'ПТС'
            # Удаление всех символов, кроме цифр и тире
            result_string = re.sub(r"[^\d-.]", "", row[-15:]).split('-')
            roof = result_string[0]
            sole = result_string[1]

        # elif 'ГК и ЛМ' in row:
        #     try:
        #         type = 'ГК и ЛМ'
        #         # Удаление всех символов, кроме цифр и тире
        #         result_string = re.sub(r"[^\d-]", "", row[-15:]).split('-')
        #         roof = result_string[0]
        #         sole = result_string[1]
        #     except IndexError:
        #         pass

        elif 'ВП' in row or '№ 2.1.13' in row or 'ГПШ' in row or 'желонк' in row or '2.1.16' in row or '2.1.17' in row \
                or 'РГД по колонне' in row or 'РГД по НКТ' in row or '2.3.2' in row or '2.3.3' in row or '2.3.1' in row \
                or '2.8.1' in row or '2.8.2' in row:
            type = row
            roof = ''
            sole = ''

        return type, roof, sole

    def copy_pvr(self, ws, work_list):
        for row in range(len(work_list)):
            for col in range(41):
                if work_list[row][col]:
                    print(row, col)
                    print(work_list[row][col])
                    ws.cell(row=row + 1, column=col + 1).value = work_list[row][col]
        # Перебираем строки и скрываем те, у которых все значения равны None
        for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
            if all(value is None for value in row[:42]):
                ws.row_dimensions[row_ind+1].hidden = True

    def add_work(self):
        from main import MyWindow

        wb = openpyxl.load_workbook('property_excel/template_gis.xlsx')
        # Выбираем активный лист
        self.ws_pvr = wb.active
        number_brigada = str(self.tabWidget.currentWidget().number_brigada_combo.currentText())
        number_telephone = self.tabWidget.currentWidget().number_telephone_edit.text()
        date_new_edit = self.tabWidget.currentWidget().date_new_edit.text()
        time_new_edit = self.tabWidget.currentWidget().time_new_edit.text()
        work_edit = self.tabWidget.currentWidget().work_edit.text()
        nkt_edit = self.tabWidget.currentWidget().nkt_edit.text()
        nkt_shoe_edit = self.tabWidget.currentWidget().nkt_shoe_edit.text()
        nkt_com_edit = self.tabWidget.currentWidget().nkt_com_edit.text()
        paker_type = self.tabWidget.currentWidget().paker_type.text()
        paker_depth = self.tabWidget.currentWidget().paker_depth.text()
        fluid = self.tabWidget.currentWidget().fluid_edit.text()
        note_to_gis = ''
        if "СКО" in work_edit.lower() or "кислот" in work_edit.lower() or 'опз' in work_edit.lower():
            note_to_gis = 'Работать на противокислотном кабеле '

        rows = self.tableWidget.rowCount()
        geophysic_dict = {
            'АКЦ': "ЗАДАЧА 2.7.1", 'СГДТ': 'ЗАДАЧА 2.7.2', 'АКЦ + СГДТ': 'ЗАДАЧА 2.7.3', 'ИНГК': 'ЗАДАЧА 2.4.3',
        'Гироскоп': 'ЗАДАЧА 2.7.4', 'РК': "ЗАДАЧА 2.4.1", 'ЭМДС': 'ЗАДАЧА 2.6.11', 'ПТС': 'ЗАДАЧА 2.6.10',
            'отбивка забоя': 'ЗАДАЧА 2.8.2', 'привязка': 'ЗАДАЧА 2.8.1'}

        gis_list = []

        for row in range(rows):
            item = self.tableWidget.item(row, 0)
            edit1 = self.tableWidget.item(row, 1)
            edit2 = self.tableWidget.item(row, 2)
            if item and edit1 and edit2:
                type_gis = item.text()
                roof_gis = edit1.text()
                sole_gis = edit2.text()

            if type_gis not in ['Гироскоп', 'АКЦ', 'АКЦ + СГДТ', 'СГДТ', 'ИНГК', 'ЭМДС', 'ПТС', 'РК', 'ГК и ЛМ']:
                if 'ВП' in type_gis or 'ГПШ' in type_gis or 'ВПШ' in type_gis:
                    task_gis = 'ЗАДАЧА 2.9.4'
                elif '№ 2.1.13' in type_gis:
                    task_gis = 'ЗАДАЧА 2.1.13'
                elif '№ 2.1.16' in type_gis:
                    task_gis = 'ЗАДАЧА 2.1.16'
                elif '№ 2.1.17' in type_gis:
                    task_gis = 'ЗАДАЧА 2.1.17'
                elif '№ 2.1.17' in type_gis:
                    task_gis = 'ЗАДАЧА 2.1.17'
                elif 'Задача 9.5.2' in type_gis:
                    task_gis = 'ЗАДАЧА 9.5.2'
                elif ' 2.3.2' in type_gis:
                    task_gis = 'ЗАДАЧА  2.3.2'
                elif 'привязка' in type_gis:
                    task_gis = 'ЗАДАЧА 2.8.1'
                elif 'отбивка забоя' in type_gis:
                    task_gis = 'ЗАДАЧА 2.8.2'


            else:
                task_gis = geophysic_dict[type_gis]

            row_list = [None, task_gis, None, None, type_gis, None, None, None, None, None, None, None, None,
                        None,
                        None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                        None, None,
                        None, None, f'{roof_gis}', None, None, None, f'{sole_gis}', None, None]
            gis_list.append(row_list)

        work_list = self.application_pvr_def(number_brigada, number_telephone, date_new_edit, time_new_edit, work_edit, nkt_edit,
                                             nkt_shoe_edit,  nkt_com_edit, paker_type, paker_depth, fluid, note_to_gis)

        for index, row in enumerate(gis_list):
            work_list[15 + index] = row


        self.copy_pvr(self.ws_pvr, work_list)

        # MyWindow.copy_pz(self, self.ws_pvr, self.table_pvr, 'application_pvr', 42)

        well_data.pause = False
        self.close()
        well_data.gis_list = []
        well_data.pvr_row = []
        self.ws_pvr.print_area = f'B1:AP{85}'

        filenames = f'{well_data.well_number._value} {well_data.well_area._value} ГИС {well_data.current_date}.xlsx'
        path = 'D:\Documents\Desktop\ГТМ\заявки ГИС'
        full_path = path + "/" + filenames
        if wb:
            wb.close()
            MyWindow.saveFileDialog(self, wb, full_path)
            # wb2.save(full_path)
            print(f"Table data saved to Excel {full_path} {well_data.number_dp}")
        if wb:
            wb.close()

    def del_row_table(self):
        row = self.tableWidget.currentRow()
        if row == -1:
            msg = QMessageBox.information(self, 'Внимание', 'Выберите строку для удаления')
            return
        self.tableWidget.removeRow(row)


    def application_pvr_def(self, number_brigada, number_telephone, date_new_edit, time_new_edit, work_edit,
                            nkt_edit, nkt_shoe_edit,
                            nkt_com_edit, paker_type, paker_depth, fluid, note_to_gis):

        column_data = f'{well_data.column_diametr._value}мм x {well_data.column_wall_thickness._value} в инт ' \
                      f'0-{well_data.shoe_column._value}м'
        if well_data.column_additional:
            column_data_add = f'{well_data.column_additional_diametr._value}мм x ' \
                              f'{well_data.column_additional_wall_thickness._value} в инт ' \
                          f'{well_data.head_column_additional._value}-{well_data.shoe_column_additional._value}м'
        else:
            column_data_add = ''
        pressuar = well_data.dict_category[list(well_data.dict_category.keys())[0]]['по давлению'].data_pressuar

        conductor = f'{well_data.column_conductor_diametr._value}мм x {well_data.column_conductor_wall_thickness._value} в инт ' \
                      f'0-{well_data.column_conductor_lenght._value}м'
        if nkt_edit != '':
            nkt_edit_vn = float(nkt_edit) - 2 * 5.5
        else:
            nkt_edit_vn = ''
        value_list = [
            ['З А Я В К А', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, '№', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None,None, None],
            ['на проведение геофизических исследований в действующих скважинах', None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Исполнитель', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, 'по договору №', None, None, None, None, None, 'ГТМ', None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Заказчик', None, None, None, 'ООО Ойл-Сервис ', None, None, None, None, None, None, None, None, None,
             None, None, None, None, 'Цех', None, None, well_data.cdng._value, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Уполномоченный представитель', None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, '№ скважины', None, None, None, None, well_data.well_number._value, None, None, None, None, None,
             None, 'куст', None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, 'Регион', None, None, None, well_data.region, None, None, None, None, None, None, None, 'Месторождение', None,
             None, None, None, None, well_data.well_area._value, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None],
            [None, 'Дата', None, None, date_new_edit, None, None, None, None, None, 'Время',
             None, None, time_new_edit, None, None, None, None, None, None, None, None, None, None, None,
             '(по регламенту за 16 часов)', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None],
            [None, 'Дата', None, None, date_new_edit, None, None, None, None, None, 'Время',
             None, None, time_new_edit, None, None, None, None, None, None, None, None, None, None, None, '(готовности по факту)',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Комплекс и интервал исследования:', None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, '№ Тех. \nкарты', None, None, 'Цель ГИРС', None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, 'интервал ПВР/ГИС, м.', None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 'от', None,
             None, None, 'до', None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Примечание к комплексу исследования', None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, note_to_gis, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Основные сведения по скважине:', None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Категория скважины по ГНВП', None, None, None, None, None, None, None, None, None, 'Рпл:', None,
             well_data.category_pressuar,
             None, None, None, None, None, 'H2S:', None, well_data.category_h2s, None, None, None, None, None, None,
             'Газовый фактор:', None,
             None, None, None, None, None,  well_data.category_gf, None,  None, None, None, None],
            [None, 'Пробуренный забой', None, None, None, None, None, None, well_data.bottomhole_drill._value, None,
             None, None, None, 'м.', None,
             'Искусственный забой', None, None, None, None, None, None, None, well_data.bottomhole_artificial._value,
             None, None, 'м.', None,
             'Текущий забой', None, None, None, None, None, well_data.current_bottom, None,  None, None, 'м.', None, None],
            [None, 'Максимальный угол', None, None, None, None, None, None, well_data.max_angle._value, None, None,
             None, None, None, None,
             'гр.', None, 'на глубине', None, None, None, well_data.max_angle_H._value, None, None, None, None, 'м.',
             None, None, None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Расстояние муфта-ротор', None, None, None, None, None, None, None, None, well_data.stol_rotora._value, None, None, None,
             'м.', None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Кондуктор: глубина спуска, м.', None, None, None, None, None, None, None, None, None,
             conductor, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Техническая колонна: глубина спуска, м.', None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, 'диаметр, мм.', None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Э/колонна: глубина спуска, м.', None, None, None, None, None, None, None, None, None,
             column_data, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Хвостовик: глубина спуска, м.', None, None, None, None, None, None, None, None, None, column_data_add, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             'Окно врезки, м.', None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Высота подъема цемента за колонной, м.', None, None, None, None, None, None, None, None, None, None,
             None, None, None, well_data.level_cement_column._value, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Толщина стенки труб последней колонны, мм', None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Диаметр НКТ нар., мм', None, None, None, None, None, None, None, nkt_edit, None, None,
             'Диаметр НКТ вн., мм', None, None, None, None, None, None, None, nkt_edit_vn, None, None, None, None,
             'Глубина спуска, м.', None, None, None, None, None, None, nkt_shoe_edit, None, None, None, None, None, None, None,
             None],
            [None, 'Низ НКТ оборудован', None, None, None, None, None, None, None, 'воронка', None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Пакер', None, None, paker_type, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, 'Глубина спуска пакера, м.', None, None, None, None, None, None, None, None, paker_depth, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Прочее оборудование', None, None, None, None, None, None, None, '', None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None],
            [None, 'Пусковые муфты', None, None, None, None, None, None, None, 'муфты', None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None],
            [None, 'Расстояние муфта-ротор, м.', None, None, None, None, None, None, None, None, None, well_data.stol_rotora._value, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Скважина заполнена:', None, None, None, None, None, None, None, 'Тип:', None, None, 'тех.вода ',
             None, None, None, None, None, None, None, None, None, None, None, None, 'Уровень, м.', None, None, None,
              None, well_data.static_level._value, None,  None, None, None, None, None, None, None, None, None],
            [None, 'Плотность, г/см3', None, None, None, None, None, fluid, None, None, None, None, 'Вязкость, сек.',
             None, None, None, None, None, None, None, None, None, None, None, 'УЭС, Омм', None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Ожидаемое пластовое давление, атм', None, None, None, None, None, None, None, None, None, None,
             None, None, pressuar, None, None, None, None, None, None, None, None, None, None, None, 'Газовый фактор, м3/т',
             None, None, None, None, None, None, None, 15.3, None, None, None, None, None, None],
            [None, 'Сведения о ранее проведенных ПВР', None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, 'интервал', None, None, None, None, None, 'тип перфоратора', None, None, None, None, None, None,
             'плотность', None, None, None, None, 'дата', None, None, None, None, None, 'индекс пласта', None, None,
             None, None, None, None, 'примечание', None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Другие данные по скважине:', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Наличие электроэнергии', None, None, None, None, None, None, None, None, 'нет', None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Способ эксплуатации', None, None, None, None, None, None, None, '', None, None, None, None,
             None, None, None, None, None, None, None, None, None, 'глубина спуска, м', None, None, None, None, None,
             '', None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Время остановки скважины', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, 'Текущий дебит, Приемистость', None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, ',', None, None, None, None, 'т/сут, м3/сут', None, None, None, None, None,
             'Обводненность', None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Начальный дебит, Приемистость', None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, ',', None, None, None, None, 'т/сут, м3/сут', None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'наличие сужений(нет/да), интервал', None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'наличие уступов(нет/да), интервал', None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None,
             'Описание работ проводимых непосредственно перед ГИРС (кислотные обработки, агрессивные растворы, продолжительность реакции обработки, время промывки после обработки реперфорация и др.):',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, work_edit, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Дополнительные сведения', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Максимально ожидаемое давление на устье скважины', None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, well_data.max_admissible_pressure._value, None,
             None, None, None, None, None, 'атм.', None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Расстояние до скважины', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, 'км.', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Ответственный представитель Заказчика на скважине во время ГИРС', None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, 'телефон', None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, number_brigada, None, None, None, None,
             number_telephone, None,
             None, None, None, None, None, None, None, None],
            [None, 'Заявку подал:', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, 'телефон', None, None, None, None, None, None,
             None, None, None, None, None, None, None, None],
            [None, f'{well_data.user[0]} {well_data.user[1]}', None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'Заявку согласовал:', None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, 'телефон', None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             '№ заявки, присвоеннный в системе ЦДС-Менеджер', None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None],
            [None, '', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],

        ]
        pvr_list = [[None, 'интервал', None, None, None, None, None, 'тип перфоратора', None, None, None, None, None, None,
             'плотность',
             None, None, None, None, 'дата', None, None, None, None, None, 'индекс пласта', None, None, None, None,
             None, None,
             'примечание', None, None, None, None, None, None, None, None, None, None]]


        for plast in well_data.plast_all:
            for interval in well_data.dict_perforation[plast]['интервал']:
                if well_data.dict_perforation[plast]['отключение']:
                    izol = 'Изолирован'
                else:
                    izol = 'рабочий'
                pvr_list.append(
                    [None, f'{interval[0]}-{interval[1]}', None, None, None, None, None, None, None, None, None, None,
                     None, None, None,
                     None, None, None, None, None, None, None, None, None, None, plast, None, None, None, None,
                     None, None,
                     izol, None, None, None, None, None, None,
                     None, None, None, None])
        col = 0
        for pvr in pvr_list:
            value_list[44 + col] = pvr
            col += 1


        return value_list



# app = QApplication(sys.argv)
# login_window = PvrApplication()
# login_window.show()
# sys.exit(app.exec_())