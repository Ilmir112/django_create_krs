import well_data
from openpyxl.styles import Border, Side, Font,  Alignment

def calc_h2s(ws3, h2s_pr, h2s_mg):
    from H2S import well_volume
    # print(well_data.dict_nkt)
    nkt_1 = list(well_data.dict_nkt.keys())[0]
    nkt_1_l = well_data.dict_nkt[nkt_1]

    try:
        nkt_2 = list(well_data.dict_nkt.keys())[1]
        nkt_2_l = well_data.dict_nkt[nkt_2]
    except:
        nkt_2 = 0
        nkt_2_l = 0

    def gno_volume():
        nkt_l = sum(list(well_data.dict_nkt.values()))
        # print(nkt_l)
        if well_data.column_additional is True:

            gno_well = 3.14 * (
                        well_data.column_diametr - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                   well_data.bottom - float(well_data.head_column_additional._value)) / 1000 / 10
        else:
            if nkt_l < well_data.shoe_column._value:
                gno_well = (3.14 * (
                            well_data.column_diametr - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                        well_data.bottom - float(well_data.head_column_additional._value)) / 1000) + (
                                       3.14 * (
                                           well_data.column_diametr - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                                   well_data.shoe_column - nkt_l) / 1000)
            else:
                gno_well = 3.14 * (
                            well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                       well_data.bottom - nkt_1_l) / 10000

        return gno_well

    try:
        nkt_3 = list(well_data.dict_nkt.keys())[2]
    except:
        nkt_3 = 0
    try:
        sucker_rod_l_25 = well_data.dict_sucker_rod['25']
    except:
        sucker_rod_l_25 = 0
    try:
        sucker_rod_l_22 = well_data.dict_sucker_rod['22']
    except:
        sucker_rod_l_22 = 0
    try:
        sucker_rod_l_19 = well_data.dict_sucker_rod['19']
    except:
        sucker_rod_l_19 = 0

    SNPKH = [
        [None, 'Расчет расхода нейтрализатора сероводорода на модификацию раствора глушения*', None, None, None,
         None],
        [None, 'Масса нейтрализатора сероводорода - по результатам расчета*', None, None, None, None],
        [None, 'Объем раствора глушения - по объему скважины от устья до забоя', None, None, None, None],
        [None, None, None, None, None, None],
        [None, '№', 'Параметр', None, 'Результат расчета', None],
        [None, 1, 'Параметры скважины', None, f'{well_data.well_number._value} {well_data.well_area._value}', None],
        [None, 1.1, 'Забой скважины', 'м', round(float(well_data.bottomhole_artificial._value), 1), 'формула'],
        [None, 1.2, 'текущий забой', 'м', round(float(well_data.bottom), 1), 'ввод'],
        [None, 1.3, 'Диаметр ЭК (ступень 1 верхняя)', 'мм', int(well_data.column_diametr._value), 'ввод'],
        [None, '1.3.1.', 'Толщина стенки ЭК (ступень 1 верхняя)', 'мм',
         round(float(well_data.column_wall_thickness._value), 1), 'ввод'],
        [None, '1.3.2.', 'Длина подвески ЭК (ступень 1 верхняя)', 'м', int(well_data.shoe_column._value), 'ввод'],
        [None, '1.3.3.', 'Диаметр ЭК (ступень 2 хвостовик)', 'мм', int(well_data.column_additional_diametr._value),
         'ввод'],
        [None, '1.3.4.', 'Толщина стенки ЭК (ступень 2 хвостовик)', 'м',
         float(well_data.column_additional_wall_thickness._value),
         'ввод'],
        [None, '1.3.5.', 'Длина подвески ЭК (ступень 2 хвостовик)', 'м',
         abs(int(well_data.head_column_additional._value) - int(well_data.shoe_column_additional._value)), 'ввод'],
        [None, '1.3.6.', 'Глубина "головы" (ступень 2 хвостовик)', 'м', int(well_data.head_column_additional._value),
         'ввод',
         ],
        [None, '1.3.7.', 'Глубина "башмака" (ступень 2 хвостовик)', 'м', int(well_data.shoe_column_additional._value),
         'формула'],
        [None, None, None, None, None, None],
        [None, 2, 'Параметры ГНО', None, None, None, None],
        [None, '2.1.', 'Общая глубина подвески НКТ', 'м', '=E22+E25', 'формула'],
        [None, '1.4.1', 'Толщина стенки подвески НКТ (ступень 1 верхняя)', 'мм', 5.5, 'ввод'],
        [None, '1.4.2', 'Внешний диаметр подвески НКТ (ступень 1 верхняя)', 'мм', nkt_1, 'ввод'],
        [None, '1.4.3', 'Длина подвески НКТ (ступень 1 верхняя)', 'м', nkt_1_l, 'ввод'],
        [None, '1.4.4', 'Толщина стенки подвески НКТ (ступень 2 нижняя)', 'мм', 0, 'ввод'],
        [None, '1.4.5', 'Внешний диаметр подвески НКТ (ступень 2 нижняя)', 'мм', nkt_2, 'ввод'],
        [None, '1.4.6', 'Длина подвески НКТ (ступень 2 нижняя)', 'м', nkt_2_l, 'ввод'],
        [None, '2.2', 'Общая глубина подвески штанг', 'м', '=E27+E28+E29', 'формула'],
        [None, '2.2.1.', 'Длина штанг 25 мм', 'м', sucker_rod_l_25, 'ввод'],
        [None, '2.2.2.', 'Длина штанг 22 мм', 'м', sucker_rod_l_22, 'ввод'],
        [None, '2.2.3.', 'Длина штанг 19 мм', 'м', sucker_rod_l_19, 'ввод'],
        [None, 3, 'Расчеты емкости', None, None, None, None],
        [None, 3.1, 'Удельный  внутренний объем ЭК', 'дм3/м', "=ROUND(10*3.14*((E9-E10*2)*0.01)^2/4, 1)", 'формула'],
        [None, 3.2, 'Удельный  внутренний объем хвостовика', 'дм3/м', '=ROUND(10*3.14*((E12-E13*2)*0.01)^2/4,1)',
         'формула'],
        [None, 3.3, 'Объем жидкости под ГНО, в т.ч.:', 'м3',
         '=ROUND(if(if(E14=0,(3.14*(E9-E10*2)^2/4/1000*(E8-E19)/1000),if(E19<E15,(3.14*(E12-E13*2)^2/4/1000*(E8-E19)/1000),((3.14*(E12-E13*2)^2/4/1000*(E8-E15)/1000) +3.14*(E9-E10*2)^2/4/1000*(E15-E19)/1000 )))<0, 0, if(E14=0,(3.14*(E9-E10*2)^2/4/1000*(E8-E19)/1000),if(E19<E15,(3.14*(E12-E13*2)^2/4/1000*(E8-E19)/1000),((3.14*(E12-E13*2)^2/4/1000*(E8-E15)/1000) +3.14*(E9-E10*2)^2/4/1000*(E15-E19)/1000 )))),)  ',
         'формула'],
        [None, '3.3.1.', 'Объем скважины', 'м3',
         '=ROUND(if(E14=0,3.14*(E9-E10*2)^2/4/1000*(E8)/1000,(3.14*(E12-E13*2)^2/4/1000*(E8-E15)/1000)+(3.14*(E9-E10*2)^2/4/1000*(E15)/1000)),1)',
         'формула'],
        [None, '3.3.1.', 'Удельное водоизмещение подвески НКТ (ступень 1 верхняя)', 'дм3/м',
         '=ROUND(10*3.14*((E21*0.01)^2-(E21*0.01-E20*2*0.01)^2)/4,1)',
         'формула'],
        [None, '3.3.2.', 'Удельное водоизмещение подвески НКТ (ступень 2 нижняя)', 'дм3/м',
         '=ROUND(10*3.14*((E24*0.01)^2-(E24*0.01-E23*2*0.01)^2)/4,1)', 'формула', ],
        [None, '3.3.3.', 'Водоизмещение подвески НКТ (объем жидкости притока при СПО)', 'м3/СПО',
         '=ROUND((E22*E35/1000) +(E25*E36/1000),1)', 'формула'],
        [None, '3.3.4.', 'Водоизмещение подвески штанг (объем жидкости притока при СПО)', 'м3/СПО',
         '=ROUND((10*3.14*((25*0.01)^2/4)*E27/1000)+(10*3.14*((22*0.01)^2/4)*E28/1000)+(10*3.14*((19*0.01)^2/4)*E29/1000),1)',
         'формула'],
        [None, None, None, None, None, None],
        [None, 4, 'Параметры добываемой жидкости и газа', None, None, None, None],
        [None, 4.1, 'Газосодержание нефти', 'м3/тонну', well_data.gaz_f_pr[0], 'ввод'],
        [None, 4.2, 'Содержание сероводорода в газе (по данным проекта разработки)', '% (об)', well_data.h2s_pr[0],
         'ввод'],
        [None, 4.3, 'Обводенность продукции', '% (масс.)', well_data.proc_water, 'ввод'],
        [None, 4.4, 'Содержание сероводорода в пластовом флюиде (устьевая проба, вода+нефть)', 'мг/дм3',
         well_data.h2s_mg[0], 'ввод'],
        [None, 4.5, 'Плотность воды', 'г/см3', 1.17, 'ввод'],
        [None, 4.6, 'Плотность нефти', 'г/см3', 0.9, 'ввод'],
        [None, None, None, None, None, None],
        [None, 5, 'Расчет массы сероводорода в жидкости притока (в объеме водоизмещения подвески НКТ и штанг)', None,
         None, None],
        [None, 5.1, 'Масса нефти ', 'т', '=ROUND((E37+E38)*(100-E43)*E46/100,1)', 'формула'],
        [None, 5.2, 'Объем сероводорода, м3', 'м3', '=ROUND(E41*E49*E42/100,1)', 'формула'],
        [None, 5.3, 'Масса сероводорода в нефти (выделяющаяся в газовую фазу при снижении давления)', 'г',
         '=(34*E50*1000/22.14)', 'формула'],
        [None, 5.4, 'Масса сероводорода в жидкости (остаточная растворенная часть)', 'г', '=ROUND((E37+E38)*E44,1)',
         'формула'],
        [None, None, None, None, None, None],
        [None, 6, 'Расчет массы сероводорода в поднасосной жидкости (в объеме скважины под ГНО)', None, None, None],
        [None, 6.1, 'Масса нефти ', 'т', '=ROUND(E33*(100-E43)*E46/100,1)    ', 'формула'],
        [None, 6.2, 'Объем сероводорода, м3', 'м3', '=ROUND(E41*E55*E42/100,1)   ', 'формула'],
        [None, 6.3, 'Масса сероводорода в нефти (доля, выделяющаяся в газовую фазу при снижении давления)', 'г',
         '=ROUND((34*E56*1000/22.14),1)  ', 'формула'],
        [None, 6.4, 'Масса сероводорода в жидкости (остаточная растворенная часть)', 'г', '=ROUND(E33*E44,1)',
         'формула'],
        [None, None, None, None, None, None],
        [None, 7, 'Масса сероводорода общая', 'г', '=ROUND(E51+E52+E57+E58,1)', 'формула'],
        [None, None, None, None, None, None],
        [None, 8, 'Параметр нейтрализатора сероводорода', None, None, None],
        [None, 8.1, 'Емкость реагента по сероводороду (определяется по результатам ЛИ конкретной марки)', 'г/г H2S',
         24, 'ввод'],
        [None, 8.2, 'Плотность товарной формы (марки) реагента (по ТУ)', 'г/см3', 1.065, 'ввод'],
        [None, None, None, None, None, None],
        [None, 9, 'Результат расчета расхода нейтрализатора сероводорода', None, None, None],
        [None, 9.1, 'Расчетная масса реагента', 'кг', '=ROUND(E63*E60/1000,1)', 'формула'],
        [None, 9.2, 'Коэффициент запаса по реагенту (решение ОГ)', 'крат', 1.25, 'ввод'],
        [None, 9.3, 'Масса реагента с запасом', 'кг', '=E67*E68', 'формула', None, None, None],
        [None, None, None, None, None, None],
        [None, 10, 'Удельный расход нейтрализатора сероводорода в растворе глушения', None, None, None],
        [None, 10.1, 'Удельный массовый расход нейтрализатора сероводорода ', 'кг/м3', '=E69/E34', 'формула'],
        [None, 10.2, 'Удельный объемный расход нейтрализатора сероводорода ', 'м3/м3', '=E72/E64', 'формула'],
        [
            'Примечание: * - расчет приблизительный, поскольку нет информации по содержанию H2S в каждой фазе (вода, нефть, газ) в одинаковых условиях',
            None, None, None, None, None, ],
        [None, None, None, None, None, None],
        [None, 'Ячейки для заполнения данными параметров скважины', None, None, None, None],
        [None, 'По данным конструкции скважины', None, None, None, None],
        [None, None, None, None, None, None],
        [None, None, None, None, None, None],
        [None, 'По данным лабораторных исследований марки нейтрализатора сероводорода', None, None, None, None,
         None],
    ]

    max_row_H2S = len(SNPKH)
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))
    ws3.column_dimensions['B'].width = 15
    ws3.column_dimensions['C'].width = 80
    ws3.column_dimensions['D'].width = 25
    ws3.column_dimensions['f'].width = 30
    ws3.column_dimensions['e'].width = 25
    for row in range(1, max_row_H2S):
        for col in range(1, 7):
            ws3.cell(row=row, column=col).value = SNPKH[row - 1][col - 1]

            if 5 <= row <= 73 and 1 < col < 6:
                ws3.cell(row=row, column=col).border = thin_border
                ws3.cell(row=row, column=col).font = Font(name='Arial', size=13, bold=True)
                ws3.cell(row=row, column=col).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                    vertical='center')

    ws3.hide = True
    ws3.page_setup.fitToPage = True
    ws3.page_setup.fitToHeight = False

    ws3.page_setup.fitToWidth = True
    ws3.print_area = 'A1:F77'


def well_volume(self):
    # print(well_data.column_additional)
    if well_data.column_additional is False:

        volume_well = 3.14 * (
                    well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000000 * (
                          well_data.bottomhole_artificial._value)
        return volume_well
    else:

        volume_well = (3.14 * (
                    well_data.column_additional_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                               well_data.bottomhole_artificial._value - float(
                           well_data.head_column_additional._value)) / 1000) + (
                              3.14 * (
                                  well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                  float(well_data.head_column_additional._value)) / 1000)
        return volume_well


def calv_h2s(self, cat_H2S, h2s_mg, h2s_pr):
    if '2' == str(cat_H2S) or '1' in str(cat_H2S):
        nkt_l = sum(list(well_data.dict_nkt.values()))

        udel_vnutr_v = 10 * 3.14 * (
                    (well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) * 0.01) ** 2 / 4
        if well_data.column_additional is True:
            udel_vn__khv = 10 * 3.14 * ((
                                                    well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) * 0.01) ** 2 / 4
        # print(f'ff{udel_vn__khv}')
        # print(f' НКТ{well_data.column_diametr._value}2 {nkt_l, well_data.head_column_additional._value}88{well_data.column_diametr, well_data.column_wall_thickness._value}0{well_data.head_column_additional._value, well_data.bottomhole_artificial._value}')
        if well_data.column_additional is False:

            v_pod_gno = 3.14 * (int(well_data.column_diametr._value) - int(
                well_data.column_wall_thickness._value) * 2) ** 2 / 4 / 1000 * (
                                    well_data.bottomhole_artificial._value - int(nkt_l)) / 1000
        elif nkt_l > float(well_data.head_column_additional._value):
            v_pod_gno = 3.14 * (
                        well_data.column_diametr._value - well_data.column_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                    float(well_data.head_column_additional._value) - nkt_l) / 1000 + 3.14 * (
                                    well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                    well_data.bottomhole_artificial._value - float(
                                well_data.head_column_additional._value)) / 1000
        elif nkt_l < float(well_data.head_column_additional._value):
            v_pod_gno = 3.14 * (
                        well_data.column_additional_diametr._value - well_data.column_additional_wall_thickness._value * 2) ** 2 / 4 / 1000 * (
                                    well_data.bottomhole_artificial._value - nkt_l) / 1000
        print(f'под ГНО{well_data.dict_nkt.keys()}')
        volume_well = well_volume(self)
        if '73' in list(well_data.dict_nkt.keys())[0]:
            nkt_1 = 73
        elif '60' in list(well_data.dict_nkt.keys())[0]:
            nkt_1 = 60
        elif '48' in list(well_data.dict_nkt.keys())[0]:
            nkt_1 = 48
        elif '89' in list(well_data.dict_nkt.keys())[0]:
            nkt_1 = 89
            print(89)

        try:
            # print(list(well_data.dict_nkt.keys()))
            nkt_2 = int(list(well_data.dict_nkt.keys())[1])
            nkt_2_l = well_data.dict_nkt[nkt_2]
        except:
            nkt_2 = 0
            nkt_2_l = 0

        udel_vodoiz_nkt = 10 * 3.14 * ((nkt_1 * 0.01) ** 2 - (nkt_1 * 0.01 - 5.5 * 2 * 0.01) ** 2) / 4
        # print(f' удел {udel_vodoiz_nkt}')
        try:
            # print(f'НКТ-{nkt_2}')
            if nkt_2 != 0:
                udel_vodoiz_nkt_2 = 10 * 3.14 * ((nkt_2 * 0.01) ** 2 + (nkt_2 * 0.01 - 5 * 2 * 0.01) ** 2) / 4
                # print(f'dnjhfzНКТ {udel_vodoiz_nkt_2}')
        except:
            udel_vodoiz_nkt = udel_vodoiz_nkt
            # print(f'dnjhfzНКТ {udel_vodoiz_nkt}')
        print(f'jggk {well_data.dict_nkt}')
        nkt_1_l = well_data.dict_nkt[list(well_data.dict_nkt.keys())[0]]
        vodoiz_nkt = nkt_1_l * udel_vodoiz_nkt / 1000
        try:
            vodoiz_nkt += nkt_2_l * udel_vodoiz_nkt_2 / 1000
        except:
            vodoiz_nkt = vodoiz_nkt
        try:
            sucker_rod_l_25 = well_data.dict_sucker_rod['25']
        except:
            sucker_rod_l_25 = 0
        try:
            sucker_rod_l_22 = well_data.dict_sucker_rod['22']
        except:
            sucker_rod_l_22 = 0
        try:
            sucker_rod_l_19 = well_data.dict_sucker_rod["19"]
        except:
            sucker_rod_l_19 = 0

        vodoiz_sucker = (10 * 3.14 * ((25 * 0.01) ** 2 / 4) * sucker_rod_l_25 / 1000) + (
                    10 * 3.14 * ((25 * 0.01) ** 2 / 4) * sucker_rod_l_22 / 1000) + (
                                    10 * 3.14 * ((25 * 0.01) ** 2 / 4) * sucker_rod_l_19 / 1000)

        oil_mass = float(v_pod_gno * (100 - well_data.proc_water) * 0.9 / 100)
        # print(f'oil {oil_mass}-{type(oil_mass)} , {well_data.gaz_f_pr[0]}-{type(well_data.gaz_f_pr[0])}')
        volume_h2s = well_data.gaz_f_pr[0] * oil_mass * (float(h2s_pr)) / 100

        h2s_mass_in_oil = (34 * volume_h2s * 1000 / 22.14)
        # print(type(vodoiz_sucker), type(vodoiz_nkt), h2s_mg, float(h2s_mg))
        h2s_mass_in_water = float(vodoiz_sucker + vodoiz_nkt) * h2s_mg
        # print(f'h2a{h2s_mass_in_water}')
        mass_oil_pog_gno = (vodoiz_sucker + vodoiz_nkt) * (100 - well_data.proc_water) * 0.9 / 100
        h2s_volume_pod_gno = mass_oil_pog_gno * well_data.gaz_f_pr[0] * h2s_pr / 100
        mass_h2s_gas = 34 * h2s_volume_pod_gno / 22.14
        mass_h2s_water = v_pod_gno * h2s_mg
        # print(f'mass{mass_h2s_water}')
        mass_h2s_all = h2s_mass_in_water + h2s_mass_in_oil + mass_h2s_gas + mass_h2s_water
        # print(f'mass_h2 {mass_h2s_all}')

        emk_reag = 24
        plotn_reag = 1.065
        raschet_mass = mass_h2s_all * emk_reag / 1000
        # print(f'ras{raschet_mass}')

        koeff_zapas = 1.25
        mass_reag_s_zapas = raschet_mass * koeff_zapas
        # print(f'mass{mass_reag_s_zapas}')
        udel_mas_raskhod = mass_reag_s_zapas / volume_well
        # print(udel_mas_raskhod)

        if udel_mas_raskhod <= 0.01:
            udel_mas_raskhod = 0.01
        return round(udel_mas_raskhod, 3)
    else:
        return 0
