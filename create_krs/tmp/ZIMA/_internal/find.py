from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtWidgets import QInputDialog, QMainWindow, QMessageBox
from openpyxl.reader.excel import load_workbook
from openpyxl.utils import column_index_from_string
from openpyxl.workbook import Workbook

import well_data
from category_correct import CategoryWindow
from data_correct import DataWindow
from main import ExcelWorker, MyWindow
from perforation_correct import PerforationCorrect

from work_py.leakage_column import LeakageWindow
from well_data import ProtectedIsDigit, ProtectedIsNonNone
from block_name import region
from work_py.advanted_file import definition_plast_work
from openpyxl.utils.cell import get_column_letter


class FindIndexPZ(QMainWindow):
    wb_pvr = Workbook()

    def __init__(self, ws):
        super().__init__()
        self.read_pz(ws)
        self.data_window = None
        self.perforation_correct_window2 = None

    def read_pz(self, ws):

        cat_well_min = []
        for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
            ws.row_dimensions[row_ind].hidden = False

            if 'Категория скважины' in row:
                cat_well_min.append(row_ind + 1)
                well_data.cat_well_min = ProtectedIsDigit(min(cat_well_min))  # индекс начала категории

            elif any(['план-заказ' in str(col).lower() or 'план работ' in str(col).lower() for col in row]) \
                    and row_ind < 50:
                well_data.cat_well_max = ProtectedIsDigit(row_ind)
                well_data.data_well_min = ProtectedIsDigit(row_ind + 1)


            elif any(['Ожидаемые показатели после' in str(col) for col in row]):
                well_data.data_x_min = ProtectedIsDigit(row_ind)
                # print(f' индекс Ожидаемые показатели {well_data.data_x_min}')
            elif '11. Эксплуатационные горизонты и интервалы перфорации:' in row:
                well_data.data_pvr_min = ProtectedIsDigit(row_ind)
            elif 'Оборудование скважины ' in row:
                well_data.data_fond_min = ProtectedIsDigit(row_ind)


            elif any(['IX. Мероприятия по предотвращению' in str(col) for col in row]) or \
                    any(['IX. Мероприятия по предотвращению аварий, инцидентов и осложнений::' in str(col) for col in
                         row]):
                well_data.data_well_max = ProtectedIsDigit(row_ind)

            elif 'НКТ' == str(row[1]).upper():

                well_data.pipes_ind = ProtectedIsDigit(row_ind + 1)



            elif 'ШТАНГИ' == str(row[1]).upper():
                well_data.sucker_rod_ind = ProtectedIsDigit(row_ind + 1)


            elif 'ХI Планируемый объём работ:' in row or 'ХI. Планируемый объём работ:' \
                    in row or 'ХIII Планируемый объём работ:' in row \
                    or 'ХI Планируемый объём работ:' in row or 'Порядок работы' in row:
                    well_data.data_x_max = ProtectedIsDigit(row_ind)


            elif any(['II. История эксплуатации скважины' in str(col) for col in row]):
                well_data.data_pvr_max = ProtectedIsDigit(row_ind)

            elif 'III. Состояние скважины к началу ремонта ' in row:
                well_data.condition_of_wells = ProtectedIsDigit(row_ind)
            elif 'Герметизация , разгерметизация  устья  скважины' in row:
                well_data.plan_correct_index = ProtectedIsDigit(row_ind)

        if well_data.cat_well_max._value == 0:
            mes = QMessageBox.warning(self, 'Ошибка', 'Не корректный файл excel, либо отсутствует строка с '
                                                      'текстом ПЛАН-ЗАКАЗ или ПЛАН-РАБОТ')
            MyWindow.pause_app()
            return

        if well_data.cat_well_min._value == 0:
            well_data.cat_well_min = ProtectedIsDigit(QInputDialog.getInt(
                self, 'индекс начала копирования', 'Программа не смогла определить строку начала копирования',
                0, 0, 800)[0])
            while well_data.cat_well_min._value > 30:
                mes = QMessageBox.warning(self, 'Ошибка', 'Не корректное значение')
                well_data.cat_well_min = ProtectedIsDigit(QInputDialog.getInt(
                    self, 'индекс начала копирования', 'Программа не смогла определить строку начала копирования',
                    0, 0, 800)[0])
        if well_data.data_well_max._value == 0:
            well_data.data_well_max = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс окончания копирования',
                                    'Программа не смогла определить строку с IX. Мероприятия по предотвращению аварий ',
                                    0, 0, 800)[0])
            while well_data.data_well_max._value < 150:
                mes = QMessageBox.warning(self, 'Ошибка', 'Не корректное значение')
                well_data.data_well_max = ProtectedIsDigit(
                    QInputDialog.getInt(self, 'индекс окончания копирования',
                                        'Программа не смогла определить строку IX. Мероприятия по предотвращению аварий',
                                        0, 0, 800)[0])
        if well_data.data_well_min._value == 0:
            well_data.data_well_min = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс начала строки после план заказ',
                                    'Программа не смогла найти начала строки после план заказ',
                                    0, 0, 800)[0])
            while well_data.data_well_min._value < well_data.cat_well_min._value or \
                    well_data.data_well_min._value > well_data.data_well_max._value:
                mes = QMessageBox.warning(self, 'Ошибка', 'Не корректное значение')
                well_data.data_well_min = ProtectedIsDigit(
                    QInputDialog.getInt(self, 'индекс начала строки после план заказ',
                                        'Программа не смогла найти начала строки после план заказ',
                                        0, 0, 800)[0])

        if well_data.cat_well_max._value == 0:
            well_data.cat_well_max = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс начала копирования',
                                    'Программа не смогла определить строку начала копирования',
                                    0, 0, 800)[0])
        if well_data.sucker_rod_none:
            if well_data.sucker_rod_ind._value == 0:
                sucker_mes = QMessageBox.question(self, 'ШТАНГИ', 'Программа определелила, что в скважине '
                                                                  'отсутствуют штанги, корректно ли это?')
                if sucker_mes == QMessageBox.StandardButton.Yes:
                    well_data.sucker_rod_ind = ProtectedIsDigit(0)
                else:
                    well_data.sucker_rod_ind = ProtectedIsDigit(
                        QInputDialog.getInt(self, 'индекс начала строки со штангами',
                                            'Программа не смогла найти строку со штангами',
                                            0, 0, 800)[0])

        if well_data.data_x_max._value == 0:
            well_data.data_x_max = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс окончания копирования ожидаемых показателей',
                                    'Программа не смогла определить строку окончания копирования'
                                    ' ожидаемых показателей',
                                    0, 0, 800)[0])


        if well_data.data_x_min._value == 0:
            well_data.data_x_min = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс начала копирования ожидаемых показателей',
                                    'Программа не смогла определить строку начала копирования'
                                    ' ожидаемых показателей',
                                    0, 0, 800)[0])

        if well_data.data_pvr_max._value == 0:
            well_data.data_pvr_max = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс историю',
                                    'Программа не смогла найти "II. История эксплуатации скважины"',
                                    0, 0, 800)[0])

        if well_data.pipes_ind._value == 0:
            well_data.pipes_ind = ProtectedIsDigit(QInputDialog.getInt(self, 'индекс начала строки с НКТ',
                                                                       'Программа не смогла найти строку с НКТ',
                                                                       0, 0, 800)[0])
        if well_data.data_pvr_min._value == 0:
            well_data.data_pvr_min = ProtectedIsDigit(QInputDialog.getInt(self, 'индекс начала начала ПВР',
                                                                          'Программа не смогла найти индекс начала ПВР',
                                                                          0, 0, 800)[0])
        if well_data.data_fond_min._value == 0:
            well_data.data_fond_min = ProtectedIsDigit(
                QInputDialog.getInt(self, 'индекс начала строки с таблицей фондовыго оборудования',
                                    'Программа не смогла найти строку с таблицей фондового оборудования',
                                    0, 0, 800)[0])
        if well_data.condition_of_wells._value == 0:
            well_data.condition_of_wells = ProtectedIsDigit(QInputDialog.getInt(
                self, 'индекс копирования',
                'Программа не смогла определить строку n\ III. '
                'Состояние скважины к началу ремонта ',
                0, 0, 800)[0])
            while well_data.condition_of_wells._value < well_data.cat_well_min._value or \
                    well_data.condition_of_wells._value > well_data.data_well_max._value:
                mes = QMessageBox.warning(self, 'Ошибка', 'Не корректное значение')
                well_data.condition_of_wells = ProtectedIsDigit(
                    QInputDialog.getInt(self, 'индекс окончания копирования',
                                        'Программа не смогла определить строку окончания копирования',
                                        0, 0, 800)[0])
        # print(well_data.cat_well_min._value, well_data.cat_well_max._value, well_data.data_well_min._value,
        #       well_data.data_pvr_max._value, well_data.condition_of_wells._value, well_data.data_well_max._value,
        #       well_data.pipes_ind._value, well_data.sucker_rod_ind._value

    # )
    def check_str_None(self, string):
        from main import MyWindow

        if MyWindow.check_str_isdigit(self, str(string)) is True:
            if str(round(float(str(string).replace(',', '.')), 1))[-1] == "0":
                return int(float(str(string).replace(',', '.')))
            else:
                return round(float(str(string).replace(',', '.')), 4)
        elif str(string).replace(' ', '') == '-' or 'отсут' in str(string) or \
                str(string).strip() == '' or string is None:
            return '0'
        elif '(мм)' in string and '(м)' in string:
            return string
        elif len(str(string).split('/')) == 2:
            lst = []
            for i in str(string).split('/'):
                b = ''
                for j in i:
                    if j in '0123456789.x':
                        b = str(b) + j
                    elif j == ',':
                        b = str(b) + '.'
                lst.append(float(b))
            return lst
        elif len(str(string).split('-')) == 2:
            lst = []
            for i in str(string).split('-'):
                # print(i)
                lst.append(float(i.replace(',', '.').strip()))
            return lst

        else:
            b = 0
            for i in str(string):
                i.replace(',', '.')
                if i in '0123456789,.x':
                    b = str(b) + i
            return float(b)

    def definition_is_None(self, data, row, col, step, m=12):
        try:

            data = data._value
            while data is None or step == m:
                data = self.ws.cell(row=row, column=col + step).value

                step += 1

            return ProtectedIsNonNone(data)
        except:

            while data is None:
                data = self.ws.cell(row=row, column=col + step).value

                if step == m:
                    break
                step += 1

            return data


class WellNkt(FindIndexPZ):

    def __init__(self, ws):
        super().__init__(ws)
        # self.read_well(ws, well_data.pipes_ind._value, well_data.condition_of_wells._value)

    def read_well(self, ws, begin_index, cancel_index):

        a_plan = 0
        well_data.nkt_mistake = False
        for row in range(begin_index, cancel_index):  # словарь  количества НКТ и метраж

            if ws.cell(row=row, column=3).value == 'План' or str(
                    ws.cell(row=row, column=3).value).lower() == 'после ремонта':
                a_plan = row
        if a_plan == 0:
            a_plan = QInputDialog.getDouble(self, 'Индекс планового НКТ',
                                            'Программа не могла определить начала строку с ПЗ НКТ - план')[0]
        # print(f'индекс a_plan {a_plan}')
        for row in range(begin_index, cancel_index + 1):
            # print(str(ws.cell(row=row, column=4).value))
            key = str(ws.cell(row=row, column=4).value)
            if key != str(None) and key != '-' and "Диам" not in key:
                value = ws.cell(row=row, column=7).value
                if value:
                    if not key is None and row < a_plan:
                        well_data.dict_nkt[key] = well_data.dict_nkt.get(
                            key, 0) + round(FindIndexPZ.check_str_None(self, value), 1)
                    elif not key is None and row >= a_plan:
                        well_data.dict_nkt_po[key] = well_data.dict_nkt_po.get(
                            key, 0) + round(FindIndexPZ.check_str_None(self, value), 1)
                # print(f'индекс a_plan {well_data.dict_nkt}')
            # well_data.shoe_nkt = float(sum(well_data.dict_nkt.values()))
        # except:
        #     well_data.nkt_mistake = True
        #     mes = QMessageBox.warning(self, 'Ошибка', 'Программа не смогла определить диаметры и длину НКТ')


class WellSucker_rod(FindIndexPZ):

    def __init__(self, ws):

        super().__init__(ws)

        # self.read_well(ws, well_data.sucker_rod_ind._value, well_data.pipes_ind._value)

    def read_well(self, ws, begin_index, cancel_index):

        # try:
        well_data.sucker_mistake = False
        b_plan = 0
        if well_data.sucker_rod_ind._value != 0:
            for row in range(begin_index, cancel_index):  # словарь  количества штанг и метраж
                if ws.cell(row=row, column=3).value == 'План' or str(
                        ws.cell(row=row, column=3).value).lower() == 'после ремонта' \
                        or str(
                    ws.cell(row=row, column=3).value).lower() == 'до ремонта':
                    b_plan = row

            if b_plan == 0 and well_data.sucker_rod_none is True:
                sucker_rod_question = QMessageBox.question(self,
                                                           'отсутствие штанг',
                                                           'Программа определило что штанг в '
                                                           'скважине нет, корректно?')
                if sucker_rod_question == QMessageBox.StandardButton.Yes:
                    well_data.sucker_rod_none = False
                else:
                    well_data.sucker_rod_none = True

                if well_data.sucker_rod_none == True:
                    b_plan = QInputDialog.getDouble(self, 'Индекс планового НКТ',
                                                    'Программа не могла определить начала строку с ПЗ штанги - план')
            # print(f'б {b_plan}')

            for row in range(begin_index, cancel_index - 1):

                key = str(ws.cell(row=row, column=4).value).replace(' ', '')
                value = ws.cell(row=row, column=7).value
                if key != str(None) and key != '-' and key != '' and 'отсут' not in str(key).lower():
                    # print(key, value)
                    if key != None and row < b_plan:
                        well_data.dict_sucker_rod[key] = well_data.dict_sucker_rod.get(key, 0) + int(
                            float(str(value).replace(',', '.'))) + 1
                    if key != None and row >= b_plan:
                        well_data.dict_sucker_rod_po[key] = well_data.dict_sucker_rod_po.get(key, 0) + int(
                            float(str(value).replace(',', '.')))
        # except:
        #     well_data.sucker_mistake = True
        #     mes = QMessageBox.warning(self, 'Ошибка', 'Программа не смогла определить диаметры и длину штанг')


class WellFond_data(FindIndexPZ):

    def __init__(self, ws):

        super().__init__(ws)
        # self.read_well(ws, well_data.data_fond_min._value, well_data.condition_of_wells._value)

    def read_well(self, ws, begin_index, cancel_index):

        well_data.old_index = 1
        for row_index, row in enumerate(ws.iter_rows(min_row=begin_index, max_row=cancel_index)):
            row_index += begin_index
            for col, cell in enumerate(row):
                value = cell.value
                if value:
                    if 'карта спуска' in str(value).lower():
                        col_plan = col
                    if 'до ремонта' in str(value).lower() and row_index < 6 + begin_index:
                        col_do = col
                    if 'колонная головка' in str(value) and 'типоразмер' in str(row[col + 2].value):
                        well_data.column_head_m = row[col_do].value
                    if 'Арматура устьевая' in str(value) and 'типоразмер' in str(row[col + 2].value):
                        well_data.wellhead_fittings = row[col_do].value

                    if 'Пакер' in str(value) and 'типоразмер' in str(row[col + 2].value):
                        if '/' in str(row[col_do].value):
                            well_data.paker_do["do"] = str(row[col_do].value).split('/')[0]
                            well_data.paker2_do["do"] = str(row[col_do].value).split('/')[1]
                        else:
                            well_data.paker_do["do"] = row[col_do].value

                        if '/' in str(row[col_plan].value):
                            well_data.paker_do["posle"] = str(row[col_plan].value).split('/')[0]
                            well_data.paker2_do["posle"] = str(row[col_plan].value).split('/')[1]
                        else:
                            well_data.paker_do["posle"] = row[col_plan].value

                    elif value == 'Насос' and row[col + 2].value == 'типоразмер':
                        if row[col_do].value:
                            if ('НВ' in str(row[col_do].value).upper() or 'ШГН' in str(row[col_do].value).upper() \
                                or 'НН' in str(row[col_do].value).upper()) or 'RHAM' in str(row[col_do].value).upper():
                                well_data.dict_pump_SHGN["do"] = row[col_do].value
                                if '/' in str(row[col_do].value):
                                    well_data.dict_pump_SHGN["do"] = [ecn for ecn in row[col_do].value.split('/')
                                                                      if 'НВ' in ecn or 'НН' in ecn or
                                                                      'ШГН' in ecn or 'RHAM' in ecn][0]

                            if ('ЭЦН' in str(row[col_do].value).upper() or 'ВНН' in str(row[col_do].value).upper()):
                                well_data.dict_pump_ECN["do"] = row[col_do].value
                                if '/' in str(row[col_do].value):
                                    well_data.dict_pump_ECN["do"] = [ecn for ecn in row[col_do].value.split('/')
                                                                     if 'ЭЦН' in ecn or 'ВНН' in ecn][0]
                                # print(well_data.dict_pump_ECN["do"])
                        if row[col_plan].value:
                            if ('НВ' in str(row[col_plan].value).upper() or 'ШГН' in str(
                                    row[col_plan].value).upper() \
                                or 'НН' in str(row[col_plan].value).upper()) \
                                    or 'RHAM' in str(row[col_plan].value).upper():
                                well_data.dict_pump_SHGN["posle"] = row[col_plan].value
                                if '/' in str(row[col_plan].value):
                                    well_data.dict_pump_SHGN["posle"] = [ecn for ecn in row[col_plan].value.split('/')
                                                                         if 'НВ' in ecn or 'НН' in ecn or
                                                                         'ШГН' in ecn or 'RHAM' in ecn][0]

                            if ('ЭЦН' in str(row[col_plan].value).upper() or 'ВНН' in str(
                                    row[col_plan].value).upper()):
                                well_data.dict_pump_ECN["posle"] = row[col_plan].value
                                if '/' in str(row[col_plan].value):
                                    well_data.dict_pump_ECN["posle"] = [ecn for ecn in row[col_plan].value.split('/')
                                                                        if 'ЭЦН' in ecn or 'ВНН' in ecn][0]

                        if well_data.dict_pump_ECN["do"] != 0:
                            well_data.dict_pump_ECN_h["do"] = FindIndexPZ.check_str_None(self,
                                                                                         ws.cell(row=row_index + 4,
                                                                                                 column=col_do + 1).value)
                            if '/' in str(ws.cell(row=row_index + 4, column=col_do + 1).value):
                                well_data.dict_pump_ECN_h["do"] = max(FindIndexPZ.check_str_None(self, ws.cell(
                                    row=row_index + 4, column=col_do + 1).value))
                        if well_data.dict_pump_SHGN["do"] != 0:

                            well_data.dict_pump_SHGN_h["do"] = FindIndexPZ.check_str_None(self,
                                                                                          ws.cell(row=row_index + 4,
                                                                                                  column=col_do + 1).value)
                            if '/' in str(ws.cell(row=row_index + 4, column=col_do + 1).value):
                                well_data.dict_pump_SHGN_h["do"] = min(FindIndexPZ.check_str_None(self, ws.cell(
                                    row=row_index + 4, column=col_do + 1).value))
                        if well_data.dict_pump_ECN["posle"] != 0:
                            well_data.dict_pump_ECN_h["posle"] = FindIndexPZ.check_str_None(self,
                                                                                            ws.cell(row=row_index + 4,
                                                                                                    column=col_plan + 1).value)
                            if '/' in str(ws.cell(row=row_index + 4, column=col_plan + 1).value):
                                well_data.dict_pump_ECN_h["posle"] = max(
                                    FindIndexPZ.check_str_None(self, ws.cell(row=row_index + 4,
                                                                             column=col_plan + 1).value))
                        if well_data.dict_pump_SHGN["posle"] != 0:
                            well_data.dict_pump_SHGN_h["posle"] = FindIndexPZ.check_str_None(self,
                                                                                             ws.cell(row=row_index + 4,
                                                                                                     column=col_plan + 1).value)
                            if '/' in str(ws.cell(row=row_index + 4, column=col_plan + 1).value):
                                well_data.dict_pump_SHGN_h["posle"] = min(
                                    FindIndexPZ.check_str_None(self, ws.cell(row=row_index + 4,
                                                                             column=col_plan + 1).value))

                    elif value == 'Н посадки, м':
                        try:
                            if well_data.paker_do["do"] != 0:
                                well_data.depth_fond_paker_do["do"] = \
                                FindIndexPZ.check_str_None(self, row[col_do].value)[0]
                                well_data.depth_fond_paker2_do["do"] = \
                                FindIndexPZ.check_str_None(self, row[col_do].value)[1]
                        except:
                            if well_data.paker_do["do"] != 0:
                                well_data.depth_fond_paker_do["do"] = row[col_do].value
                        try:
                            if well_data.paker_do["posle"] != 0:
                                well_data.depth_fond_paker_do["posle"] = \
                                FindIndexPZ.check_str_None(self, row[col_plan].value)[0]
                                well_data.depth_fond_paker2_do["posle"] = \
                                FindIndexPZ.check_str_None(self, row[col_plan].value)[1]
                        except:
                            if well_data.paker_do["posle"] != 0:
                                well_data.depth_fond_paker_do["posle"] = row[col_plan].value
        if well_data.wellhead_fittings in [None, '']:
            well_data.wellhead_fitting, _ = QInputDialog.getText(self, 'Ошибка',
                                                                 'ПРограмма не могла найти устьевую арматуру, '
                                                                 'Введите вид фонтанной арматуры')
            print(f'арматура {well_data.wellhead_fitting}')


class WellHistory_data(FindIndexPZ):

    def __init__(self, ws):

        super().__init__(ws)
        self.leakage_window = None
        self.ws = ws
        # self.read_well(self.ws, well_data.data_pvr_max._value, well_data.data_fond_min._value)

    def read_well(self, ws, begin_index, cancel_index):

        well_data.max_expected_pressure = ProtectedIsNonNone('не корректно')
        well_data.max_admissible_pressure = ProtectedIsNonNone('не корректно')

        # print(begin_index, cancel_index)
        for row_index, row in enumerate(ws.iter_rows(min_row=begin_index, max_row=cancel_index)):
            for col, cell in enumerate(row):
                value = cell.value
                if value:

                    if 'нэк' in str(value).lower() or 'негерм' in str(
                            value).lower() or 'нарушение э' in str(
                        value).lower() or \
                            'нарушение г' in str(value).lower():
                        well_data.leakiness_Count += 1
                        well_data.leakiness = True
                    if ('авар' in str(
                            value).lower() or 'расхаж' in str(
                        value).lower() or 'лар' in str(value).lower()) \
                            and 'акт о расследовании аварии прилагается' not in str(value).lower():
                        well_data.emergency_well = True
                        well_data.emergency_count += 1
                    if 'сужен' in str(value).lower() or 'не проход' in str(value).lower():
                        well_data.problemWithEk = True

                    if 'Начало бурения' in str(value):
                        well_data.date_drilling_run = row[col + 2].value

                    elif 'Конец бурения' == value:
                        well_data.date_drilling_cancel = row[col + 2].value

                        well_data.date_drilling_cancel = FindIndexPZ.definition_is_None(self,
                                                                                        well_data.date_drilling_cancel,
                                                                                        row_index + begin_index,
                                                                                        col + 1, 1)

                    elif 'Максимально ожидаемое давление на устье' == value:
                        well_data.max_expected_pressure = ProtectedIsDigit(row[col + 1].value)
                        well_data.max_expected_pressure = FindIndexPZ.definition_is_None(self,
                                                                                         well_data.max_expected_pressure,
                                                                                         row_index + begin_index,
                                                                                         col + 1, 1)
                    elif 'Первоначальное давление опрессовки э/колонны' == value:
                        well_data.first_pressure = ProtectedIsDigit(row[col + 3].value)
                        # well_data.first_pressure = FindIndexPZ.definition_is_None(self,
                        #                                                                  well_data.max_expected_pressure,
                        #                                                                  row_index + begin_index,
                        #                                                                  col + 1, 1)

                    elif 'Максимально допустимое давление'.lower() in str(value).lower():
                        well_data.max_admissible_pressure = ProtectedIsDigit(row[col + 1].value)
                        well_data.max_admissible_pressure = FindIndexPZ.definition_is_None(self,
                                                                                           well_data.max_admissible_pressure,
                                                                                           row_index + begin_index,
                                                                                           col + 1, 1)


class WellCondition(FindIndexPZ):
    leakage_window = None

    def __init__(self, ws):

        super().__init__(ws)

        self.ws = ws

        # self.read_well(ws, well_data.condition_of_wells._value, well_data.data_well_max._value)

    def read_well(self, ws, begin_index, cancel_index):
        from main import MyWindow

        well_data.static_level = ProtectedIsNonNone('не корректно')
        well_data.dinamic_level = ProtectedIsNonNone('не корректно')

        well_data.dinamic_level = ProtectedIsDigit(0)
        for row_index, row in enumerate(ws.iter_rows(min_row=begin_index, max_row=cancel_index)):
            row_index += begin_index
            for col, cell in enumerate(row):
                value = cell.value

                if 'нэк' in str(
                        value).lower() or 'негерм' in str(value).lower() or 'нарушение э' in str(
                    value).lower() or 'нарушение г' in str(value).lower():
                    well_data.leakiness_Count += 1
                    well_data.leakiness = True
                if ('авар' in str(value).lower() or 'расхаж' in str(value).lower() or 'лар' in str(value)) \
                        and 'акт о расследовании аварии прилагается' not in str(value):
                    well_data.emergency_well = True
                    well_data.emergency_count += 1
                if value:
                    if "Hст " in str(value):
                        if '/' in str(row[col + 1].value):
                            well_data.static_level = ProtectedIsDigit(row[col + 1].value.split('/')[0])
                        else:
                            well_data.static_level = ProtectedIsDigit(row[col + 1].value)
                    if "грп" in str(value).lower():
                        well_data.grp_plan = True

                    if "Ндин " in str(value):
                        well_data.dinamic_level = ProtectedIsDigit(row[col + 1].value)
                    if "% воды " in str(value):
                        well_data.proc_water = str(row[col + 1].value).strip().replace('%', '')
                        well_data.proc_water = FindIndexPZ.definition_is_None(self, well_data.proc_water, row_index,
                                                                              col + 1, 1)
                    if 'Vжг' in str(value):
                        try:
                            well_volume_in_PZ = str(row[col + 1].value).replace(',', '.')
                            # print(f'строка {well_volume_in_PZ}')
                            # well_volume_in_PZ = FindIndexPZ.definition_is_None(self,well_volume_in_PZ, row_index, col + 1, 1)
                            well_data.well_volume_in_PZ.append(round(float(well_volume_in_PZ), 1))
                        except:
                            well_volume_in_PZ, _ = QInputDialog.getDouble(self, 'Объем глушения',
                                                                          'ВВедите объем глушения согласно ПЗ', 50, 1,
                                                                          70)
                            well_data.well_volume_in_PZ.append(well_volume_in_PZ)

        if well_data.grp_plan:
            grp_plan_quest = QMessageBox.question(self, 'Подготовка к ГРП', 'Программа определела что в скважине'
                                                                            f'планируется ГРП, верно ли?')
            if grp_plan_quest == QMessageBox.StandardButton.Yes:
                well_data.grp_plan = True
            else:
                well_data.grp_plan = False

        if well_data.leakiness is True:
            leakiness_quest = QMessageBox.question(self, 'нарушение колонны',
                                                   'Программа определела что в скважине'
                                                   f'есть нарушение - {well_data.leakiness_Count}, верно ли?')
            if leakiness_quest == QMessageBox.StandardButton.Yes:
                well_data.leakiness = True
                if WellCondition.leakage_window is None:
                    WellCondition.leakage_window = LeakageWindow()
                    WellCondition.leakage_window.setWindowTitle("Геофизические исследования")
                    # WellCondition.leakage_window.setGeometry(200, 400, 300, 400)
                    WellCondition.leakage_window.show()

                    MyWindow.pause_app()
                    well_data.dict_leakiness = WellCondition.leakage_window.add_work()
                    # print(f'словарь нарушений {well_data.dict_leakiness}')
                    well_data.pause = True
                    WellCondition.leakage_window = None  # Discard reference.


                else:
                    WellCondition.leakage_window.close()  # Close window.
                    WellCondition.leakage_window = None  # Discard reference.


            else:
                well_data.leakiness = False


class Well_expected_pick_up(FindIndexPZ):

    def __init__(self, ws):

        super().__init__(ws)
        self.ws = ws
        # self.read_well(ws, well_data.data_x_min._value, well_data.data_x_max._value)

    def read_well(self, ws, begin_index, cancel_index):

        for row_index, row in enumerate(ws.iter_rows(min_row=begin_index, max_row=cancel_index)):
            row_index += begin_index
            # print(row_index)
            for col, cell in enumerate(row[0:15]):
                # print(row_index)
                value = cell.value
                if value:

                    if 'прием' in str(value).lower() or 'qж' in str(value).lower():
                        well_data.expected_Q = row[col + 1].value
                        # print(well_data.expected_Q)
                        well_data.expected_Q = FindIndexPZ.definition_is_None(self, well_data.expected_Q, row_index,
                                                                              col + 1, 1)
                        # print(f'после {well_data.expected_Q}')
                    if 'зак' in str(value).lower() or 'давл' in str(value).lower() or 'p' in str(value).lower():
                        well_data.expected_P = row[col + 1].value
                        well_data.expected_P = FindIndexPZ.definition_is_None(self, well_data.expected_P, row_index,
                                                                              col + 1, 1)

                    if 'qж' in str(value).lower():
                        well_data.Qwater = str(row[col + 1].value).strip().replace(' ', '').replace('м3/сут', '')
                        well_data.Qwater = FindIndexPZ.definition_is_None(self, well_data.Qwater, row_index, col + 1, 1)

                    if 'qн' in str(value).lower():
                        well_data.Qoil = str(row[col + 1].value).replace(' ', '').replace('т/сут', '')
                        well_data.Qoil = FindIndexPZ.definition_is_None(self, well_data.Qoil, row_index, col + 1, 1)
                    if 'воды' in str(value).lower() and "%" in str(value).lower():
                        try:
                            proc_water = str(row[col + 1].value).replace(' ', '').replace('%', '')

                            proc_water = FindIndexPZ.definition_is_None(self, proc_water, row_index, col + 1, 1)
                            well_data.proc_water = int(float(proc_water)) if float(proc_water) > 1 else round(
                                float(proc_water) * 100,
                                0)
                        except:
                            print(f'ошибка в определение')

            try:
                well_data.expected_pick_up[well_data.expected_Q] = well_data.expected_P
            except:
                print('Ошибка в определении ожидаемых показателей')


class Well_data(FindIndexPZ):

    def __init__(self, ws):

        super().__init__(ws)
        self.ws = ws
        # self.read_well(self.ws, well_data.cat_well_max._value, well_data.data_pvr_min._value)

    def read_well(self, ws, begin_index, cancel_index):
        from main import MyWindow

        well_data.well_area = ProtectedIsNonNone('не корректно')
        well_data.well_number = ProtectedIsNonNone('не корректно')
        well_data.inv_number = ProtectedIsNonNone('не корректно')
        well_data.cdng = ProtectedIsNonNone('не корректно')
        well_data.bottomhole_drill = ProtectedIsNonNone('не корректно')
        well_data.bottomhole_artificial = ProtectedIsNonNone('не корректно')
        well_data.max_angle = ProtectedIsNonNone('не корректно')
        well_data.max_angle_H = ProtectedIsNonNone('не корректно')
        well_data.stol_rotora = ProtectedIsNonNone('не корректно')
        well_data.column_conductor_diametr = ProtectedIsNonNone('не корректно')
        well_data.column_conductor_wall_thickness = ProtectedIsNonNone('не корректно')
        well_data.column_conductor_lenght = ProtectedIsNonNone('не корректно')
        well_data.level_cement_direction = ProtectedIsNonNone('не корректно')
        well_data.column_direction_diametr = ProtectedIsNonNone('отсут')
        well_data.column_direction_wall_thickness = ProtectedIsNonNone('отсут')
        well_data.column_direction_lenght = ProtectedIsNonNone('отсут')
        well_data.level_cement_conductor = ProtectedIsNonNone('отсут')
        well_data.column_conductor_wall_thickness = ProtectedIsNonNone('отсут')
        well_data.column_conductor_lenght = ProtectedIsNonNone('отсут')
        well_data.column_diametr = ProtectedIsNonNone('не корректно')
        well_data.column_wall_thickness = ProtectedIsNonNone('не корректно')
        well_data.shoe_column = ProtectedIsNonNone('не корректно')
        well_data.level_cement_column = ProtectedIsNonNone('не корректно')
        well_data.pressuar_mkp = ProtectedIsNonNone('не корректно')
        well_data.column_additional_diametr = ProtectedIsNonNone('не корректно')
        well_data.column_additional_wall_thickness = ProtectedIsNonNone('не корректно')
        well_data.head_column_additional = ProtectedIsNonNone('не корректно')
        well_data.shoe_column_additional = ProtectedIsNonNone('не корректно')

        for row_index, row in enumerate(ws.iter_rows(min_row=begin_index, max_row=cancel_index)):
            row_index += begin_index

            for col, cell in enumerate(row):
                value = cell.value
                if value:
                    if 'площадь' in str(value):  # определение номера скважины
                        well_data.well_number = ProtectedIsDigit(row[col - 1].value)
                        well_data.well_area = ProtectedIsNonNone(row[col + 1].value)
                    elif 'месторождение ' in str(value):  # определение номера скважины
                        well_data.well_oilfield = ProtectedIsNonNone(row[col + 2].value)
                    elif 'Инв. №' in str(value):
                        well_data.inv_number = ProtectedIsNonNone(row[col + 1].value)
                    elif 'цех' == value:
                        well_data.cdng = ProtectedIsDigit(row[col + 1].value)
                        # print(f' ЦДНГ {well_data.cdng._value}')
                    elif 'пробуренный забой' in str(value).lower():
                        well_data.bottomhole_drill = ProtectedIsDigit(row[col + 2].value)
                        well_data.bottomhole_drill = FindIndexPZ.definition_is_None(self, well_data.bottomhole_drill,
                                                                                    row_index, col, 2)

                        well_data.bottomhole_artificial = ProtectedIsDigit(row[col + 4].value)
                        # print(f'пробуренный забой {well_data.bottomhole_artificial}')
                        well_data.bottomhole_artificial = \
                            FindIndexPZ.definition_is_None(self, well_data.bottomhole_artificial, row_index, col, 5)
                        # print(f'пробуренный забой {well_data.bottomhole_artificial}')

                    elif 'зенитный угол' in str(value).lower():
                        well_data.max_angle = ProtectedIsDigit(row[col + 4].value)
                        for index, col1 in enumerate(row):
                            if 'на глубине' in str(col1.value):
                                well_data.max_angle_H = ProtectedIsDigit(row[index + 1].value)

                    elif 'текущий забой' in str(value).lower() and len(value) < 15:
                        well_data.current_bottom = row[col + 2].value
                        well_data.current_bottom = \
                            FindIndexPZ.definition_is_None(self, well_data.current_bottom, row_index, col, 2)

                        well_data.bottom = well_data.current_bottom
                    elif '10. Расстояние от стола ротора до среза муфты э/колонны ' in str(value):
                        well_data.stol_rotora = FindIndexPZ.definition_is_None(
                            self, ProtectedIsDigit(row[col + 5].value), row_index, col, 1)

                    elif 'Направление' in str(value) and 'Шахтное направление' not in str(value) and \
                            ws.cell(row=row_index + 1, column=col + 1).value != None and \
                            FindIndexPZ.check_str_None(self, row[col + 3].value) != '0':
                        well_data.column_direction_True = True
                        if well_data.column_direction_True:
                            for col1, cell in enumerate(row):
                                if 'Уровень цемента' in str(cell.value):
                                    if 'уст' in str(row[col1 + 2].value).lower() or str(row[col1 + 2].value).isdigit():
                                        well_data.level_cement_direction = ProtectedIsDigit(0)
                                    else:
                                        well_data.level_cement_direction = ProtectedIsDigit(
                                            str(row[col1 + 2].value.split('-')[0]).replace(" ", ""))
                        else:
                            well_data.level_cement_direction = ProtectedIsNonNone('отсут')
                        try:
                            column_direction_data = row[col + 3].value.split('(мм),')
                            try:
                                well_data.column_direction_diametr = ProtectedIsDigit(column_direction_data[0])
                            except:
                                well_data.column_direction_diametr = ProtectedIsNonNone('не корректно')

                            try:
                                well_data.column_direction_wall_thickness = ProtectedIsDigit(
                                    column_direction_data[1].replace(' ', ''))
                            except:
                                well_data.column_direction_wall_thickness = ProtectedIsNonNone('не корректно')
                            try:
                                try:
                                    well_data.column_direction_lenght = ProtectedIsDigit(
                                        column_direction_data[2].split('-')[1].replace('(м)', '').replace(" ", ""))
                                except:
                                    well_data.column_direction_lenght = ProtectedIsDigit(
                                        column_direction_data[2].replace('(м)', '').replace(" ", ""))

                            except:
                                well_data.column_direction_lenght = ProtectedIsNonNone('не корректно')
                        except:
                            well_data.column_direction_diametr = ProtectedIsNonNone('не корректно')
                            well_data.column_direction_wall_thickness = ProtectedIsNonNone('не корректно')
                            well_data.column_direction_lenght = ProtectedIsNonNone('не корректно')

                    elif 'Кондуктор' in str(value) and \
                            FindIndexPZ.check_str_None(self, row[col + 3].value) != '0':

                        for col1, cell in enumerate(row):
                            if 'Уровень цемента' in str(cell.value):
                                try:
                                    if 'уст' in str(row[col1 + 2].value).lower() or str(row[col1 + 2].value).isdigit():
                                        well_data.level_cement_conductor = ProtectedIsDigit(0)
                                    else:
                                        well_data.level_cement_conductor = ProtectedIsDigit(
                                            str(row[col1 + 2].value.split('-')[0]).replace(' ', ''))
                                except:
                                    well_data.level_cement_conductor = ProtectedIsNonNone('не корректно')
                        try:
                            column_conductor_data = str(row[col + 3].value).split('(мм),')
                            try:
                                well_data.column_conductor_diametr = ProtectedIsDigit(column_conductor_data[0].strip())
                            except:
                                well_data.column_conductor_diametr = ProtectedIsNonNone('не корректно')

                            try:
                                well_data.column_conductor_wall_thickness = \
                                    ProtectedIsDigit(column_conductor_data[1].strip())
                            except:
                                well_data.column_conductor_wall_thickness = ProtectedIsNonNone('не корректно')
                            try:
                                try:
                                    well_data.column_conductor_lenght = ProtectedIsDigit(
                                        column_conductor_data[2].split('-')[1].replace('(м)', '').strip())
                                except:
                                    well_data.column_conductor_lenght = ProtectedIsDigit(
                                        str(column_conductor_data[2].replace('(м)', '')).strip())
                            except:
                                well_data.column_conductor_lenght = ProtectedIsNonNone('не корректно')

                        except:
                            well_data.column_conductor_diametr = ProtectedIsNonNone('не корректно')
                            well_data.column_conductor_wall_thickness = ProtectedIsNonNone('не корректно')
                            well_data.column_conductor_lenght = ProtectedIsNonNone('не корректно')
                    elif str(
                            value) == '4. Эксплуатационная колонна (диаметр(мм), толщина стенки(мм), глубина спуска(м))':

                        try:
                            data_main_production_string = str(ws.cell(row=row_index + 1, column=col + 1).value).split(
                                '(мм),', )
                            try:
                                well_data.column_diametr = ProtectedIsDigit(
                                    float(str(data_main_production_string[0]).replace(',', '.')))
                            except:
                                well_data.column_diametr = ProtectedIsNonNone('не корректно')
                            try:
                                well_data.column_wall_thickness = ProtectedIsDigit(
                                    float(str(data_main_production_string[1]).replace(',', '.')))
                            except:
                                well_data.column_wall_thickness = ProtectedIsNonNone('не корректно')
                            try:
                                if len(data_main_production_string[-1].split('-')) == 2:

                                    well_data.shoe_column = ProtectedIsDigit(
                                        FindIndexPZ.check_str_None(self,
                                                                   data_main_production_string[-1].split('-')[-1]))
                                else:
                                    well_data.shoe_column = ProtectedIsDigit(
                                        FindIndexPZ.check_str_None(self, data_main_production_string[-1]))
                            except:
                                well_data.shoe_column = ProtectedIsNonNone('не корректно')
                        except ValueError:
                            well_data.column_diametr = ProtectedIsNonNone('не корректно')
                            well_data.column_wall_thickness = ProtectedIsNonNone('не корректно')
                            well_data.shoe_column = ProtectedIsNonNone('не корректно')

                    elif 'Уровень цемента за колонной' in str(value):
                        well_data.level_cement_column = ProtectedIsDigit(row[col + 3].value)
                        well_data.level_cement_column = FindIndexPZ.definition_is_None(self,
                                                                                       well_data.level_cement_column,
                                                                                       row_index,
                                                                                       col, 1)
                    elif 'Рмкп ( э/к и' in str(cell):
                        well_data.pressuar_mkp = ProtectedIsNonNone(row[col + 2].value)
                    elif 'онструкция хвостовика' in str(value):

                        data_column_additional = FindIndexPZ.check_str_None(self, ws.cell(row=row_index + 2,
                                                                                          column=col + 2).value)
                        # print(f'доп колона {data_column_additional.strip(), FindIndexPZ.check_str_None(self,data_column_additional.strip())}')
                        if data_column_additional != '0':
                            well_data.column_additional = True
                        if well_data.column_additional is True:
                            try:
                                well_data.head_column_additional = ProtectedIsDigit(data_column_additional[0])
                            except:
                                well_data.head_column_additional = ProtectedIsNonNone('не корректно')
                            try:
                                well_data.shoe_column_additional = ProtectedIsDigit(data_column_additional[1])
                            except:
                                well_data.shoe_column_additional = ProtectedIsNonNone('не корректно')

                            try:
                                try:
                                    data_add_column = FindIndexPZ.check_str_None(self,
                                                                                 ws.cell(row=row_index + 2,
                                                                                         column=col + 4).value)
                                    # print(f' доп колонна {data_add_column}')
                                    well_data.column_additional_diametr = ProtectedIsDigit(data_add_column[0])

                                except:
                                    well_data.column_additional_diametr = ProtectedIsDigit(
                                        FindIndexPZ.check_str_None(self,
                                                                   ws.cell(row=row_index + 2, column=col + 4).value))

                                try:
                                    data_add_column = FindIndexPZ.check_str_None(self,
                                                                                 ws.cell(row=row_index + 2,
                                                                                         column=col + 4).value)
                                    well_data.column_additional_wall_thickness = ProtectedIsDigit(data_add_column[1])
                                except:

                                    well_data.column_additional_wall_thickness = ProtectedIsDigit(
                                        FindIndexPZ.check_str_None(self,
                                                                   ws.cell(row=row_index + 2, column=col + 6).value))

                            except:
                                well_data.column_additional_wall_thickness = ProtectedIsNonNone('не корректно')
                                well_data.column_additional_diametr = ProtectedIsNonNone('не корректно')
                        else:
                            well_data.column_additional_diametr = ProtectedIsNonNone('отсут')
                            well_data.column_additional_wall_thickness = ProtectedIsNonNone('отсут')
                            well_data.head_column_additional = ProtectedIsNonNone('отсут')
                            well_data.shoe_column_additional = ProtectedIsNonNone('отсут')

        if self.data_window is None:
            self.data_window = DataWindow(self)
            self.data_window.setWindowTitle("Сверка данных")
            self.data_window.setGeometry(100, 100, 300, 400)

            self.data_window.show()
            MyWindow.pause_app()
            well_data.pause = True
            self.data_window = None

        if well_data.max_angle._value > 45 or 'gnkt' in well_data.work_plan:
            angle_true_question = QMessageBox.question(self, 'Зенитный угол', 'Зенитный угол больше 45 градусов, '
                                                                              'есть данные иклинометрии?')
            if angle_true_question == QMessageBox.StandardButton.Yes:
                well_data.angle_data = Well_data.read_angle_well()

                print(well_data.angle_data)

        well_data.nkt_diam = 73 if well_data.column_diametr._value > 110 else 60
        well_data.nkt_template = 59.6 if well_data.column_diametr._value > 110 else 47.9

        if well_data.column_additional:
            if well_data.current_bottom > well_data.shoe_column_additional._value:
                well_data.open_trunk_well = True
            else:
                well_data.open_trunk_well = False
        else:
            if well_data.current_bottom > well_data.shoe_column._value:
                well_data.open_trunk_well = True
            else:
                well_data.open_trunk_well = False
        if str(well_data.well_number._value) in ['882', '667', "3375", '2064', '1160', '1034',
                                                 '1685', '1686', "1571", "3374", "2360", "3354",
                                                 '878', '124', '549', '168']:
            QMessageBox.warning(self, 'Канатные технологии', f'Скважина согласована на канатные технологии')
            well_data.konte_true = True

    @staticmethod
    def read_angle_well():
        fname, _ = QtWidgets.QFileDialog.getOpenFileName(None, 'Выберите файл', '.',
                                                         "Файлы Exсel (*.xlsx);;Файлы Exсel (*.xls)")
        # Загрузка файла Excel
        wb = load_workbook(fname)
        ws = wb.active
        angle_data = []
        depth_column = ''
        row_data = ''
        angle_column = ''
        curvature_column = ''
        for index_row, row in enumerate(ws.iter_rows(min_row=1, max_row=50, values_only=True)):
            for col, value in enumerate(row):
                if not value is None:
                    if 'глубина' in str(value).lower() and col <= 7:
                        depth_column = col
                        row_data = index_row
                        print(f'угол{depth_column, row_data}')
                    elif ('Угол, гpад' in str(value) and col <= 7) or 'зенитный' in str(value).lower():
                        angle_column = col
                        print(f'угол_ e {angle_column}')
                    elif 'кривизна' in str(value).lower() or 'гр./10' in str(value).lower():
                        curvature_column = col
                        print(f'угол_e{curvature_column}')
        if depth_column == '':
            depth_column, ok = QInputDialog.getInt(None, 'номер столбца', 'Программа не смогла найти номер столбца с '
                                                                          'указанием значений глубин')
        if row_data == '':
            row_data, ok = QInputDialog.getInt(None, 'номер столбца', 'Программа не смогла найти номер строки с '
                                                                      'указанием данных строки')
        if angle_column == '':
            angle_column, ok = QInputDialog.getInt(None, 'номер столбца', 'Программа не смогла найти номер столбца с '
                                                                          'указанием значений угла')
        if curvature_column == '':
            curvature_column, ok = QInputDialog.getInt(None, 'номер столбца',
                                                       'Программа не смогла найти номер столбца с '
                                                       'указанием интенсивности набора угла')

        # Вставка данных в таблицу
        for index_row, row in enumerate(ws.iter_rows(min_row=row_data, values_only=True)):
            if str(row[depth_column]).replace(',', '').replace('.', '').isdigit():
                angle_data.append((row[depth_column], row[angle_column], row[curvature_column]))
        return angle_data


class Well_perforation(FindIndexPZ):
    def __init__(self, ws):

        super().__init__(ws)
        self.ws = ws

        # self.read_well(self.ws, well_data.data_pvr_min._value, well_data.data_pvr_max._value + 1)

    def read_well(self, ws, begin_index, cancel_index):
        from work_py.alone_oreration import is_number, calculationFluidWork
        from main import MyWindow
        # print(f'перфорация доп {well_data.dict_perforation}')
        well_data.old_version = True
        col_old_open_index = 0
        bokov_stvol = False
        osnov_stvol = False
        col_plast_index = -1
        col_vert_index = 0
        col_roof_index = 0
        col_sole_index =0
        col_open_index = 0
        col_close_index = 0
        col_udlin_index = 0
        col_pressuar_index = 0
        col_date_pressuar_index = 0

        if len(well_data.dict_perforation) == 0:
            for row in ws.iter_rows(min_row=begin_index + 1, max_row=cancel_index + 2, values_only=True):
                # print(row)
                for col_index, column in enumerate(row[:14]):
                    if 'оризонт'.lower() in str(column).lower() or 'пласт/'.lower() in str(column).lower():
                        col_plast_index = col_index - 1
                        print(f'пласт {col_plast_index}')
                    if 'по вертикали'.lower() in str(column).lower():
                        col_vert_index = col_index - 1
                        # print(f'вер {col_index}')
                    if 'кровля'.lower() in str(column).lower():
                        col_roof_index = col_index - 1
                        # print(f'кров {col_index}')
                    if 'подошва'.lower() in str(column).lower():
                        # print(f'подо {col_index}')
                        col_sole_index = col_index - 1
                    if 'вскрытия'.lower() in str(column).lower():
                        # print(f'вскр {col_index}')
                        col_open_index = col_index - 1
                    if 'отключ'.lower() in str(column).lower() and col_index < 8:
                        # print(f'октл {col_index}')
                        col_close_index = col_index - 1
                    if 'удлине'.lower() in str(column).lower():
                        # print(f'удл {col_index}')
                        col_udlin_index = col_index - 1
                    if 'Рпл'.lower() in str(column).lower():
                        print(f'Рпл {col_index}')
                        col_pressuar_index = col_index - 1
                    if 'замера' in str(column).lower():
                        print(f'заме {col_index}')
                        col_date_pressuar_index = col_index - 1
                    if 'вскрыт'.lower() in str(column).lower() and 'откл'.lower() in str(column).lower():
                        well_data.old_version = True
                        col_old_open_index = col_index - 1
                    if "сновной" in str(column).lower():
                        osnov_stvol = True
                    if "боков" in str(column).lower():
                        bokov_stvol = True
            column_list = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M']
            if col_date_pressuar_index == 0:
                col_date_pressuar_index = column_index_from_string(QInputDialog.getItem(self,
                                                                 'Ошибка', 'Программа не смогла определить колонку '
                                                                  'в таблице ПВР где указано в дата замера',
                                                              column_list, 11)[0]) - 2
                print(f'fjjар {col_date_pressuar_index}')
            if col_pressuar_index == 0:
                col_pressuar_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано в Рпл',
                                                              column_list, 10)[0]) - 2
            if col_udlin_index == 0:
                col_udlin_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано в удлинение',
                                                              column_list, 9)[0]) - 2
            if col_close_index == 0:
                col_close_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано дата отключения',
                                                             column_list, 6)[0]) - 2
            if col_open_index == 0:
                col_open_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано дата вскрытия',
                                                              column_list, 5)[0]) - 2
            if col_sole_index == 0:
                col_sole_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано Подошва ИП',
                                                             column_list, 4)[0]) - 2
            if col_roof_index == 0:
                col_roof_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано кровля ИП',
                                                              column_list, 3)[0]) - 2
            if col_vert_index == 0:
                col_vert_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано вертикаль',
                                                              column_list, 2)[0]) - 2
            if col_plast_index == -1:
                col_plast_index = column_index_from_string(QInputDialog.getItem(self, 'Ошибка', 'Программа не смогла определить колонку '
                                                                              'в таблице ПВР где указано пласт',
                                                              column_list, 1)[0]) - 2
                print(f'fjjар {col_plast_index}')
            if osnov_stvol is True and bokov_stvol is True:
                mes = QMessageBox.warning(self, 'ОШИБКА', 'Ошибка в определении рабочий интервалов перфорации')
                begin_index = QInputDialog.getInt(self, 'Индекс начала', 'ВВедите индекс начала рабочих интервалов ПВР',
                                                  0, 0, 300)[0] - 3

                cancel_index = QInputDialog.getInt(self, 'Индекс начала',
                                                   'ВВедите индекс окончания рабочих интервалов ПВР', 0, 0, 300)[0] - 2

            perforations_intervals = []
            for row_index, row in enumerate(
                    ws.iter_rows(min_row=begin_index + 3, max_row=cancel_index + 2)):
                lst = []

                if str(row[col_roof_index + 1].value).replace('.', '').replace(',', '').isdigit():
                    if row[1].value != None:
                        plast = row[1].value
                        lst.append(plast)
                    else:
                        lst.append(plast)

                    for col in row[2:13]:
                        lst.append(col.value)

                if all([str(i).strip() == 'None' or i is None for i in lst]) is False:
                    perforations_intervals.append(lst)

            for ind, row in enumerate(perforations_intervals):
                plast = row[col_plast_index]
                # print(f'пласт {plast}')
                if plast is None:
                    plast = perforations_intervals[ind - 1][col_plast_index]
                # print(f'пластs {plast}')

                if any(['проект' in str((i)).lower() or 'не пер' in str((i)).lower() for i in row]) is False and all(
                        [str(i).strip() is None for i in row]) is False and is_number(row[col_roof_index]) is True \
                        and is_number(row[col_sole_index]) is True:
                    # print(f'5 {row}')

                    if is_number(str(row[col_vert_index]).replace(',', '.')) is True:
                        well_data.dict_perforation.setdefault(plast,
                                                              {}).setdefault('вертикаль',
                                                                             []).append(float(
                            str(row[col_vert_index]).replace(',', '.')))
                    if any(['фильтр' in str(i).lower() for i in row]):
                        well_data.dict_perforation.setdefault(plast, {}).setdefault('отрайбировано', True)
                    else:
                        well_data.dict_perforation.setdefault(plast, {}).setdefault('отрайбировано', False)
                    well_data.dict_perforation.setdefault(plast, {}).setdefault('Прошаблонировано', False)
                    roof_int = round(float(str(row[col_roof_index]).replace(',', '.')), 1)
                    sole_int = round(float(str(row[col_sole_index]).replace(',', '.')), 1)
                    well_data.dict_perforation.setdefault(plast, {}).setdefault('интервал', []).append(
                        (roof_int, sole_int))
                    well_data.dict_perforation_short.setdefault(plast, {}).setdefault('интервал', []).append(
                        (roof_int, sole_int))
                    # for interval in list(well_data.dict_perforation[plast]["интервал"]):
                    # print(interval)

                    well_data.dict_perforation.setdefault(plast, {}).setdefault('вскрытие', []).append(
                        row[col_open_index])

                    if col_old_open_index != col_open_index:
                        if row[col_close_index] is None or row[col_close_index] == '-':
                            print(f'вскрыт {plast} {row[col_close_index], col_close_index}')

                            well_data.dict_perforation.setdefault(plast, {}).setdefault('отключение', False)
                            well_data.dict_perforation_short.setdefault(plast, {}).setdefault('отключение', False)
                        else:
                            print(f'отключ {plast} {row[col_close_index], col_close_index}')
                            well_data.dict_perforation.setdefault(plast, {}).setdefault('отключение', True)
                            well_data.dict_perforation_short.setdefault(plast, {}).setdefault('отключение', True)
                    else:
                        if isinstance(row[col_old_open_index], datetime):
                            well_data.dict_perforation.setdefault(plast, {}).setdefault('отключение', False)
                            well_data.dict_perforation_short.setdefault(plast, {}).setdefault('отключение', False)
                        else:
                            well_data.dict_perforation.setdefault(plast, {}).setdefault('отключение', True)
                            well_data.dict_perforation_short.setdefault(plast, {}).setdefault('отключение', True)

                    zhgs = 1.01
                    if str(row[col_pressuar_index]).replace(',', '').replace('.', '').isdigit() and row[col_vert_index]:
                        data_p = float(str(row[col_pressuar_index]).replace(',', '.'))
                        well_data.dict_perforation.setdefault(plast, {}).setdefault('давление',
                                                                                    []).append(round(data_p, 1))
                        well_data.dict_perforation_short.setdefault(plast, {}).setdefault('давление',
                                                                                          []).append(round(data_p, 1))
                        zhgs = calculationFluidWork(float(str(row[col_vert_index]).replace(',', '.')), float(data_p))
                    else:
                        well_data.dict_perforation.setdefault(plast, {}).setdefault('давление',
                                                                                    []).append('0')
                        well_data.dict_perforation_short.setdefault(plast, {}).setdefault('давление',
                                                                                          []).append('0')
                    if zhgs:
                        well_data.dict_perforation.setdefault(plast, {}).setdefault('рабочая жидкость', []).append(zhgs)
                    if row[col_date_pressuar_index]:
                        well_data.dict_perforation.setdefault(
                            plast, {}).setdefault('замер', []).append(row[col_date_pressuar_index])

                elif any([str((i)).lower() == 'проект' for i in row]) is True and all(
                        [str(i).strip() is None for i in row]) is False and is_number(row[col_roof_index]) is True \
                        and is_number(
                    float(
                        str(row[col_roof_index]).replace(',',
                                                         '.'))) is True:  # Определение проектных интервалов перфорации
                    roof_int = round(float(str(row[col_roof_index]).replace(',', '.')), 1)
                    sole_int = round(float(str(row[col_sole_index]).replace(',', '.')), 1)

                    if row[col_vert_index] != None:
                        well_data.dict_perforation_project.setdefault(
                            plast, {}).setdefault('вертикаль', []).append(round(float(str(
                            row[col_vert_index]).replace(',', '.')), 1))
                    well_data.dict_perforation_project.setdefault(
                        plast, {}).setdefault('интервал', []).append((roof_int, sole_int))

                    if row[col_pressuar_index] != None:
                        well_data.dict_perforation_project.setdefault(plast, {}).setdefault('давление', []).append(
                            round(FindIndexPZ.check_str_None(self, row[col_pressuar_index]), 1))
                    well_data.dict_perforation_project.setdefault(plast, {}).setdefault('рабочая жидкость', []).append(
                        calculationFluidWork(row[col_vert_index], row[col_pressuar_index]))

            if len(well_data.dict_perforation_project) != 0:
                well_data.plast_project = list(well_data.dict_perforation_project.keys())

            # объединение интервалов перфорации если они пересекаются
            for plast, value in well_data.dict_perforation.items():
                intervals = value['интервал']
                merged_segments = list()
                for roof_int, sole_int in sorted(list(intervals), key=lambda x: x[0]):

                    if not merged_segments or roof_int > merged_segments[-1][1]:
                        merged_segments.append((roof_int, sole_int))
                    else:
                        merged_segments[-1] = [merged_segments[-1][0], max(sole_int, merged_segments[-1][1])]

                well_data.dict_perforation[plast]['интервал'] = merged_segments

        if self.perforation_correct_window2 is None:
            self.perforation_correct_window2 = PerforationCorrect(self)
            self.perforation_correct_window2.setWindowTitle("Сверка данных перфорации")
            # self.perforation_correct_window2.setGeometry(200, 400, 100, 400)

            self.perforation_correct_window2.show()
            MyWindow.pause_app()
            well_data.pause = True
            self.perforation_correct_window2 = None
            definition_plast_work(self)
        else:
            self.perforation_correct_window2.close()
            self.perforation_correct_window2 = None

        if len(well_data.dict_perforation_project) != 0:
            well_data.plast_project = list(well_data.dict_perforation_project.keys())


class Well_Category(FindIndexPZ):

    def __init__(self, ws):

        super(Well_Category, self).__init__(ws)
        # self.read_well(ws, well_data.cat_well_min._value, well_data.data_well_min._value)

    def read_well(self, ws, begin_index, cancel_index):
        from main import MyWindow
        if well_data.data_in_base is False:
            print(f'индекс катего {begin_index, cancel_index}')
            for row in range(begin_index, cancel_index):
                for col in range(1, 13):
                    cell = ws.cell(row=row, column=col).value
                    if cell:
                        if str(cell) in ['атм'] and ws.cell(row=row, column=col - 2).value:
                            well_data.cat_P_1.append(ws.cell(row=row, column=col - 2).value)
                            # print(well_data.cat_P_P)
                            well_data.cat_P_P.append(ws.cell(row=row, column=col - 1).value)

                        elif str(cell) in ['%', 'мг/л', 'мг/дм3', 'мг/м3', 'мг/дм', 'мгдм3']:
                            if str(cell) == '%':
                                if ws.cell(row=row, column=col - 2).value is None:
                                    well_data.cat_h2s_list.append(ws.cell(row=row-1, column=col - 2).value)
                                else:
                                    well_data.cat_h2s_list.append(ws.cell(row=row, column=col - 2).value)
                                if str(ws.cell(row=row, column=col - 1).value).strip() in ['', '-', '0', 'None'] or \
                                        'отс' in str(ws.cell(row=row, column=col - 1).value).lower():
                                    well_data.h2s_pr.append(0)
                                else:
                                    well_data.h2s_pr.append(float(str(ws.cell(row=row, column=col - 1).value).replace(',', '.')))
                            if str(cell) in ['мг/л', 'мг/дм3', 'мг/дм', 'мгдм3']:
                                if str(ws.cell(row=row, column=col - 1).value).strip() in ['', '-', '0', 'None'] or \
                                        'отс' in str(ws.cell(row=row, column=col - 1).value).lower():
                                    well_data.h2s_mg.append(0)
                                    print(well_data.h2s_mg)
                                else:
                                    print(ws.cell(row=row, column=col - 1).value)
                                    well_data.h2s_mg.append(float(str(ws.cell(row=row, column=col - 1).value).replace(',', '.')))

                            if str(cell) in ['мг/м3'] and ws.cell(row=row-1, column=col - 1).value not in \
                                ['мг/л', 'мг/дм3', 'мг/дм', 'мгдм3'] and ws.cell(row=row+1, column=col - 1).value not in  \
                                ['мг/л', 'мг/дм3', 'мг/дм', 'мгдм3']:
                                if str(ws.cell(row=row, column=col - 1).value).strip() in ['', '-', '0', 'None'] or \
                                        'отс' in str(ws.cell(row=row, column=col - 1).value).lower():
                                    well_data.h2s_mg.append(0)
                                    print(f'1{well_data.h2s_mg}')
                                else:
                                    well_data.h2s_mg.append(float(str(
                                            FindIndexPZ.check_str_None(self, str(ws.cell(row=row, column=col - 1).value).replace(',', '.')))) / 1000)
                                    print(f'2{well_data.h2s_mg}')

                        elif str(cell) == 'м3/т':
                            well_data.cat_gaz_f_pr.append(ws.cell(row=row, column=col - 2).value)
                            if 'отс' in str(ws.cell(row=row, column=col - 1).value):
                                well_data.gaz_f_pr.append(0)
                            else:
                                well_data.gaz_f_pr.append(float(
                                    str(ws.cell(row=row, column=col - 1).value).replace(',', '.')))
            if len(well_data.cat_h2s_list) == 0:
                mes = QMessageBox.warning(self, 'ОШИБКА', 'Приложение не смогла найти значение '
                                                          'сероводорода в %')
                well_data.pause = True
                self.pause_app()



            if self.data_window is None:
                self.data_window = CategoryWindow(self)
                self.data_window.setWindowTitle("Сверка данных")
                # self.data_window.setGeometry(200, 200, 200, 200)
                self.data_window.show()
                MyWindow.pause_app()
                well_data.pause = True
            else:
                self.data_window.close()
                self.data_window = None

            if len(well_data.h2s_pr) == 0:
                mes = QMessageBox.warning(self, 'Ошибка', 'Программа не смогла найти данные по содержания '
                                                          'сероводорода в процентах')
                h2s_pr, _ = QInputDialog.getDouble(self, 'сероводород в процентах',
                                                   'Введите значение серовородода в процентах', 0, 0, 100, 5)

                well_data.h2s_pr.append(h2s_pr)

            well_data.category_pressuar = well_data.cat_P_1[0]
            # print(f'категория по давлению {well_data.category_pressuar}')
            well_data.category_h2s = well_data.cat_h2s_list[0]
            well_data.category_gf = well_data.cat_gaz_f_pr[0]

            thread = ExcelWorker()

            well_data.without_damping = thread.check_well_existence(
                well_data.well_number._value, well_data.well_area._value, well_data.region)

            try:
                categoty_pressure_well, categoty_h2s_well, categoty_gf, data = thread.check_category(
                    well_data.well_number, well_data.well_area, well_data.region)

                if categoty_pressure_well:
                    if str(categoty_pressure_well) != str(well_data.category_pressuar):
                        mes = QMessageBox.warning(None, 'Некорректная категория давления',
                                                  f'согласно классификатора от {data} категория скважина '
                                                  f'по давлению {categoty_pressure_well}')
                if categoty_h2s_well:
                    if str(well_data.cat_h2s_list[0]) != str(well_data.category_h2s):
                        # print(str(well_data.cat_h2s_list[0]), well_data.category_h2s)
                        #
                        mes = QMessageBox.warning(None, 'Некорректная категория давления',
                                                  f'согласно классификатора от {data} категория скважина '
                                                  f'по сероводороду {categoty_h2s_well}')
                if categoty_gf:
                    if categoty_gf != well_data.category_gf:
                        mes = QMessageBox.warning(None, 'Некорректная категория давления',
                                                  f'согласно классификатора от {data} категория скважина '
                                                  f'по газовому фактору {categoty_gf}')
            except:
                mes = QMessageBox.warning(self, 'Ошибка', 'Скважина не найдена в классификаторе')
