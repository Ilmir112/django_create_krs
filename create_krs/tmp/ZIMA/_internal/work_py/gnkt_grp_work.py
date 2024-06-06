import sqlite3
from datetime import datetime

import psycopg2
from PyQt5.QtWidgets import QInputDialog, QMainWindow, QTabWidget, QWidget, QTableWidget, QApplication, QLabel, \
    QLineEdit, QGridLayout, QComboBox, QPushButton, QMessageBox
# from PyQt5.uic.properties import QtWidgets
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from collections import namedtuple

import well_data
from gnkt_data.gnkt_data import gnkt_1, gnkt_2, gnkt_dict, read_database_gnkt
from krs import TabPageGno
from perforation_correct import PerforationCorrect

import block_name
import main
import plan
from block_name import razdel_1
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook

from work_py.alone_oreration import well_volume


class TabPageDp(QWidget):
    def __init__(self):
        super().__init__()
        self.gnkt_number_label = QLabel('Номер флота ГНКТ')
        self.gnkt_number_combo = QComboBox(self)
        self.gnkt_number_combo.addItems(gnkt_dict["Ойл-сервис"])
        if self.gnkt_number_combo.currentText() == 'ГНКТ №1':
            self.gnkt = gnkt_1
        elif self.gnkt_number_combo.currentText() == 'ГНКТ №2':
            self.gnkt = gnkt_2

        self.lenght_gnkt_label = QLabel('длина ГНКТ')
        self.lenght_gnkt_edit = QLineEdit(self)

        self.iznos_gnkt_label = QLabel('Износ трубы')
        self.iznos_gnkt_edit = QLineEdit(self)

        self.pvo_number_label = QLabel('Номер ПВО')
        self.pvo_number_edit = QLineEdit(self)

        self.pipe_mileage_label = QLabel('Пробег трубы')
        self.pipe_mileage_edit = QLineEdit(self)

        self.previous_well_label = QLabel('Предыдущая скважина')
        self.previous_well_combo = QComboBox(self)


        self.current_bottom_label = QLabel('необходимый текущий забой')
        self.current_bottom_edit = QLineEdit(self)
        self.current_bottom_edit.setText(f'{well_data.current_bottom}')

        self.fluid_label = QLabel("уд.вес жидкости глушения", self)
        self.fluid_edit = QLineEdit(self)
        if well_data.work_plan == 'gnkt_opz':
            self.fluid_edit.setText('1.18')
        else:
            self.fluid_edit.setText('1.01')
        if well_data.work_plan == 'gnkt_after_grp':
            self.osvoenie_label = QLabel('Необходимость освоения')
            self.osvoenie_combo = QComboBox(self)
            self.osvoenie_combo.addItems(['Да', 'Нет'])

        self.grid = QGridLayout(self)
        self.grid.addWidget(self.gnkt_number_label, 0, 2, 1, 5)
        self.grid.addWidget(self.gnkt_number_combo, 1, 2, 1, 5)
        self.grid.addWidget(self.lenght_gnkt_label, 2, 3)
        self.grid.addWidget(self.lenght_gnkt_edit, 3, 3)
        self.grid.addWidget(self.iznos_gnkt_label, 2, 4)
        self.grid.addWidget(self.iznos_gnkt_edit, 3, 4)

        self.grid.addWidget(self.pipe_mileage_label, 2, 5)
        self.grid.addWidget(self.pipe_mileage_edit, 3, 5)

        self.grid.addWidget(self.pvo_number_label, 2, 6)
        self.grid.addWidget(self.pvo_number_edit, 3, 6)

        self.grid.addWidget(self.previous_well_label, 2, 7)
        self.grid.addWidget(self.previous_well_combo, 3, 7)

        self.grid.addWidget(self.current_bottom_label, 4, 2)
        self.grid.addWidget(self.current_bottom_edit, 5, 2)
        self.grid.addWidget(self.fluid_label, 4, 3)
        self.grid.addWidget(self.fluid_edit, 5, 3)
        if well_data.work_plan == 'gnkt_after_grp':
            self.grid.addWidget(self.osvoenie_label, 4, 4)
            self.grid.addWidget(self.osvoenie_combo, 5, 4)
        self.gnkt_number_combo.currentTextChanged.connect(self.update_number_gnkt)
        self.previous_well_combo.currentTextChanged.connect(self.update_data_gnkt)

    def update_data_gnkt(self):
        previus_well = self.previous_well_combo.currentText()
        try:
            if previus_well:
                conn = psycopg2.connect(**well_data.postgres_conn_gnkt)

                cursor = conn.cursor()

                if 'ойл-сервис' in well_data.contractor.lower():
                    contractor = 'oil_service'
                print(previus_well)

                cursor.execute("""
                    SELECT * FROM gnkt_{contractor} WHERE well_number = %s;
                """.format(contractor=contractor), (previus_well,))

                result_gnkt = cursor.fetchone()
                print(result_gnkt)

                self.lenght_gnkt_edit.setText(str(result_gnkt[3]))
                self.iznos_gnkt_edit.setText(str(result_gnkt[5]))
                self.pipe_mileage_edit.setText(str(result_gnkt[6]))
                self.pvo_number_edit.setText(str(result_gnkt[10]))

                conn.close()
        except:
            print('Ошибка подключения')

    def update_number_gnkt(self, gnkt_number):
        if gnkt_number != '':

            well_previus_list = read_database_gnkt(well_data.contractor, gnkt_number)

            self.previous_well_combo.clear()
            self.previous_well_combo.addItems(list(map(str, well_previus_list)))

            # conn = sqlite3.connect('data_base\data_base_gnkt\gnkt_base.dp')
            # cursor = conn.cursor()
            # if 'ойл-сервис' in well_data.contractor.lower():
            #     contractor = 'oil_service'
            #
            # cursor.execute(f"SELECT * FROM gnkt_{contractor} WHERE gnkt_number =?", (gnkt_number,))
            #
            # result_gnkt = cursor.fetchone()
            #
            # self.lenght_gnkt_edit.setText(f'{result_gnkt[3]}')
            # self.iznos_gnkt_edit.setText(f'{result_gnkt[5]}')
            # self.pipe_mileage_edit.setText(f'{result_gnkt[6]}')
            # self.pvo_number_edit.setText(f'{result_gnkt[10]}')




class TabWidget(QTabWidget):
    def __init__(self):
        super().__init__()
        self.addTab(TabPageDp(), 'Данные по ГНКТ')


class GnktOsvWindow2(QMainWindow):
    def __init__(self, parent=None):
        super(QMainWindow, self).__init__(parent)

        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        self.tabWidget = TabWidget()
        self.dict_perforation = well_data.dict_perforation

        self.buttonAdd = QPushButton('Добавить данные в план работ')
        self.buttonAdd.clicked.connect(self.add_work)
        vbox = QGridLayout(self.centralWidget)
        vbox.addWidget(self.tabWidget, 0, 0, 1, 2)
        vbox.addWidget(self.buttonAdd, 2, 0)

    def add_work(self):
        gnkt_number_combo = self.tabWidget.currentWidget().gnkt_number_combo.currentText()
        well_data.gnkt_number = gnkt_number_combo

        lenght_gnkt_edit = self.tabWidget.currentWidget().lenght_gnkt_edit.text()

        well_data.pipe_fatigue = 0


        well_data.gnkt_length = lenght_gnkt_edit
        iznos_gnkt_edit = self.tabWidget.currentWidget().iznos_gnkt_edit.text().replace(',', '.')
        pipe_mileage_edit = self.tabWidget.currentWidget().pipe_mileage_edit.text()
        well_data.pipe_mileage = pipe_mileage_edit
        well_data.iznos = iznos_gnkt_edit
        current_bottom_edit = self.tabWidget.currentWidget().current_bottom_edit.text()
        GnktOsvWindow2.current_bottom_edit = int(float(current_bottom_edit))
        fluid_edit = self.tabWidget.currentWidget().fluid_edit.text()
        GnktOsvWindow2.fluid_edit = round(float(fluid_edit), 2)
        if well_data.work_plan == 'gnkt_after_grp':
            osvoenie_combo_need = self.tabWidget.currentWidget().osvoenie_combo.currentText()
            GnktOsvWindow2.osvoenie_combo_need = osvoenie_combo_need
        pvo_number = self.tabWidget.currentWidget().pvo_number_edit.text()
        well_data.pvo = pvo_number
        previous_well_combo = self.tabWidget.currentWidget().previous_well_combo.currentText()
        well_data.previous_well = previous_well_combo

        diametr_length = 38
        well_data.diametr_length = 38

        if '' in [gnkt_number_combo, lenght_gnkt_edit, iznos_gnkt_edit, fluid_edit, pvo_number]:
            mes = QMessageBox.warning(self, 'Некорректные данные', f'Не все данные заполнены')
            return

        work_list = self.schema_well(current_bottom_edit, fluid_edit, gnkt_number_combo,
                                     lenght_gnkt_edit, iznos_gnkt_edit, pvo_number, diametr_length, pipe_mileage_edit)

        well_data.pause = False
        self.close()
        return work_list

    def count_row_height(self, ws2, work_list, sheet_name):

        from openpyxl.utils.cell import range_boundaries, get_column_letter

        colWidth = [2.85546875, 14.42578125, 16.140625, 22.85546875, 17.140625, 14.42578125, 13.0, 13.0, 17.0,
                    14.42578125, 13.0, 21, 12.140625, None]

        text_width_dict = {35: (0, 100), 50: (101, 200), 70: (201, 300), 110: (301, 400), 120: (401, 500),
                           130: (501, 600), 150: (601, 700), 170: (701, 800), 190: (801, 900), 230: (901, 1500)}

        boundaries_dict = {}

        for ind, _range in enumerate(ws2.merged_cells.ranges):
            boundaries_dict[ind] = range_boundaries(str(_range))

        for key, value in boundaries_dict.items():
            ws2.unmerge_cells(start_column=value[0], start_row=value[1],
                              end_column=value[2], end_row=value[3])
        ins_ind = 1

        for i in range(1, len(work_list) + 1):  # Добавлением работ
            if str(work_list[i - 1][1]).isdigit() and i > 39:  # Нумерация
                work_list[i - 1][1] = str(ins_ind)
                ins_ind += 1
            for j in range(1, 13):
                cell = ws2.cell(row=i, column=j)

                if cell and str(cell) != str(work_list[i - 1][j - 1]):
                    if str(work_list[i - 1][j - 1]).replace('.', '').isdigit() and \
                            str(work_list[i - 1][j - 1]).count('.') != 2:
                        cell.value = str(work_list[i - 1][j - 1]).replace('.', ',')
                        # print(f'цифры {cell.value}')
                    else:
                        cell.value = work_list[i - 1][j - 1]

        # print(merged_cells_dict)
        if sheet_name != 'Ход работ':
            for key, value in boundaries_dict.items():
                # print(value)
                ws2.merge_cells(start_column=value[0], start_row=value[1],
                                end_column=value[2], end_row=value[3])

        elif sheet_name == 'Ход работ':
            for i, row_data in enumerate(work_list):
                # print(f'gghhg {work_list[i][2]}')
                for column, data in enumerate(row_data):
                    if column == 2:
                        if not data is None:
                            text = data
                            for key, value in text_width_dict.items():
                                if value[0] <= len(text) <= value[1]:
                                    ws2.row_dimensions[i + 1].height = int(key)
                    elif column == 1:
                        if not data is None:
                            text = data
                            # print(text)
                            for key, value in text_width_dict.items():
                                if value[0] <= len(text) <= value[1]:
                                    ws2.row_dimensions[i + 1].height = int(key)
                    if column != 0:
                        ws2.cell(row=i + 1, column=column + 1).border = well_data.thin_border
                    if column == 1 or column == 11:
                        ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(wrap_text=True,
                                                                                     horizontal='center',
                                                                                     vertical='center')
                        ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=False)
                    else:
                        ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                                     vertical='center')
                        ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=False)
                        if 'примечание' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'внимание' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'мероприятия' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'порядок работ' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'По доп.согласованию с Заказчиком' in str(
                            ws2.cell(row=i + 1, column=column + 1).value).lower():
                            # print('есть жирный')
                            ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=True)

                if len(work_list[i][1]) > 5:
                    ws2.merge_cells(start_column=2, start_row=i + 1, end_column=12, end_row=i + 1)
                    ws2.cell(row=i + 1, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
                    ws2.cell(row=i + 1, column=2).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1',
                                                                     fill_type='solid')
                    ws2.cell(row=i + 1, column=2).font = Font(name='Arial', size=13, bold=True)

                else:
                    ws2.merge_cells(start_column=3, start_row=i + 1, end_column=11, end_row=i + 1)
                    ws2.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')

            for col in range(13):
                ws2.column_dimensions[get_column_letter(col + 1)].width = colWidth[col]

            ws2.print_area = f'B1:L{self.table_widget.rowCount() + 45}'
            ws2.page_setup.fitToPage = True
            ws2.page_setup.fitToHeight = False
            ws2.page_setup.fitToWidth = True
            ws2.print_options.horizontalCentered = True
            # зададим размер листа
            ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
            # содержимое по ширине страницы
            ws2.sheet_properties.pageSetUpPr.fitToPage = True
            ws2.page_setup.fitToHeight = False

        for row_ind, row in enumerate(ws2.iter_rows(values_only=True)):
            for col, value in enumerate(row):
                if 'А.Р. Хасаншин' in str(value):
                    coordinate = f'{get_column_letter(col + 1)}{row_ind - 1}'
                    self.insert_image(ws2, '_internal/imageFiles/Хасаншин.png', coordinate)
                elif 'Д.Д. Шамигулов' in str(value):
                    coordinate = f'{get_column_letter(col + 1)}{row_ind - 2}'
                    self.insert_image(ws2, '_internal/imageFiles/Шамигулов.png', coordinate)
                elif 'Зуфаров' in str(value):
                    coordinate = f'{get_column_letter(col - 2)}{row_ind}'
                    self.insert_image(ws2, '_internal/imageFiles/Зуфаров.png', coordinate)
                elif 'М.К.Алиев' in str(value):
                    coordinate = f'{get_column_letter(col - 1)}{row_ind - 2}'
                    self.insert_image(ws2, '_internal/imageFiles/Алиев махир.png', coordinate)
                elif 'З.К. Алиев' in str(value):
                    coordinate = f'{get_column_letter(col - 1)}{row_ind - 2}'
                    self.insert_image(ws2, '_internal/imageFiles/Алиев Заур.png', coordinate)
                    break
        print(f'{sheet_name} - вставлена')

    def save_to_gnkt(self):

        sheets = ["Титульник", 'Схема', 'Ход работ']
        tables = [self.table_title, self.table_schema, self.table_widget]

        for i, sheet_name in enumerate(sheets):
            worksheet = self.wb_gnkt_frez[sheet_name]
            table = tables[i]

            work_list = []
            for row in range(table.rowCount()):
                row_lst = []
                # self.ins_ind_border += 1
                for column in range(table.columnCount()):

                    item = table.item(row, column)
                    if not item is None:

                        row_lst.append(item.text())
                        # print(item.text())
                    else:
                        row_lst.append("")
                work_list.append(row_lst)
            self.wb.count_row_height(self, worksheet, work_list, sheet_name)

        ws6 = self.wb.create_sheet(title="СХЕМЫ КНК_44,45")
        main.MyWindow.insert_image(self, ws6, '_internal/imageFiles/schema_well/СХЕМЫ КНК_44,45.png', 'A1', 550, 900)
        ws7 = self.wb.create_sheet(title="СХЕМЫ КНК_38,1")
        main.MyWindow.insert_image(self, ws7, '_internal/imageFiles/schema_well/СХЕМЫ КНК_38,1.png', 'A1', 550, 900)

        # path = 'workiii'
        if 'Зуфаров' in well_data.user:
            path = 'D:\Documents\Desktop\ГТМ'
        else:
            path = ""
        filenames = f"{well_data.well_number} {well_data.well_area} кат {well_data.cat_P_1} {self.work_plan}.xlsx"
        full_path = path + '/' + filenames
        # print(f'10 - {ws2.max_row}')
        # print(wb2.path)
        # print(f' кате {well_data.cat_P_1}')

        if well_data.bvo is True:
            ws5 = self.wb.create_sheet('Sheet1')
            ws5.title = "Схемы ПВО"
            ws5 = self.wb["Схемы ПВО"]
            self.wb.move_sheet(ws5, offset=-1)
            # schema_list = self.check_pvo_schema(ws5, ins_ind + 2)

        if self.wb:
            self.wb.remove(self.wb['Sheet'])

            main.MyWindow.saveFileDialog(self, self.wb, full_path)

            self.wb.close()
            print(f"Table data saved to Excel {full_path} {well_data.number_dp}")
        if self.wb:
            self.wb.close()

    def create_title_list(self, ws2):

        # print(f'цднг {well_data.cdng._value}')
        well_data.region = block_name.region(well_data.cdng._value)
        self.region = well_data.region

        title_list = [
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ЗАКАЗЧИК:', None, None, None, None, None, None, None, None, None, None],
            [None, 'ООО «Башнефть-Добыча»', None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ПЛАН РАБОТ НА СКВАЖИНЕ С ПОМОЩЬЮ УСТАНОВКИ С ГИБКОЙ ТРУБОЙ', None, None, None,
             None, None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, '№ скважины:', f'{well_data.well_number}', 'куст:', None, 'Месторождение:', None, None,
             well_data.well_oilfield, None, None],
            [None, None, 'инв. №:', well_data.inv_number, None, None, None, None, 'Площадь: ', well_data.well_area,
             None,
             1],
            [None, None, None, None, None, None, None, 'цех:',  f'{well_data.cdng}', None, None, None]]

        razdel = razdel_1(self, well_data.region)

        for row in razdel:  # Добавлением работ
            title_list.append(row)

        for row in [
            [None, None, None, None, None, "дата", datetime.now().strftime('%d.%m.%Y'), None, None, None, None]]:
            title_list.insert(-1, row)
        # print(title_list)
        index_insert = 11
        ws2.cell(row=1, column=2).alignment = Alignment(wrap_text=False, horizontal='left',
                                                        vertical='center')
        ws2.cell(row=1, column=4).alignment = Alignment(wrap_text=False, horizontal='left',
                                                        vertical='center')
        # ws2.column_dimensions[get_column_letter(1)].width = 15
        # ws2.column_dimensions[get_column_letter(2)].width = 20
        # ws2.column_dimensions[get_column_letter(3)].width = 20
        a = None
        for row in range(len(title_list)):  # Добавлением работ
            if row not in range(8, 13):
                ws2.row_dimensions[row].height = 35
            for col in range(1, 12):
                ws2.column_dimensions[get_column_letter(col)].width = 15
                cell = ws2.cell(row=row + index_insert, column=col)
                # print(f' Х {title_list[i ][col - 1]}')
                if title_list[row - 1][col - 1] != None:
                    ws2.cell(row=row + index_insert, column=col).value = str(title_list[row - 1][col - 1])
                ws2.cell(row=row + index_insert, column=col).font = Font(name='Arial', size=11, bold=False)
                ws2.cell(row=row + index_insert, column=col).alignment = Alignment(wrap_text=False, horizontal='left',
                                                                                   vertical='center')
                if 'ПЛАН РАБОТ' in str(title_list[row - 1][col - 1]):
                    ws2.merge_cells(start_row=row + index_insert, start_column=2, end_row=row + index_insert,
                                    end_column=11)
                    ws2.merge_cells(start_row=row - 4 + index_insert, start_column=2, end_row=row - 4 + index_insert,
                                    end_column=4)
                    ws2.cell(row=row + index_insert, column=col).font = Font(name='Arial', size=14, bold=True)
                    ws2.cell(row=row + index_insert, column=col).alignment = Alignment(wrap_text=False,
                                                                                       horizontal='center',
                                                                                       vertical='center')
                if 'СОГЛАСОВАНО:' in str(title_list[row - 1][col - 1]):
                    a = row + index_insert

            # for row in range(len(title_list)):  # Добавлением работ
            if a:
                if a > row:
                    # print(f'сссооссоссо {row + index_insert}')
                    ws2.merge_cells(start_row=row + index_insert, start_column=2, end_row=row + index_insert,
                                    end_column=6)
                    ws2.merge_cells(start_row=row + index_insert, start_column=8, end_row=row + index_insert,
                                    end_column=11)

        ws2.print_area = f'B1:K{44}'
        ws2.page_setup.fitToPage = True
        ws2.page_setup.fitToHeight = False
        ws2.page_setup.fitToWidth = True
        ws2.print_options.horizontalCentered = True
        # зададим размер листа
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4

    def schema_well(self, current_bottom_edit, fluid_edit, gnkt_number_combo,
                    gnkt_lenght, iznos_gnkt_edit, pvo_number, diametr_length, pipe_mileage_edit):
        self.gnkt = self.tabWidget.currentWidget()
        for plast_ind in well_data.plast_work:
            try:
                self.plast_work = plast_ind
                plast_work = self.plast_work

                self.pressuar = list(well_data.dict_perforation[plast_work]["давление"])[0]

                zamer = list(well_data.dict_perforation[plast_work]['замер'])[0]
                vertikal = min(map(float, list(well_data.dict_perforation[plast_work]["вертикаль"])))
                break
            except:
                pass

        koef_anomal = round(float(self.pressuar) * 101325 / (float(vertikal) * 9.81 * 1000), 1)
        nkt = int(list(well_data.dict_nkt.keys())[0])
        if nkt == 73:
            nkt_widht = 5.5
        elif nkt == 89:
            nkt_widht = 7.34
        elif nkt == 60:
            nkt_widht = 5

        lenght_nkt = sum(list(map(int, well_data.dict_nkt.values())))

        volume_vn_gnkt = round(30.2 ** 2 * 3.14 / (4 * 1000), 2)

        volume_gnkt = round(float(gnkt_lenght) * volume_vn_gnkt / 1000, 1)
        if well_data.column_additional:
            well_volume_ek = well_volume(self, well_data.head_column_additional._value)
        else:
            well_volume_ek = well_volume(self, well_data.current_bottom)
        well_volume_dp = well_volume(self, well_data.current_bottom) - well_volume_ek

        volume_pm_ek = round(
            3.14 * (well_data.column_diametr._value - 2 * well_data.column_wall_thickness._value) ** 2 / 4 / 1000, 2)
        volume_pm_dp = round(3.14 * (well_data.column_additional_diametr._value - 2 *
                                     well_data.column_additional_wall_thickness._value) ** 2 / 4 / 1000, 2)


        if well_data.column_additional:
            column_data_add_diam = well_data.column_additional_diametr._value
            column_data_add_wall_thickness = well_data.column_additional_wall_thickness._value

            column_data_add_vn_volume = round(
                well_data.column_additional_diametr._value - 2 * well_data.column_additional_wall_thickness._value, 1)
            column_add_head = well_data.head_column_additional._value
            column_add_shoe = well_data.shoe_column_additional._value

        else:
            column_data_add_diam = ''
            column_data_add_wall_thickness = ''

            column_data_add_vn_volume = ''
            column_add_head = ''
            column_add_shoe = ''
        if well_data.curator == 'ОР':
            expected_title = 'Ожидаемый приемистость скважины'
            Qoil = f'{well_data.expected_Q}м3/сут'
            Qwater = f'{well_data.expected_P}атм'
            proc_water = ''
        else:
            expected_title = 'Ожидаемый дебит скважины'
            Qoil = f'{well_data.Qoil}т/сут'
            Qwater = f'{well_data.Qwater}м3/сут'
            proc_water = f'{well_data.proc_water}%'

        wellhead_fittings = well_data.wellhead_fittings
        if well_data.work_plan == 'gnkt_after_grp':
            if 'грп' in str(well_data.wellhead_fittings).lower():
                wellhead_fittings = well_data.wellhead_fittings
            else:
                wellhead_fittings = f'АУШГН-{well_data.column_diametr._value}/' \
                                    f'АУГРП {well_data.column_diametr._value}*14'

        schema_well_list = [
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 'СХЕМА СКВАЖИНЫ', None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Данные о размерности труб', None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тип ПВО', None, None,
             f'4-х секционный превентор БП 80-70.00.00.000 (700атм) К2 № {pvo_number}', None, None, None, None, None,
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тип ФА', None, None,
             wellhead_fittings, None, None, None,
             None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тип КГ', None, None,
             well_data.column_head_m, None,
             None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, f'ЭК {well_data.column_diametr._value}мм', None,
             None,
             'Стол ротора', None, f'{well_data.stol_rotora._value}м',
             'Øнаруж мм', 'толщ, мм', 'Øвнут, мм', 'Интервал спуска, м', None, 'ВПЦ.\nДлина', 'Объем', None],
            [None, None, None, None, None, None, None, None, None, f'0-{well_data.shoe_column._value}м', None, None,
             'Ø канавки', None, 211, None, None,
             None, None, None, None, 'л/п.м.', 'м3'],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Шахтное направление', None, None,
             "", None, None, "", "", '', None, None],
            [None, None, None, None, None, None, None, None, None, f'НКТ {nkt}мм', None, None, 'Направление', None,
             None,
             f'{well_data.column_direction_diametr._value}', well_data.column_direction_wall_thickness._value,
             round(well_data.column_direction_diametr._value - 2 * well_data.column_direction_wall_thickness._value, 1),
             f'0-', well_data.column_direction_lenght._value, f'{well_data.level_cement_direction._value}', None, None],
            [None, None, None, None, None, None, None, None, None, f'0-{lenght_nkt}м', None, None, 'Кондуктор',
             None, None, well_data.column_conductor_diametr._value, well_data.column_conductor_wall_thickness._value,
             f'{round(well_data.column_conductor_diametr._value - 2 * well_data.column_conductor_wall_thickness._value)}',
             f'0-', well_data.column_conductor_lenght._value,
             well_data.level_cement_conductor._value, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Экспл. колонна', None, None,
             f'{well_data.column_diametr._value}', f'{well_data.column_wall_thickness._value}',
             f'{round(float(well_data.column_diametr._value - 2 * well_data.column_wall_thickness._value), 1)}',
             f'0-', well_data.shoe_column._value, well_data.level_cement_column._value, volume_pm_ek, well_volume_ek],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, "", ""],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, "", ""],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, "", ""],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тех.колонна', None, None,
             column_data_add_diam,
             column_data_add_wall_thickness, column_data_add_vn_volume, column_add_head, column_add_shoe, None,
             volume_pm_dp, well_volume_dp],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'колонна НКТ', None, None, nkt,
             nkt_widht, nkt - 2 * nkt_widht, 0, lenght_nkt, None, '=R19^2*3.14/4/1000', '=U19*V19/1000'],
            [None, None, None, None, None, None, None, None, None, None, None, None, f'{well_data.paker_do["do"]}',
             None, None, None, None,
             50, well_data.depth_fond_paker_do["do"], well_data.depth_fond_paker_do["do"] + 2, 2, None, None],
            [None, None, None, None, None, None, None, None, None, 'пакер', None, None, 'без патрубка', None, None,
             None, 0, 0, well_data.depth_fond_paker_do["do"], well_data.depth_fond_paker_do["do"], 0, 0, 0],
            [None, None, None, None, None, None, None, None, None, f'на гл {well_data.depth_fond_paker_do["do"]}м',
             None, None,
             'воронка', None, None, nkt, None,
             None, well_data.depth_fond_paker_do["do"], None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Данные о перфорации', None, None,
             None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, 'воронка', None, None, 'Пласт\nгоризонт', None,
             'Глубина пласта по вертикали', None, 'Интервал перфорации', None, None, None, 'вскрытия/\nотключения',
             'Рпл. атм', None],
            [None, None, None, None, None, None, None, None, None, f'на гл.{lenght_nkt}м', None, None, None, None, None,
             None, 'от', None, 'до', None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Данные о забое', None, None, None,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Тек. забой по ПЗ ', None, None,
             None, None, None, None, None, None, well_data.bottom, None],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             'необходимый текущий забой ', None, None, None, None, None, None, None, None, current_bottom_edit, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Искусственный забой  ', None,
             None, None, None, None, None, None, None, well_data.bottomhole_artificial._value, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Дополнительная информация', None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Категория скважины', None, None,
             None, None, well_data.category_pressuar, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Содержание H2S, мг/л', None, None,
             None,
             None, well_data.h2s_mg[0], None, None, None, None, None],
            [None, None, None, None, None, None, None, None,
             None, None, None, None, 'Газовый фактор', None, None, None,
             None, well_data.gaz_f_pr[0], None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Коэффициент аномальности', None,
             None, None, None, koef_anomal, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Плотность жидкость глушения',
             None, None, None, None, fluid_edit, None, 'в объеме', None, well_volume_ek, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, expected_title, None,
             None, None, None, Qoil, None, Qwater, None, proc_water, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Максимальный угол наклона', None,
             None, None, None, well_data.max_angle._value, None, 'на глубине', None, well_data.max_angle_H._value,
             None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Макс. набор кривизны более', None,
             None, None, None, 'вертикальная', None, 'на глубине', None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Дата начало / окончания бурения',
             None, None, None, None, self.date_dmy(well_data.date_drilling_run), None,
             self.date_dmy(well_data.date_drilling_cancel),
             None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Дата ввода в эксплуатацию', None,
             None, None, None, '', None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Р в межколонном пространстве',
             0, None, None, None, 0, None, ' ', None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Первоначальное Р опр-ки ЭК', None,
             None, None, None, well_data.first_pressure._value, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Результат предыдущей опрес-и ЭК',
             None, None, None, None, well_data.max_admissible_pressure._value, None, '', None, 'гермет.', None],
            [None, None, None, None, None, None, None, None, 'Тек.забой', None, None, None,
             'Макс.допустимое Р опр-ки ЭК', None, None, None, None, well_data.max_admissible_pressure._value,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, current_bottom_edit, None, None, None,
             'Макс. ожидаемое Р на устье ',
             None, None, None, None, well_data.max_expected_pressure._value, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, f'флот {gnkt_number_combo}',
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 'длина трубы', 'Øнаруж мм', 'толщ, мм',
             'Øвнут, мм', 'Объем', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, 'л/п.м.', None,
             'м3', None, '%', None, 'м', None, None],
            [None, None, None, None, None, None, None, None, None, None, None, gnkt_lenght,
             diametr_length, 3.68, '=M67-2*N67',
             '=ROUND(O67*O67*3.14/4/1000,2)', None, '=ROUND(L67*P67/1000, 1)', None, iznos_gnkt_edit, None,
             pipe_mileage_edit, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 'объем наземной линии, м3', None,
             'объем до границы перфорации, м3', None, 'Объем продавки, м3 ', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, 1, None, '=L67*P67/1000', None,
             '=L70+N70', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None],
        ]
        if well_data.paker_do['do'] == 0:
            schema_well_list[21] = [None, None, None, None, None, None, None, None, None,  None,
             None, None,
             'воронка', None, None, nkt, None,
             None, well_data.depth_fond_paker_do["do"], None, None, None, None]
            schema_well_list[20] = [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None]
            schema_well_list[19] = [None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                                    None,
                                    None, None, None, None, None, None, None, None]

        pvr_list = []
        for plast in sorted(well_data.plast_all, key = lambda x: self.get_start_depth(
                well_data.dict_perforation[x]['интервал'][0])):
            count_interval = 0
            for interval in well_data.dict_perforation[plast]['интервал']:
                count_interval += 1
                if well_data.dict_perforation[plast]['отключение']:
                    izol = 'Изолирован'
                else:
                    izol = 'рабочий'
                if well_data.paker_do['do'] != 0:
                    if well_data.dict_perforation[plast]['кровля'] < well_data.depth_fond_paker_do['do']:
                        izol = 'над пакером'
                try:
                    pressuar = well_data.dict_perforation[plast]['давление'][0]
                    zamer = well_data.dict_perforation[plast]['замер'][0]
                except:
                    pressuar = None
                    zamer = None
                pvr_list.append(
                    [None, None, None, None, None, None, None, None, None, None, None, None, plast, None, vertikal,
                     None, interval[0],
                     None, interval[1], None, izol, pressuar,
                     zamer, None], )
            well_data.dict_perforation[plast]['счет_объединение'] = count_interval


        for index, pvr in enumerate(pvr_list):
            schema_well_list[26 + index] = pvr

        well_data.current_bottom = round(float(current_bottom_edit), 1)
        return schema_well_list



    def date_dmy(self, date_str):
        if type(date_str) is str:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        else:
            date_obj = date_str
            # print(date_obj)
        # print(date_str)

        if isinstance(date_obj, datetime):
            return date_obj.strftime('%d.%m.%Y')
        else:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            # print(f' даь {date_obj}')
        return date_obj.strftime('%d.%m.%Y')

    # Функция для получения глубины начала интервала
    def get_start_depth(self, interval):
        return interval[0]
    def calc_fluid(self):

        fluid_list = []
        try:

            fluid_p = 0.83
            for plast in well_data.plast_work:
                if float(list(well_data.dict_perforation[plast]['рабочая жидкость'])[0]) > fluid_p:
                    fluid_p = list(well_data.dict_perforation[plast]['рабочая жидкость'])[0]
            fluid_list.append(fluid_p)

            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите расчетный удельный вес жидкости глушения в '
                                                           'конце жидкости',
                                                           max(fluid_list), 0.87, 2, 2)
        except:
            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите удельный вес рабочей жидкости',
                                                           0, 0.87, 2, 2)
        return fluid_work_insert


if __name__ == '__main__':
    app = QApplication([])
    window = GnktOsvWindow2()
    window.show()
    app.exec_()
