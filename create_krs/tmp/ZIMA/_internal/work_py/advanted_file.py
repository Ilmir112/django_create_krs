from PyQt5.QtWidgets import QInputDialog, QMessageBox

import plan
import well_data
from openpyxl.drawing.image import Image
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from PyQt5 import QtCore
def skm_interval(self, template):
   

    str_raid = []
    if well_data.paker_do["posle"] != 0:
        str_raid.append([float(well_data.depth_fond_paker_do["posle"]) - 20, float(well_data.depth_fond_paker_do["posle"]) + 20])

    if well_data.leakiness:
        for nek in list(well_data.dict_leakiness['НЭК']['интервал'].keys()):
            if int(float(nek.split('-')[1])) + 20 < well_data.current_bottom:
                str_raid.append([int(float(nek.split('-')[0])) - 90, int(float(nek.split('-')[1])) + 20])
            else:
                str_raid.append([int(float(nek.split('-')[0])) - 90,
                                 well_data.current_bottom - 2])
    if all([well_data.dict_perforation[plast]['отрайбировано'] is False for plast in well_data.plast_work]):
        str_raid.append([well_data.perforation_roof - 90, well_data.skm_depth])
        if well_data.leakiness:
            for nek in list(well_data.dict_leakiness['НЭК']['интервал'].keys()):
                # print(f' наруш {nek}')
                if float(nek.split('-')[1]) + 20 < well_data.current_bottom:
                    str_raid.append([int(float(nek.split('-')[0])) - 90, int(float(nek.split('-')[1])) + 20])
                else:
                    str_raid.append([int(float(nek.split('-')[0])) - 90,
                                     well_data.current_bottom - 2])

    elif all(
            [well_data.dict_perforation[plast]['отрайбировано'] is True for plast in well_data.plast_work]):
        str_raid = []

        perforating_intervals = []

        for plast in well_data.plast_all:
            for interval in well_data.dict_perforation[plast]['интервал']:
                perforating_intervals.append(list(interval))


        str_raid.extend(remove_overlapping_intervals(perforating_intervals))

    if well_data.dict_perforation_project is None and any(
            [plast in well_data.plast_all for plast in list(well_data.dict_perforation_project.keys())]) is False:
        if well_data.dict_perforation_project[plast]['интервал'][1] < well_data.current_bottom:
            str_raid.append([well_data.dict_perforation_project[plast]['интервал'][1] + 10,
                             well_data.dict_perforation_project[plast]['интервал'][1] + 50])
        else:
            str_raid.append([well_data.dict_perforation_project[plast]['интервал'][1] + 10,
                             well_data.current_bottom - 2])



    # print(f'скреперо {str_raid}')
    merged_segments = merge_overlapping_intervals(str_raid)
    # print(f'скреперо после {merged_segments}')
    merged_segments_new = []



    for interval in merged_segments:

        if template in ['ПСШ ЭК', 'ПСШ без хвоста', 'ПСШ открытый ствол']:
            if well_data.skm_depth >= interval[1] and interval[1] > interval[0]:
                merged_segments_new.append(interval)
            elif well_data.skm_depth <= interval[1] and well_data.skm_depth >= interval[0] and interval[1] > interval[0]:
                merged_segments_new.append([interval[0], well_data.skm_depth])


        elif template in ['ПСШ СКМ в доп колонне c хвостом', 'ПСШ СКМ в доп колонне без хвоста',
                          'ПСШ СКМ в доп колонне + открытый ствол'] and well_data.skm_depth >= interval[1]:

            if interval[0] > float(well_data.head_column_additional._value) and interval[1] > float(
                    well_data.head_column_additional._value) and well_data.skm_depth >= interval[1] \
                    and interval[1] > interval[0]:
                # print(f'1 {interval, merged_segments}')
                merged_segments_new.append(interval)

            elif interval[0] < float(well_data.head_column_additional._value) and interval[1] > float(
                    well_data.head_column_additional._value) and well_data.skm_depth <= interval[1] \
                    and well_data.skm_depth >= interval[0] and interval[1] > interval[0]:
                # print(f'2 {interval, merged_segments}')

                merged_segments_new.append((well_data.head_column_additional._value + 2, well_data.skm_depth))
            elif interval[0] < float(well_data.head_column_additional._value) and interval[1] > float(
                well_data.head_column_additional._value) and well_data.skm_depth >= interval[1] and interval[1] > interval[0]:

                merged_segments_new.append((well_data.head_column_additional._value + 2, interval[1]))
                # print(f'3 {interval, merged_segments}')

        elif template in ['ПСШ Доп колонна СКМ в основной колонне']:
            if interval[0] < float(well_data.head_column_additional._value) and interval[1] < float(
                    well_data.head_column_additional._value) and well_data.skm_depth >= interval[1] and interval[1] > interval[0]:
                merged_segments_new.append(interval)

            elif interval[0] < float(well_data.head_column_additional._value) and interval[1] > float(
                    well_data.head_column_additional._value) and well_data.skm_depth <= interval[1] \
                    and well_data.skm_depth >= interval[0] and interval[1] > interval[0]:
                # merged_segments.remove(interval)
                merged_segments_new.append((interval[0], well_data.skm_depth))


    return merged_segments_new


# Функция исключения из интервалов скреперования интервалов ПВР
def remove_overlapping_intervals(perforating_intervals, skm_interval = None):
   

    if skm_interval is None:
        # print(f' перфорация_ {perforating_intervals}')
        skipping_intervals = []
        if well_data.paker_do["posle"] != 0:
            if well_data.skm_depth > well_data.depth_fond_paker_do["posle"] + 20:
                skipping_intervals.append(
                    [float(well_data.depth_fond_paker_do["posle"]) - 20, float(well_data.depth_fond_paker_do["posle"]) + 20])
                # print(f'1 {skipping_intervals}')
            elif well_data.skm_depth > well_data.depth_fond_paker_do["posle"]:
                skipping_intervals.append(
                    [float(well_data.depth_fond_paker_do["posle"]) - 20, well_data.skm_depth])
                # print(f'2 {skipping_intervals}')

        if well_data.leakiness:
            for nek in list(well_data.dict_leakiness['НЭК']['интервал'].keys()):
                if int(float(nek.split('-')[1])) + 20 < well_data.skm_depth:
                    skipping_intervals.append([int(float(nek.split('-')[0])) - 90, int(float(nek.split('-')[1])) + 20])
                else:
                    skipping_intervals.append([int(float(nek.split('-')[0])) - 90,
                                     well_data.skm_depth])
        # print(f'глубина СКМ {well_data.skm_depth, skipping_intervals}')
        # print(perforating_intervals)
        for pvr in sorted(perforating_intervals, key=lambda x: x[0]):
            if pvr[1] <= well_data.skm_depth:
                # print(pvr, well_data.skm_depth)
                if pvr[1] + 40 < well_data.skm_depth and pvr[0] < well_data.skm_depth:
                    skipping_intervals.append([pvr[0] - 90, pvr[0] - 2])
                    if well_data.skm_depth >= pvr[1] + 40:
                        if [pvr[1] + 2, pvr[1] + 40] not in skipping_intervals:
                            skipping_intervals.append([pvr[1] + 2, pvr[1] + 40])
                    else:
                        if [pvr[1] + 2, well_data.skm_depth] not in skipping_intervals:
                            skipping_intervals.append([pvr[1] + 2, well_data.skm_depth])
                elif pvr[1] + 40 > well_data.skm_depth and pvr[0] < well_data.skm_depth:
                    if [pvr[0] - 90, pvr[0] - 2] not in skipping_intervals:
                        skipping_intervals.append([pvr[0] - 90, pvr[0] - 2])
                    if [pvr[1] + 1, well_data.skm_depth] not in skipping_intervals:
                        skipping_intervals.append([pvr[1] + 1, well_data.skm_depth])

        # print(f'СКМ на основе ПВР{sorted(skipping_intervals, key=lambda x: x[0])}')

        skipping_intervals = merge_overlapping_intervals(sorted(skipping_intervals, key=lambda x: x[0]))
        skipping_intervals_new = []
        for skm in sorted(skipping_intervals, key=lambda x: x[0]):
            kroly_skm = int(skm[0])
            pod_skm = int(skm[1])

            skm_range = list(range(kroly_skm, pod_skm + 1))
            for pvr in sorted(perforating_intervals, key=lambda x: x[0]):
                # print(int(pvr[0]) in skm_range, skm_range[0], int(pvr[0]))
                if int(pvr[0]) in skm_range and int(pvr[1]) in skm_range and skm_range[0]+1 <= int(pvr[0] - 1):
                    if skm_range[0]+1 < int(pvr[0]) - 2:
                        skipping_intervals_new.append((skm_range[0]+1, int(pvr[0] - 2)))
                        skm_range = skm_range[skm_range.index(int(pvr[1])):]



            skipping_intervals_new.append((skm_range[0], pod_skm))
    else:
        skipping_intervals_new = skm_interval

    print(f'после разделения {skipping_intervals_new}')
    return skipping_intervals_new


def raiding_interval(ryber_key):

    str_raid = []
    crt = 0
    for plast in well_data.plast_all:

        if well_data.dict_perforation[plast]['отрайбировано'] is False and \
                well_data.dict_perforation[plast]['кровля'] <= well_data.current_bottom:

            for interval in well_data.dict_perforation[plast]['интервал']:
                if interval[1] < well_data.current_bottom:
                    if well_data.column_additional is False or \
                            (well_data.column_additional and \
                             well_data.head_column_additional._value > well_data.current_bottom):

                        if float(interval[1]) + 20 <= well_data.current_bottom and \
                                well_data.shoe_column._value >= float(interval[1]) + 20:
                            crt = [float(interval[0]) - 20, float(interval[1]) + 20]
                        elif float(interval[1]) + 20 >= well_data.shoe_column._value and \
                                well_data.shoe_column._value > well_data.current_bottom:
                            crt = [float(interval[0]) - 20, well_data.shoe_column._value]
                        elif float(interval[1]) + 20 >= well_data.shoe_column._value and \
                                well_data.shoe_column._value <= well_data.current_bottom:
                            crt = [float(interval[0]) - 20,  well_data.current_bottom]
                        elif float(interval[1]) + 20 < well_data.shoe_column._value and \
                                float(interval[1] + 20) > well_data.current_bottom:
                            crt = [float(interval[0]) - 20, well_data.current_bottom]
                    else:
                        if float(interval[1]) + 20 <= well_data.current_bottom and \
                                well_data.shoe_column_additional._value >= float(interval[1]) + 20:
                            crt = [float(interval[0]) - 20, float(interval[1]) + 20]
                        elif float(interval[1]) + 20 >= well_data.shoe_column_additional._value:
                            crt = [float(interval[0]) - 20, well_data.shoe_column._value]
                        elif float(interval[1]) + 20 < well_data.shoe_column_additional._value and \
                                float(interval[1] + 20) > well_data.current_bottom:
                            crt = [float(interval[0]) - 20, well_data.current_bottom]
                if crt != 0 and crt not in str_raid:
                    str_raid.append(crt)

    if len(well_data.drilling_interval) != 0:
        # print(well_data.drilling_interval)
        for interval in well_data.drilling_interval:
            # print(interval)
            if float(interval[1]) + 20 <= well_data.current_bottom:
                str_raid.append((float(interval[0]) - 20, float(interval[1]) + 20))
            else:
                str_raid.append((float(interval[0]) - 20, well_data.current_bottom))
  # print(well_data.leakiness)
    if len(well_data.dict_leakiness) != 0:

        for nek in list(well_data.dict_leakiness['НЭК']['интервал'].keys()):
            if well_data.dict_leakiness['НЭК']['интервал'][nek]['отрайбировано'] is False:

                if float(nek.split('-')[1]) + 30 <= well_data.current_bottom and float(nek.split('-')[0]) + 30 <= well_data.current_bottom:
                    crt = (float(nek.split('-')[0]) - 30, float(nek.split('-')[1]) + 30)
                else:
                    crt = (float(nek.split('-')[0]) - 30, well_data.current_bottom)
                str_raid.append(crt)


    merged_segments = merge_overlapping_intervals(str_raid)
    merged_segments_new = []
    if ryber_key == 'райбер в ЭК' and well_data.column_additional:
        for str in merged_segments:
            if str[0] < well_data.head_column_additional._value and str[0] < str[1]:
                merged_segments_new.append(str)

    elif ryber_key == 'райбер в ДП' and well_data.column_additional:
        for str in merged_segments:
            if str[1] > well_data.head_column_additional._value and str[0] < str[1]:
                merged_segments_new.append(str)
    else:
        for str in merged_segments:
            if str[0] < str[1]:
                merged_segments_new = merged_segments
    return merged_segments_new

def change_True_raid(self, str_raid):
    if well_data.dict_leakiness:
        for nek in list(well_data.dict_leakiness['НЭК']['интервал'].keys()):
            for str in str_raid:
                # print(str[0], list(nek)[0], str[1])
                if str[0] <= float(nek.split('-')[0]) <= str[1]:
                    well_data.dict_leakiness['НЭК']['интервал'][nek]['отрайбировано'] = True
    if well_data.plast_all:
        for plast in well_data.plast_all:
            for interval in list((well_data.dict_perforation[plast]['интервал'])):
                for str in str_raid:
                    if str[0] <= list(interval)[0] <= str[1]:
                        well_data.dict_perforation[plast]['отрайбировано'] = True
def merge_overlapping_intervals(intervals):

    merged = []
    intervals = sorted(intervals, key=lambda x: x[0])
    for interval in intervals:
        if not merged or interval[0] > merged[-1][1]:
            merged.append(interval)
        else:
            if interval[0] < interval[1]:
                merged[-1] = (merged[-1][0], max(merged[-1][1], interval[1]))
    # print(f'интервалы СКМ {merged}')

    return merged
def raid(string):
    if len(string) == 1:
        return f'{int(float(string[0][0]))} - {int(float(string[0][1]))}'
    if len(string) == 0:
        return 'разбуренного цем моста'
    elif len(string) > 1:
        d = ''
        for i in list(string):
            d += f'{int(float(i[0]))} - {int(float(i[1]))}, '

    return d[:-2]



def definition_plast_work(self):
    # Определение работающих пластов
    plast_work = set()
    perforation_roof = well_data.current_bottom
    perforation_sole = 0

    for plast, value in well_data.dict_perforation.items():
        for interval in value['интервал']:

            if well_data.dict_perforation[plast]["отключение"] is False:
                plast_work.add(plast)

            if well_data.dict_perforation[plast]["отключение"] is False:
                roof = min(list(map(lambda x: x[0], list(well_data.dict_perforation[plast]['интервал']))))
                sole = max(list(map(lambda x: x[1], list(well_data.dict_perforation[plast]['интервал']))))
                well_data.dict_perforation[plast]["кровля"] = roof
                well_data.dict_perforation[plast]["подошва"] = sole
                if perforation_roof >= roof and well_data.current_bottom > roof:
                    perforation_roof = roof
                if perforation_sole < sole and well_data.current_bottom > sole:
                    perforation_sole = sole
            else:
                well_data.dict_perforation[plast]["кровля"] = \
                    min(list(map(lambda x: x[0], list(well_data.dict_perforation[plast]['интервал']))))
                well_data.dict_perforation[plast]["подошва"] = \
                    max(list(map(lambda x: x[1] if x[1] < well_data.current_bottom else 0,
                                 list(well_data.dict_perforation[plast]['интервал']))))

    well_data.perforation_roof = perforation_roof
    well_data.perforation_sole = perforation_sole
    # print(dict_perforation)
    well_data.plast_all = list(well_data.dict_perforation.keys())
    well_data.plast_work = list(plast_work)
def count_row_height(ws, ws2, work_list, merged_cells_dict, ind_ins):
    from openpyxl.utils.cell import range_boundaries, get_column_letter

    boundaries_dict = {}

    text_width_dict = {35: (0, 100), 50: (101, 200), 70: (201, 300), 95: (301, 400), 110: (401, 500),
                       130: (501, 600), 150: (601, 700), 170: (701, 800), 190: (801, 900), 210: (901, 1500)}
    for ind, _range in enumerate(ws.merged_cells.ranges):
        boundaries_dict[ind] = range_boundaries(str(_range))

    rowHeights1 = [ws.row_dimensions[i].height for i in range(ws.max_row)]
    colWidth = [ws.column_dimensions[get_column_letter(i + 1)].width for i in range(0, 13)] + [None]
    # print(colWidth)
    for i, row_data in enumerate(work_list):
        for column, data in enumerate(row_data):
            if column == 2:
                if not data is None:
                    text = data
                    for key, value in text_width_dict.items():
                        if value[0] <= len(text) <= value[1]:
                            ws2.row_dimensions[i + 1].height = int(key)

    head = plan.head_ind(0, ind_ins)
    # print(f'head - {head}')

    plan.copy_true_ws(ws, ws2, head)

    for i in range(1, len(work_list) + 1):  # Добавлением работ
        for j in range(1, 13):
            cell = ws2.cell(row=i, column=j)

            if cell and str(cell) != str(work_list[i - 1][j - 1]):
                # print(work_list[i - 1][j - 1])
                if str(work_list[i - 1][j - 1]).replace('.', '').isdigit() and \
                        str(work_list[i - 1][j - 1]).count('.') != 2:
                    try:
                        cell.value = str(work_list[i - 1][j - 1]).replace('.', ',')
                    except:
                        pass
                else:
                    try:
                        cell.value = work_list[i - 1][j - 1]
                    except:
                        pass

                if i >= ind_ins:
                    if j != 1:
                        cell.border = well_data.thin_border
                    if j == 11:
                        cell.font = Font(name='Arial', size=11, bold=False)
                    # if j == 12:
                    #     cell.value = work_list[i - 1][j - 1]
                    else:
                        cell.font = Font(name='Arial', size=13, bold=False)
                    ws2.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                    vertical='center')
                    ws2.cell(row=i, column=11).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                     vertical='center')
                    ws2.cell(row=i, column=12).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                     vertical='center')
                    ws2.cell(row=i, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                    vertical='center')
                    if 'примечание' in str(cell.value).lower() \
                            or 'заявку оформить за 16 часов' in str(cell.value).lower() \
                            or 'ЗАДАЧА 2.9.' in str(cell.value).upper() \
                            or 'ВСЕ ТЕХНОЛОГИЧЕСКИЕ ОПЕРАЦИИ' in str(cell.value).upper() \
                            or 'за 48 часов до спуска' in str(cell.value).upper():
                        # print('есть жирный')
                        ws2.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=True)
                    elif 'порядок работы' in str(cell.value).lower() or \
                            'Наименование работ' in str(cell.value):
                        ws2.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=True)
                        ws2.cell(row=i, column=j).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
    # print(merged_cells_dict)
    for row, col in merged_cells_dict.items():
        if len(col) != 2:
            # print(row)
            ws2.merge_cells(start_row=row + 1, start_column=3, end_row=row + 1, end_column=10)
    for key, value in boundaries_dict.items():
        # print(value)
        ws2.merge_cells(start_column=value[0], start_row=value[1],
                        end_column=value[2], end_row=value[3])

    # вставка сохраненных изображение по координатам ячеек
    if well_data.image_list:
        # print(f' схемы {well_data.image_list}')
        for img in well_data.image_list:
            logo = Image(img[0])
            logo.width, logo.height = img[2][0] * 0.48, img[2][1] * 0.72
            ws2.add_image(logo, img[1])

    # print(f'высота строк работ {ins_ind}')
    # print(f'высота строк работ {len(rowHeights1)}')
    for index_row, row in enumerate(ws2.iter_rows()):  # Копирование высоты строки
        if all([col is None for col in row]):
            ws2.row_dimensions[index_row].hidden = True
        try:
            if index_row < ind_ins:
                ws2.row_dimensions[index_row].height = rowHeights1[index_row]
        except:
            pass
        if index_row == 2:
            for col_ind, col in enumerate(row):
                if col_ind <= 12:
                    ws2.column_dimensions[get_column_letter(col_ind + 1)].width = colWidth[col_ind]
    ws2.column_dimensions[get_column_letter(11)].width = 20
    ws2.column_dimensions[get_column_letter(12)].width = 20

    ws2.column_dimensions[get_column_letter(6)].width = 18

    return 'Высота изменена'