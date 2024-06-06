import sqlite3
from datetime import datetime

import psycopg2
from PyQt5.QtWidgets import QInputDialog, QMainWindow, QTabWidget, QWidget, QTableWidget, QApplication, QLabel, \
    QLineEdit, QGridLayout, QComboBox, QPushButton,QMessageBox
from main import MyWindow

import well_data
from gnkt_data.gnkt_data import gnkt_1, gnkt_2, gnkt_dict, read_database_gnkt, insert_data_base_gnkt
from gnkt_opz import GnktOpz
from krs import TabPageGno, GnoWindow
from perforation_correct import PerforationCorrect

import block_name
import main
import plan
from block_name import razdel_1
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.reader.excel import load_workbook
from openpyxl.workbook import Workbook
from open_pz import CreatePZ
from work_py.alone_oreration import well_volume

from work_py.gnkt_grp_work import GnktOsvWindow2


class TabPageDp(QWidget):
    def __init__(self, work_plan):
        super().__init__()
        self.work_plan = work_plan

        self.gnkt_number_label = QLabel('Номер флота ГНКТ')
        self.gnkt_number_combo = QComboBox(self)
        self.gnkt_number_combo.addItems(gnkt_dict["Ойл-сервис"])

        self.lenght_gnkt_label = QLabel('длина ГНКТ')
        self.lenght_gnkt_edit = QLineEdit(self)


        self.iznos_gnkt_label = QLabel('Износ трубы')
        self.iznos_gnkt_edit = QLineEdit(self)

        self.pipe_mileage_label = QLabel('Пробег трубы')
        self.pipe_mileage_edit = QLineEdit(self)


        self.pvo_number_label = QLabel('Номер ПВО')
        self.pvo_number_edit = QLineEdit(self)

        self.previous_well_label = QLabel('Предыдущая скважина')
        self.previous_well_combo = QComboBox(self)
        well_previus_list = read_database_gnkt(well_data.contractor, self.gnkt_number_combo.currentText())
        self.previous_well_combo.addItems(well_previus_list)

        self.current_bottom_label = QLabel('необходимый текущий забой')
        self.current_bottom_edit = QLineEdit(self)
        self.current_bottom_edit.setText(f'{well_data.current_bottom}')

        self.fluid_label = QLabel("уд.вес жидкости глушения", self)
        self.fluid_edit = QLineEdit(self)
        if well_data.fluid_work == '':
            self.fluid_edit.setText(f'{TabPageGno.calc_fluid(self.work_plan, well_data.current_bottom)}')
        else:
            self.fluid_edit.setText(f'{well_data.fluid_work}')

        self.osvoenie_label = QLabel('Необходимость освоения')
        self.osvoenie_combo = QComboBox(self)
        self.osvoenie_combo.addItems('Да', 'Нет')

        self.grid = QGridLayout(self)
        self.grid.addWidget(self.gnkt_number_label, 2, 2, 3, 1)
        self.grid.addWidget(self.gnkt_number_combo, 3, 2, 3, 1)
        self.grid.addWidget(self.lenght_gnkt_label, 2, 3)
        self.grid.addWidget(self.lenght_gnkt_edit, 3, 3)
        self.grid.addWidget(self.iznos_gnkt_label, 2, 4)
        self.grid.addWidget(self.iznos_gnkt_edit, 3, 4)
        self.grid.addWidget(self.pvo_number_label, 2, 5)
        self.grid.addWidget(self.pvo_number_edit, 3, 5)
        self.grid.addWidget(self.previous_well_label, 2, 6)
        self.grid.addWidget(self.previous_well_combo, 3, 6)

        self.grid.addWidget(self.current_bottom_label, 4, 2)
        self.grid.addWidget(self.current_bottom_edit, 5, 2)
        self.grid.addWidget(self.fluid_label, 4, 3)
        self.grid.addWidget(self.fluid_edit, 5, 3)
        self.grid.addWidget(self.osvoenie_label, 4, 4)
        self.grid.addWidget(self.osvoenie_combo, 5, 4)
        self.gnkt_number_combo.textChanged.connect(self.update_number_gnkt)

    def update_number_gnkt(self, number_gnkt):

        try:

            conn = psycopg2.connect(**well_data.postgres_conn_gnkt)
            cursor = conn.cursor()
            cursor.execute(f"SELECT * FROM КГМ WHERE today (%s), ?", (number_gnkt, self.previous_well_edit.text()))

            result_gnkt = cursor.fetchone()

            self.lenght_gnkt_edit.setText(f'{result_gnkt[3]}')
            self.iznos_gnkt_edit.setText(f'{round(result_gnkt[5], 1)}')
            self.pipe_mileage_edit.setText(f'{result_gnkt[6]}')
            self.pvo_number_edit.setText(f'{result_gnkt[8]}')
        except psycopg2.Error as e:
            print('Ошибка подключения')

        finally:
            # Закройте курсор и соединение
            if cursor:
                cursor.close()
            if conn:
                conn.close()


class TabWidget(QTabWidget):
    def __init__(self, work_plan):
        super().__init__()
        self.addTab(TabPageDp(work_plan), 'ГНКТ освоение после ГРП')


class GnktOsvWindow(QMainWindow):
    wb = None

    def __init__(self, sheet, table_title, table_schema, table_widget, work_plan):
        super(QMainWindow, self).__init__()
        from work_py.gnkt_frez import Work_with_gnkt

        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        # self.tabWidget = tabWidget
        self.table_widget = table_widget
        self.table_title = table_title
        self.table_schema = table_schema

        self.dict_perforation = well_data.dict_perforation
        self.wb = load_workbook('property_excel/tepmpale_gnkt_osv_grp.xlsx')
        GnktOsvWindow.wb = self.wb
        self.ws_schema = self.wb.active

        self.work_plan = work_plan

        self.work_window = None
        self.wb.sheetnames.insert(0, "Титульник")
        self.ws_title = self.wb.create_sheet("Титульник", 0)
        self.ws_work = self.wb.create_sheet(title="Ход работ")

        Work_with_gnkt.create_title_list(self, self.ws_title)

        head = plan.head_ind(well_data.cat_well_min._value, well_data.cat_well_max._value)

        plan.copy_true_ws(sheet, self.ws_title, head)

        if self.work_window is None:
            self.work_window = GnktOsvWindow2(self)
            self.work_window.setWindowTitle("Данные по ГНКТ")
            self.work_window.setGeometry(200, 400, 100, 400)
            self.work_window.show()
            main.MyWindow.pause_app()
            well_data.pause = True

            self.work_schema = self.work_window.add_work()

            # print(self.work_schema)
            self.work_window = None

        else:
            self.work_window.close()
            self.work_window = None
        # schema_well = self.schema_well(self.ws_schema, )
        self.copy_pvr(self.ws_schema, self.work_schema)
        # self.wb.save(f"{well_data.well_number._value} {well_data.well_area._value} ГНКТ освоение.xlsx")
        # print('файл сохранен')
        for row_number, row in enumerate(self.ws_title.iter_rows(values_only=True)):
            for col_number, cell in enumerate(row):
                if 'по H2S' in str(self.ws_title.cell(row=row_number+1, column=col_number+1).value) and 'по H2S' not in str(
                        self.ws_title.cell(row=row_number + 2, column=col_number).value):
                    self.ws_title.merge_cells(start_column=col_number + 1, start_row=row_number + 1,
                                    end_column=col_number + 1, end_row=row_number + 2)
                    self.ws_title.merge_cells(start_column=col_number + 3, start_row=row_number + 1,
                                    end_column=col_number + 3, end_row=row_number + 2)

        main.MyWindow.copy_pz(self, self.ws_title, table_title, self.work_plan, 13, 1)
        main.MyWindow.copy_pz(self, self.ws_schema, table_schema, self.work_plan, 23, 2)

        main.MyWindow.copy_pz(self, self.ws_work, table_widget, self.work_plan, 12, 3)
        if self.work_plan == 'gnkt_opz':
            if self.work_window is None:
                self.work_window = GnktOpz(table_widget, well_data.gnkt_number, GnktOsvWindow2.fluid_edit)
                self.work_window.show()
                well_data.pause = True
                MyWindow.pause_app()

                work_well = self.work_window.add_work()
                well_data.pause = True
                self.work_window = None
            else:
                self.work_window.close()  # Close window.
                self.work_window = None

        elif self.work_plan == 'gnkt_after_grp':
            work_well = self.gnkt_work(
                GnktOsvWindow2.fluid_edit, well_data.pvo, GnktOsvWindow2.current_bottom_edit,
             GnktOsvWindow2.osvoenie_combo_need)
        if work_well:
            main.MyWindow.populate_row(self, 0, work_well, table_widget)
            CreatePZ.add_itog(self, self.ws_work, self.table_widget.rowCount() + 1, self.work_plan)



    def gnkt_work(self, fluid_work_insert, pvo_number_edit, current_bottom_edit, osvoenie_combo_need):
        from cdng import events_gnvp_frez

        fluid_work, well_data.fluid_work_short = GnoWindow.calc_work_fluid(self, fluid_work_insert)

        distance, _ = QInputDialog.getInt(None, 'Расстояние НПТЖ', 'Введите Расстояние до ПНТЖ')

        block_gnvp_list = events_gnvp_frez(self, distance, float(fluid_work_insert))
        gnkt_opz = [
            [None, 'Порядок работы', None,  None, None, None, None, None, None, None, None, None],
            [None,  None, 'Наименование работ',  None, None, None, None, None, None, None,
             'Ответственный', 'Нормы'],
            [None,
             None, 'Ознакомить бригаду с планом работ и режимными параметрами '
                   'дизайна по промывке и СПО. Провести инструктаж по промышленной безопасности',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, None, 'Принять скважину у Заказчика по акту '
                         '(состояние ф/арматуры и кустовой площадки.).',
             None, None, None, None, None, None, None, 'Мастер ГНКТ'],
            [None, None, 'Расставить оборудование и технику согласно '
                         '«Типовой схемы расстановки оборудования и спецтехники при '
                         'проведении капитального ремонта скважин с использованием установки «Койлтюбинг».',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', None],
            [None, 3, 'Произвести завоз технологической жидкости в V не менее 10м3 '
                      f'плотностью {fluid_work}. Солевой раствор солевой раствор в объеме ГНКТ',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', None],
            [None, 4, 'При наличии, согласно плана заказа Н2S, добавить в завезенную промывочную '
                      'жидкость нейтрализатора сероводорода "Реком-102" в концентрации 0,5л на 10м³',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', None],

            [None, 6, 'Внимание: при проведении работ по ОПЗ с кислотными составами, весь состав вахты '
                      'обязан применять СИЗ (Инструкция П1-01.03 И-0128 ЮЛ-305 ООО"Башнефть-Добыча")',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады ГНКТ', None],
            [None, 7, 'Примечание: на месте проведения работ по ОПЗ кислотами и их смесями должен быть '
                      'аварийный запас спецодежды, спецобуви и других средств индивидуальной защиты, запас '
                      'чистой пресной воды и средств нейтрализации кислоты (мел, известь, хлорамин).',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', None],
            [None, 'МОНТАЖ И ОПРЕССОВКА', None,
             None, None, None, None, None, None, None,
             None, None, ],
            [None, 9, 'Собрать Компоновку Низа Колонны-1 далее КНК-1: '
                      '(насадка-промывочная D-38мм + сдвоенный обратный клапан.)',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, бур-к КРС, машинист подъёмника, пред. УСРСиСТ', 0.5],
            [None, 10, f'Произвести монтаж 4-х секционного превентора БП 80-70.00.00.000 (700атм) {pvo_number_edit} '
                       f'и инжектора н'
                       f'а устье скважины согласно «Схемы обвязки № 5 устья противовыбросовым оборудованием при '
                       f'производстве работ по промывке скважины с установкой «ГНКТ» утвержденная главным '
                       f'инженером от 14.10.2021г. Произвести обвязку установки ГНКТ, насосно-компрессорного '
                       f'агрегата, желобной циркуляционной системы.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, представ.БВО (вызов по телефонограмме при необходимости)', 2],
            [None, 11, f'Внимание: Все требования ПБ и ОТ должны быть доведены до сведения работников, '
                       f'персонал должен быть проинформирован о начале проведения опрессовок. Все опрессовки '
                       f'производить согласно инструкции опрессовки ПВО и инструкции опрессовки нагнетательной '
                       f'и выкидной линии перед производством работ на скважине с Колтюбинговыми установками.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 12, f'При отрицательной температуре окружающей среды, нагреть до t - 50C и прокачать '
                       f'по ГНКТ солевой раствор в объеме ГНКТ для предотвращения замерзания раствора внутри '
                       f'г/трубы (получения ледяной пробки).',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 2],
            [None, 13, f'При закрытой центральной задвижке фондовой арматуры опрессовать ГНКТ и все '
                       f'нагнетательные линии на {round(well_data.max_admissible_pressure._value * 1.5, 1)}атм. '
                       f'Опрессовать ПВО, обратные клапана и выкидную линию от '
                       f'устья скважины до желобной ёмкости (надёжно закрепить, оборудовать дроссельными задвижками) '
                       f'опрессовать на {well_data.max_admissible_pressure._value}атм с выдержкой 30мин. '
                       f'Опрессовку ПВО зафиксировать в вахтовом журнале. Установить на малом и большом затрубе '
                       f'технологический манометр. Провести УТЗ и инструктаж. Опрессовку проводить в присутствии мастера, '
                       f'бурильщика, машиниста подъемника и представителя супервайзерской службы. Получить разрешение на '
                       f'проведение работ.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады, представит. Заказчика', 2],
            [None, 14, f'Ограничения веса и скоростей при СПО',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады, представитель Заказчика', None, ],
            [None, f'СПУСК ГНКТ В СКВАЖИНУ', None,
             None, None, None, None, None, None, None,
             None, None],

            [None, 22, f'Открыв скважину и записав число оборотов задвижки – зафиксировать дату и время. '
                       f'Спустить КНК-1 в скважину с ПЕРИОДИЧЕСКОЙ прокачкой рабочей жидкостью и проверкой веса на '
                       f'подъём до получения посадки с целью определения глубины "головы" проппанта.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, ', 4.5],
            [None, 23,
             f'ВНИМАНИЕ: При получении посадки до гл. {well_data.depth_fond_paker_do["do"]}м и наличии разгрузки на промывочный '
             f'инструмент более 500кг  (уведомить Заказчика – '
             f'составить АКТ на посадку). Приподнять КНК-1 на 20м выше глубины посадки. '
             f'Произвести вывод НКА на рабочий режим, восстановить '
             f'устойчивую циркуляцию промывочной жидкости (расход 180-190л/мин), '
             f'произвести промывку лифта НКТ до гл.{well_data.depth_fond_paker_do["do"]}м с постоянным контролем промывочной жидкости '
             f'в обратной ёмкости на наличие мех. примесей. Скорость спуска при промывке НКТ от проппанта до '
             f'гл.{well_data.depth_fond_paker_do["do"]}м не более 5м/мин. Контрольная проверка веса при вымыве проппанта - '
             f'через каждые 100м промывки на высоту не менее 5-10м со скоростью подъёма ГНКТ при проверке веса не '
             f'более 5м/мин. Внимание: после промывки НКТ до гл.2087м приподнять ГТ на 20м выше '
             f'пакера и '
             f'промыться до выхода чистой тех. жидкости (без мех.примесей) и '
             f'только после этого продолжить промывку ниже пакера', 2.5],
            [None,  f'НОРМАЛИЗАЦИЯ ЗАБОЯ СКВАЖИНЫ', None,
             None, None, None, None, None, None, None,
             'Мастер ГНКТ,', None],
            [None, 25,
             f'После отбивки текущего забоя (головы проппанта) произвести подъем КНК -1 до гл.{well_data.depth_fond_paker_do["do"] - 20}м.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 12.5],
            [None, 26,
             f'Произвести запуск и вывести Азотный комплекс и НКА на рабочий режим. Получить стабильную круговую циркуляцию '
             f'азотированной смеси, - циркуляция по жидкости не менее 120л/мин; 10м3/мин по азоту в течении 30мин '
             f'с контролем на мех примеси в обратной ёмкости.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', 1],
            [None, 27,
             'ВНИМАНИЕ. В процессе промывки скважины, параметры азотированной промывочной смеси могут изменяться '
             '(от 80 до 200л/мин по жидкости и от 8 до 20м3/мин по азоту) в зависимости от '
             'качества выноса посторонних частиц с забоя, данный процесс находится под постоянным контролем у ст.мастера ГНКТ.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 0.5],
            [None, 28, 'Произвести доспуск КНК-1 с циркуляцией на азотированной смеси со скоростью 2 м/мин, '
                       'с проверкой веса на подъём со скоростью не более 3 м/мин через каждые 30м интервала до гл. '
                       'посадки. Добиться устойчивой циркуляции азотированной смеси.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', None],
            [None, 29, 'ВНИМАНИЕ: При промывке, во избежании риска прихвата ГНКТ, производить сопровождение '
                       'вымытой пачки со скоростью 2-3м/мин. Производить промывку до выхода чистой тех. '
                       'жидкости и только после этого продолжать промывку.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 28, 'Произвести допуск компоновки с промывкой со скоростью 2 м/мин, с проверкой веса '
                       f'на подъём со скоростью не более 3 м/мин через каждые 10-20м интервала промывки '
                       f'до глубины {current_bottom_edit}м. В случае отсутствия проходки, '
                       f'согласовать максимально достигнутый забой с Заказчиком.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ, состав бригады', 0.7],
            [None, 29,
             f'При достижении глубины {current_bottom_edit}м произвести прокачку гелевой пачки в V=2-3м3, '
             f'промыть скважину в течении 120мин до выхода чистой, без посторонних примесей, '
             f'промывочной жидкости.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 2.5],
            [None, 29, f'ПОДТВЕРЖДЕНИЕ НОРМАЛИЗОВАННОГО ЗАБОЯ',
             None, None, None, None, None, None, None,
             None, None],
            [None, 29,
             f'Приподнять КНК-1 на ГНКТ не прекращая циркуляции до гл.{well_data.depth_fond_paker_do["do"] - 20}м. '
             f'Убедиться в отсутствии мех. примесей в промывочной жидкости, остановить подачу '
             f'жидкости НКА и ПАУ.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 0.5],
            [None, 29, f'Произвести тех.отстой скважины для оседания твёрдых частиц в течении 2х часов.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 2],
            [None, 29, f'По истечении 2х часов, произвести допуск КНК-1 на г/трубе в скважину '
                       f'«без циркуляции» до гл. {current_bottom_edit}м. '
                       f'Забой должен соответствовать ранее нормализованному. Составить АКТ на забой совмесно '
                       f'с представителем Заказчика. При отсутствии нормализованного забоя на гл.{current_bottom_edit}м, '
                       f'по согласованию с Заказчиком, провести работы по нормализации забоя.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 0.5],

            [None, 29, f'При наличии забоя на гл.{current_bottom_edit}м, '
                       f'по согласованию с Заказчиком: ',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Произвести замер избыточного давления в течении 2ч при условии заполнения '
                       f'ствола ствола жидкостью уд.весом {fluid_work}г/см3. Произвести перерасчет '
                       f'забойного давления, Согласовать с заказчиком глушение скважин и необходимый '
                       f'удельный вес жидкости глушения, допустить КНК до {current_bottom_edit}м ',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 2.5],
            [None, f'ГЛУШЕНИЕ СКВАЖИНЫ', None,
             None, None, None, None, None, None, None,
             None, None],
            [None, 29, f'Произвести перевод на тех жидкость расчетного удельного веса '
                       f'в объеме {self.calc_volume_jumping()}м3 (объем хвостовика + объем НКТ + 20 % запаса) '
                       f'с ПРОТЯЖКОЙ ГНКТ СНИЗУ ВВЕРХ с выходом по малому затрубу. Т'
                       f'ех отстой 2ч. В случае отрицательного результата по глушению скважины '
                       f'произвести перерасчет ЖГС и повторить операцию.'
                       f'ПРИ ПРОВЕДЕНИИ ВЕСТИ ГЛУШЕНИЯ КОНТРОЛЬ ЗА БАЛАНСОМ МЕЖДУ ОБЪЕМОМ'
                       f' ЗАКАЧИВАЕМОЙ И ВЫХОДЯЩЕЙ ЖИДКОСТЬЮ',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 4.5],
            [None, 29, f'Закрыв скважину и записав число оборотов задвижки – зафиксировать дату и время.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None,  f'ОТБИВКА ЗАБОЯ ПО ГК и ЛМ', None,
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'По согласованию с Заказчиком, проведение ГИС - отбивка забоя по ГК, МЛМ.'
                       f'Заявку на промыслово-геофизические исследования для подтверждения забоя подавать за 24 часа',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 4],
            [None,  f'ДЕМОНТАЖ И ОСВОБОЖДЕНИЕ ТЕРРИТОРИИ', None,
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'После закрытия задвижки - отдуть г/трубу азотом.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Произвести демонтаж превентора и инжектора, установки ГНКТ. '
                       f'Очистить желобные ёмкости от проппанта в мешки – приготовить к вывозу. Составить '
                       f'Акт на количество вымытого проппанта. Произвести демонтаж рабочих линий, рабочей площадки.'
                       f'Внимание: произвести вывоз мешков с вымытым проппантом и отработанной '
                       f'технологической жидкости на пункты утилизации, соглаованный с Заказчиком.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 2.5],
            [None, 29, f'Сдать скважину представителю Заказчика Составить АКТ.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Контроль выхода малого затруба',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Во время промывки - выход малого затруба постоянно должен находиться под '
                       f'контролем. На желобной ёмкости постоянно осуществляется наблюдение за '
                       f'наличием проппанта и мех. примесей на выходной линии. Перед началом '
                       f'промывки – необходимо отрегулировать штуцерный монифольд так, как это '
                       f'необходимо – уровень промывочной жидкости в циркуляционной ёмкости не должен уменьшаться. '
                       f'Уровень жидкости должен находиться под постоянным наблюдением, '
                       f'чтобы избежать потери жидкости в пласт. Во время промывки уровень жидкости должен немного '
                       f'увеличиваться или оставаться неизменным. Промывку от проппанта производить со '
                       f'скоростью не более 2м/мин, с проверкой веса на подъём через каждые 30м промывки, '
                       f'не превышая скорость подъёма г/трубы 3м/мин.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Действия при приватах ГНКТ.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29,
             f'ВНИМАНИЕ: При наличии посадок КНК - спуск производить с остановками для промежуточных промывок. '
             f'В случае прихвата ГНКТ в скважине - проинформировсть ответственного представителя Заказчика и '
             f'руководство ГНКТ ООО "Ойл-сервис". Дальнейшие действия производить в присутствии представителя '
             f'Заказчика с составлением АКТа согласно "Плана-Схемы действий при прихватах ГНКТ" '
             f'ТЕХНОЛОГИЧЕСКОЙ ИНСТРУКЦИИ ОАО «Башнефть добыча»',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Использование хим. реагентов в процессе работ',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29,
             f'При отрицательных температурах окружающего воздуха перед открытием задвижки прокачать по '
                   f'ГНКТ (горячий солевой раствор в объеме ГНКТ) для предотвращения замерзания раствора внутри ГНКТ '
                   f'(получения ледяной пробки). ВНИМАНИЕ: Во время промывки возможен резкий вынос большого '
                   f'объёма проппанта из пласта, что может привести к потере циркуляции и последующему прихвату ГНКТ. '
                   f'Данную ситуацию можно проследить; при этом вес ГНКТ резко понизиться, а циркуляционное давление начнёт повышаться. '
                   f'а) Необходимо спуск г/трубы приостановить – произвести промывку с добавлением понизителя трения (0.4-1 л/м3) '
                   f'до стабилизации рабочего давления, но не менее 4 пачек по 4 м3 каждая. '
                   f'Продолжить промывку. б) В случае поглощения промывочной жидкости в процессе промывки интервала перфорации - '
                   f'производить прокачку по г/трубе (вязких пачек V-от 2 м3 до 4 м3), а на забое '
                   f'(пачку V- 4 м)3 и сопровождение вязкой пачки в пакер через каждые 2м промывки интервала '
                   f'до полного выноса проппанта на желобную ёмкость.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29,
             f'После закрытия задвижки - приготовить и прокачать по г/трубе по циркуляции на желобную ёмкость '
             f'пачку – (ингибитора коррозии в объёме 40л.) - для недопустимости коррозийных отложений в г/трубе. '
             f'Предположительный расход хим.реагентов на скважину: \n'
             f'1) Понизитель трения Лубритал - 30л (концентрация 1л/м3) \n'
             f'2) Загуститель ВГ-4 - 20л (для загеливания тех жидкости и прокачки вязких пачек концентрация 5кг/м3)\n'
             f'3) Лимонная кислота - 300кг (для разложения геля 1 кг/м3, для кислотных ванн 100кг/м3) \n'
             f'4) Ингибитор коррозии - 40л (концентрация 1л/м3)',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'При наличии корки проппанта',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'При отсутствии проходки в процессе промывки и наличии корки спёкшегося проппанта, '
                       f'по согласованию с Заказчиком, произвести подъём КНК-1 на ГНКТ из скважины, заменить '
                       f'перо на насадку для резки проппанта. Спустить насадку на г/трубе до глубины посадки '
                       f'промывочного пера и произвести резку корки проппанта насадкой, но не более 10м от '
                       f'места посадки. После чего, поднять насадку на г/трубе из скважины, произвести замену '
                       f'насадки на КНК-1. Произвести спуск промывочного пера на г/трубе до текущего забоя. '
                       f'Продолжить промывку далее.- В случае неполного выноса проппанта при промывке произвести '
                       f'подъём компоновки 20м выше пакера и произвести промывку до полного выноса проппанта.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Ограничения веса и скоростей при СПО:',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29,
             f'Максимальный расчётный вес ГНКТ при подъёме с забоя – {round(current_bottom_edit * 2.2 * 1.1, 0)}кг; '
             f'при спуске – {round(current_bottom_edit * 2.2 * 0.9, 0)}кг.; в неподвижном состоянии - '
             f'{round(current_bottom_edit * 2.2, 0)}кг.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Скорость спуска по интервалам:\n'
                       f'в устьевом оборудовании не более 0.5м/мин;\n'
                       f'в интервале 2 -{well_data.depth_fond_paker_do["do"] - 20}м не более 10-15м/мин '
                       f'(первичный-последующий спуск);\n'
                       f'в интервале {well_data.depth_fond_paker_do["do"] - 20}-{well_data.perforation_roof - 10}м не '
                       f'более 2м/мин;\n'
                       f'в интервале {well_data.perforation_roof - 10}-{well_data.perforation_sole + 10}м не более 10 м/мин;\n'
                       f'в интервале {well_data.perforation_sole + 10}-{current_bottom_edit}м не более 2-5 м/мин;',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Скорость подъёма по интервалам:\n'
                       f'в интервале забой-{well_data.perforation_sole + 10}м не более 10 м/мин;\n'
                       f'в интервале {well_data.perforation_sole + 10}-{well_data.perforation_roof - 10} не более 2 м/мин;\n'
                       f'в интервале {well_data.depth_fond_paker_do["do"] - 20}-2м не более 15-20 м/мин;\n'
                       f'в устьевом оборудовании не более 0.5 м/мин.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'В процессе спуска производить приподъёмы для проверки веса на высоту не '
                       f'менее 20м со скоростью не более 5м/мин через каждые 300-500м спуска (первичный-последующий спуск).',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29, f'Перед каждой промывкой и после проверять веса ГТ (вверх, вниз, собств.)',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],

            [None, 29,
             f'Не допускать увеличение нагрузки на г/трубу в процессе спуска. РАЗГРУЗКА Г/ТРУБЫ НЕ БОЛЕЕ 500 кг от '
             f'собственного веса на этой глубине.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 29,
             f'При проведении технологического отстоя - не оставлять ГНКТ без движения: производить расхаживания'
             f' г/трубы на 20м '
             f'вверх и на 20м вниз со скоростью СПО не более 3м/мин. При отрицательной температуре окружающей среды, '
             f'во избежании получения ледяной '
             f'пробки в г/трубе при проведении тех.отстоя ни в коем случае не прекращать минимальную циркуляцию '
             f'жидкости по г/трубе.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None]
        ]
        osvoenie_list = [[None,  f'ОСВОЕНИЕ СКВАЖИНЫ', None,
             None, None, None, None, None, None, None,
             None, None],
            [None, 29, f'Установить КНК-1 на гл.{well_data.depth_fond_paker_do["do"] - 20}м, произвести вывод на '
                       f'режим мобильного азотного комплекса. Дождаться выхода пузыря азота. '
                       f'В случае отсутствия выхода пузыря азота более 1-1.5 часа - начать постепенный приподъём ГНКТ, не '
                       f'прекращая отдувки, до выхода пузыря азота. После получения прорыва азота и выхода пузыря - '
                       f'произвести допуск ГНКТ с одновременной отдувкой азотом до гл. {current_bottom_edit}м',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 1],
            [None, 29,
             f'Произвести освоение скважины один цикл в течении 4х часов с расходом по азоту в процессе освоения: \n'
             f'Первый час освоения - расход азота не менее 14м3/мин; \n'
             f'Второй час освоения - расход азота не менее 16м3/мин; ',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 4],
            [None, 29, f'Последующие 2 часа освоения производить на выбранном оптимальном режиме освоения '
                       f'(по согласованию с Заказчиком) позволяющим производить оптимальный отбор флюида из пласта. '
                       f'В ПРОЦЕССЕ ОСВОЕНИЯ (не оставлять г/трубу без движения) ПРОИЗВОДИТЬ ПРИПОДЪЁМЫ '
                       f'ГНКТ ЧЕРЕЗ 15-20 МИНУТ ВВЕРХ-ВНИЗ НА 20м. '
                       f'В процессе освоения проводить замеры притока каждый час и общее количество '
                       f'отобранной из пласта жидкости. В сводке отразить: процентное содержание нефти '
                       f'в возвратной жидкости, средний приток пластового флюида, количество отобранной из '
                       f'пласта жидкости. Отбор проб производить каждый  час. Пробы отбирать через пробоотборник, '
                       f'расположенный согласно схеме обвязки устья. По отдельному запросу геологической службы, '
                       f'пробы предоставлять оператору ЦДНГ для определения КВЧ и процентного содержания воды, '
                       f'с составлением акта передачи проб.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, f'ПОДТВЕРЖДЕНИЕ НОРМАЛИЗОВАННОГО ЗАБОЯ', None,
             None, None, None, None, None, None, None,
             None, None],
            [None, 29, f'По окончании освоения, остановить закачку азота, '
                       f'дождаться полного выхода азота из скважины и произвести тех.отстой '
                       f'скважины для оседания твёрдых частиц – в течении 2 часов',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 2],
            [None, 29, f'По истечении 2х часов, произвести допуск КНК-1 на г/трубе в скважину '
                       f'«без циркуляции» до гл. {current_bottom_edit}м.'
                       f' Забой должен соответствовать – {current_bottom_edit}м. '
                       f'Составить АКТ на забой совмесно с представителем Заказчика. '
                       f'При отсутствии нормализованного забоя на гл. {current_bottom_edit}м '
                       f'(по согласованию с Заказчиком) - провести работы по нормализации забоя.',
             None, None, None, None, None, None, None,
             'Мастер ГНКТ', 0.5],
        ]
        if osvoenie_combo_need == 'Да':
            for index, row in enumerate(osvoenie_list):
                gnkt_opz.insert(index + 32, row)

        for row in block_gnvp_list[::-1]:
            gnkt_opz.insert(0, row)
        return gnkt_opz

    def calc_volume_jumping(self):
        from work_py.alone_oreration import volume_vn_ek, volume_vn_nkt
        volume = round((volume_vn_ek(well_data.current_bottom) *
                        (well_data.current_bottom - well_data.depth_fond_paker_do["do"]) / 1000 +
                        volume_vn_nkt(well_data.dict_nkt) *
                        well_data.depth_fond_paker_do["do"] / 1000) * 1.2, 1)
        return volume

    def copy_pvr(self, ws, work_list):
        for row in range(len(work_list)):
            for col in range(23):
                if work_list[row][col]:
                    # print(row, col)
                    ws.cell(row=row + 1, column=col + 1).value = work_list[row][col]
        # Перебираем строки и скрываем те, у которых все значения равны None
        for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
            if all(value is None for value in row[:42]):
                ws.row_dimensions[row_ind + 1].hidden = True



    def save_to_gnkt(self):
        from work_py.gnkt_frez import Work_with_gnkt
        from main import MyWindow

        sheets = ["Титульник", 'СХЕМА', 'Ход работ']
        tables = [self.table_title, self.table_schema, self.table_widget]

        for i, sheet_name in enumerate(sheets):
            worksheet = GnktOsvWindow.wb[sheet_name]
            table = tables[i]
            work_list = []

            for row in range(table.rowCount()):
                row_lst = []

                for column in range(table.columnCount()):
                    item = table.item(row, column)
                    if not item is None:
                        row_lst.append(item.text())
                    else:
                        row_lst.append("")
                work_list.append(row_lst)
            Work_with_gnkt.count_row_height(self, worksheet, work_list, sheet_name)

        # ws6 = GnktOsvWindow.wb.create_sheet(title="СХЕМЫ КНК_44,45")
        # main.MyWindow.insert_image(self, ws6, '_internal/imageFiles/schema_well/СХЕМЫ КНК_44,45.png', 'A1', 550, 900)
        ws7 = GnktOsvWindow.wb.create_sheet(title="СХЕМЫ КНК_38,1")
        main.MyWindow.insert_image(self, ws7, '_internal/imageFiles/schema_well/СХЕМЫ КНК_38,1.png', 'A1', 550, 900)


        if 'Зуфаров' in well_data.user:
            path = 'D:\Documents\Desktop\ГТМ'
        else:
            path = ''

        filenames = f"{well_data.well_number._value} {well_data.well_area._value} кат {well_data.cat_P_1[0]} " \
                    f"{self.work_plan}.xlsx"
        full_path = path + '/' + filenames

        insert_data_base_gnkt(well_data.contractor, filenames, well_data.gnkt_number, int(well_data.gnkt_length),
                              float(well_data.diametr_length),
                              float(well_data.iznos)*1.014,
                              int(well_data.pipe_mileage) + int(well_data.current_bottom * 1.1),
                              well_data.pipe_fatigue, int(well_data.pvo), well_data.previous_well)

        if well_data.bvo is True:
            ws5 = GnktOsvWindow.wb.create_sheet('Sheet1')
            ws5.title = "Схемы ПВО"
            ws5 = GnktOsvWindow.wb["Схемы ПВО"]
            GnktOsvWindow.wb.move_sheet(ws5, offset=-1)
            schema_list = MyWindow.check_pvo_schema(self, ws5, 1)

        if GnktOsvWindow.wb:
            main.MyWindow.saveFileDialog(self, GnktOsvWindow.wb, full_path)
            Work_with_gnkt.wb_gnkt_frez.close()
            print(f"Table data saved to Excel {full_path} {well_data.number_dp}")
        if self.wb:
            self.wb.close()

    def insert_image_schema(self, ws):

        if well_data.paker_do["do"] != 0:
            coordinate_nkt_with_paker = 'F6'
            main.MyWindow.insert_image(self, ws, '_internal/imageFiles/schema_well/НКТ с пакером.png',
                                       coordinate_nkt_with_paker, 100, 470)
        else:
            coordinate_nkt_with_voronka = 'F6'
            main.MyWindow.insert_image(self, ws, '_internal/imageFiles/schema_well/НКТ с воронкой.png',
                                       coordinate_nkt_with_voronka, 70, 470)

        coordinate_propant = 'F43'
        if self.work_plan == 'gnkt_after_grp':
            main.MyWindow.insert_image(self, ws, '_internal/imageFiles/schema_well/пропант.png', coordinate_propant, 90, 500)

        n = 0
        m = 0
        for plast in well_data.plast_all:
            count_interval = well_data.dict_perforation[plast]['счет_объединение']

            ws.merge_cells(start_column=23, start_row=27 + m,
                                       end_column=23, end_row=27 + count_interval + m - 1)
            ws.merge_cells(start_column=22, start_row=27 + m,
                                       end_column=22, end_row=27 + count_interval + m - 1)
            ws.merge_cells(start_column=21, start_row=27 + m,
                                       end_column=21, end_row=27 + count_interval + m - 1)
            m += count_interval
            roof_plast = well_data.dict_perforation[plast]['кровля']
            sole_plast = well_data.dict_perforation[plast]['подошва']
            try:
                if roof_plast > well_data.depth_fond_paker_do["do"] and roof_plast < well_data.current_bottom:
                    interval_str = f'{plast} {roof_plast}-{sole_plast}'
                    coordinate_pvr = f'F{48 + n}'

                    ws.cell(row=48 + n, column=10).value = interval_str
                    ws.merge_cells(start_column=10, start_row=48 + n,
                                   end_column=12, end_row=48 + n + 2)
                    ws.cell(row=48 + n, column=10).font = Font(name='Arial', size=12, bold=True)
                    ws.cell(row=48 + n, column=10).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                         vertical='center')
                    n += 3
                    main.MyWindow.insert_image(self, ws, '_internal/imageFiles/schema_well/ПВР.png', coordinate_pvr, 85, 70)
            except:
                mes = QMessageBox.critical(self, 'Ошибка', f'программа не смогла вставить интервал перфорации в схему'
                                                           f'{roof_plast}-{sole_plast}')

        coordinate_voln = f'E18'
        main.MyWindow.insert_image(self, ws, '_internal/imageFiles/schema_well/переход.png', coordinate_voln, 150, 60)


    def date_dmy(self, date_str):
        date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
        # print(date_obj)
        # print(date_str)

        if isinstance(date_obj, datetime):
            return date_obj.strftime('%d.%m.%Y')
        else:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            # print(f' даь {date_obj}')
        return date_obj.strftime('%d.%m.%Y')

    def calc_fluid(self):

        fluid_list = []
        try:

            fluid_p = 0.83
            for plast in well_data.plast_work:
                if float(list(well_data.dict_perforation[plast]['рабочая жидкость'])[0]) > fluid_p:
                    fluid_p = list(well_data.dict_perforation[plast]['рабочая жидкость'])[0]
            fluid_list.append(fluid_p)

            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите расчетный удельный вес жидкости глушения в '
                                                           'конце жидкости',
                                                           max(fluid_list), 0.87, 2, 2)
        except:
            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите удельный вес рабочей жидкости',
                                                           0, 0.87, 2, 2)
        return fluid_work_insert


if __name__ == '__main__':
    app = QApplication([])
    window = GnktOsvWindow()
    window.show()
    app.exec_()
