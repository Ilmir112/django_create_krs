from datetime import datetime

from PyQt5.QtWidgets import QInputDialog, QMainWindow, QApplication

from openpyxl.utils import get_column_letter

import well_data
from perforation_correct_gnkt_frez import PerforationCorrectGnktFrez

import block_name
import main
import plan
from block_name import razdel_1
from openpyxl.styles import Border, Side, PatternFill, Font, Alignment
from openpyxl.workbook import Workbook
from gnkt_data.gnkt_data import dict_saddles
from work_py.alone_oreration import volume_jamming_well, well_volume, volume_nkt_metal, volume_nkt
from work_py.gnkt_grp import GnktOsvWindow
from work_py.gnkt_grp_work import GnktOsvWindow2
from .data_informations import dict_data_cdng, calc_pntzh

class Work_with_gnkt(QMainWindow):
    wb_gnkt_frez = Workbook()

    def __init__(self, ws, table_title, table_schema, table_widget):
        from open_pz import CreatePZ

        super(QMainWindow, self).__init__()
        self.table_widget = table_widget
        self.table_title = table_title
        self.table_schema = table_schema

        self.dict_perforation = well_data.dict_perforation
        self.ws = ws
        self.work_plan = 'gnkt_frez'
        self.perforation_correct_window2 = None
        self.perforation_correct_window = None

        if self.perforation_correct_window2 is None:
            self.perforation_correct_window2 = PerforationCorrectGnktFrez(self)
            self.perforation_correct_window2.setWindowTitle("Сверка данных по муфтам")
            # self.perforation_correct_window2.setGeometry(200, 400, 100, 400)
            self.perforation_correct_window2.show()
            main.MyWindow.pause_app()
            well_data.pause = True
            self.dict_ports = self.perforation_correct_window2.addRowTable()
            self.perforation_correct_window2 = None

        else:
            self.perforation_correct_window2.close()
            self.perforation_correct_window2 = None

        if self.perforation_correct_window is None:
            self.perforation_correct_window = GnktOsvWindow2(self)
            self.perforation_correct_window.setWindowTitle("Данные по ГНКТ")
            self.perforation_correct_window.setGeometry(200, 400, 100, 400)
            self.perforation_correct_window.show()
            well_data.pause = True
            main.MyWindow.pause_app()
            self.perforation_correct_window = None

        else:
            self.perforation_correct_window.close()
            self.perforation_correct_window = None

        # print(f' порты {self.dict_ports}')
        self.manufacturer = list(self.dict_ports.keys())[0]
        # print(f' порты {self.manufacturer}')
        self.type_column = list(self.dict_ports[self.manufacturer].keys())[0]
        # print(f' порты {self.type_column}')
        self.ports_data = self.dict_ports[self.manufacturer][self.type_column]
        # print(f' порты {self.ports_data}')
        self.top_muft = list(self.ports_data.keys())[-1]
        self.bottom_muft = list(self.ports_data.keys())[0]
        self.ws_title = Work_with_gnkt.wb_gnkt_frez.create_sheet(title="Титульник")

        self.ws_schema = Work_with_gnkt.wb_gnkt_frez.create_sheet(title="Схема")
        self.ws_work = Work_with_gnkt.wb_gnkt_frez.create_sheet(title="Ход работ")

        head = plan.head_ind(well_data.cat_well_min._value, well_data.cat_well_max._value)

        plan.copy_true_ws(self.ws, self.ws_title, head)

        create_title = self.create_title_list(self.ws_title)
        schema_well = self.schema_well(self.ws_schema)

        main.MyWindow.copy_pz(self, self.ws_title, table_title, 'gnkt_frez', 13, 1)
        main.MyWindow.copy_pz(self, self.ws_schema, table_schema, 'gnkt_frez', 47, 2)
        main.MyWindow.copy_pz(self, self.ws_work, table_widget, 'gnkt_frez', 12, 3)
        work_well = self.work_gnkt_frez(self.ports_data, self.plast_work)
        main.MyWindow.populate_row(self, 0, work_well, table_widget)

        CreatePZ.add_itog(self, self.ws_work, self.table_widget.rowCount() + 1, self.work_plan)

    def count_row_height(self, ws2, work_list, sheet_name):

        from openpyxl.utils.cell import range_boundaries, get_column_letter

        colWidth = [2.85546875, 14.42578125, 16.140625, 22.85546875, 17.140625, 14.42578125, 13.0, 13.0, 17.0,
                    14.42578125, 13.0, 21, 12.140625, None]

        text_width_dict = {35: (0, 100), 50: (101, 200), 70: (201, 300), 110: (301, 400), 120: (401, 500),
                           130: (501, 600), 150: (601, 700), 170: (701, 800), 190: (801, 900), 230: (901, 1500)}

        boundaries_dict = {}

        for ind, _range in enumerate(ws2.merged_cells.ranges):
            boundaries_dict[ind] = range_boundaries(str(_range))

        for key, value in boundaries_dict.items():
            ws2.unmerge_cells(start_column=value[0], start_row=value[1],
                              end_column=value[2], end_row=value[3])

        ins_ind = 1

        for i in range(1, len(work_list) + 1):  # Добавлением работ
            if sheet_name == 'Ход работ':
                if len(str(work_list[i - 1][1])) <= 3 and str(work_list[i - 1][1]) != '№':  # Нумерация
                    work_list[i - 1][1] = str(ins_ind)
                    ins_ind += 1
                else:
                    ins_ind = 1
            for j in range(1, 13):
                cell = ws2.cell(row=i, column=j)

                if cell and str(cell) != str(work_list[i - 1][j - 1]):
                    if str(work_list[i - 1][j - 1]).replace('.', '').isdigit() and \
                            str(work_list[i - 1][j - 1]).count('.') != 2:
                        cell.value = str(work_list[i - 1][j - 1]).replace('.', ',')
                        # print(f'цифры {cell.value}')
                    else:
                        cell.value = work_list[i - 1][j - 1]

        # print(merged_cells_dict)
        if sheet_name != 'Ход работ':
            for key, value in boundaries_dict.items():
                # print(value)
                ws2.merge_cells(start_column=value[0], start_row=value[1],
                                end_column=value[2], end_row=value[3])
            if sheet_name == 'СХЕМА':

                GnktOsvWindow.insert_image_schema(self, ws2)
                ws2.print_area = f'B3:AP{70}'
                # print(ws2, type(ws2))
                # ws2.page_setup.fitToPage = True
                # ws2.page_setup.fitToHeight = False
                # ws2.page_setup.fitToWidth = True
                # ws2.print_options.horizontalCentered = True

        elif sheet_name == 'Ход работ':
            for i, row_data in enumerate(work_list):
                # print(f'gghhg {work_list[i][2]}')
                for column, data in enumerate(row_data):
                    if column == 2:
                        if not data is None:
                            text = data
                            for key, value in text_width_dict.items():
                                if value[0] <= len(text) <= value[1]:
                                    ws2.row_dimensions[i + 1].height = int(key)
                    elif column == 1:
                        if not data is None:
                            text = data
                            # print(text)
                            for key, value in text_width_dict.items():
                                if value[0] <= len(text) <= value[1]:
                                    ws2.row_dimensions[i + 1].height = int(key)
                    if column != 0:
                        ws2.cell(row=i + 1, column=column + 1).border = well_data.thin_border
                    if column == 1 or column == 11:
                        ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(wrap_text=True,
                                                                                     horizontal='center',
                                                                                     vertical='center')
                        ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=False)
                    else:
                        ws2.cell(row=i + 1, column=column + 1).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                                     vertical='center')
                        ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=False)
                        if 'примечание' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'внимание' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'мероприятия' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'порядок работ' in str(ws2.cell(row=i + 1, column=column + 1).value).lower() or \
                                'По доп.согласованию с Заказчиком' in str(
                            ws2.cell(row=i + 1, column=column + 1).value).lower():
                            # print('есть жирный')
                            ws2.cell(row=i + 1, column=column + 1).font = Font(name='Arial', size=13, bold=True)

                if len(work_list[i][1]) > 5:
                    ws2.merge_cells(start_column=2, start_row=i + 1, end_column=12, end_row=i + 1)
                    ws2.cell(row=i + 1, column=2).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                        vertical='center')
                    ws2.cell(row=i + 1, column=2).fill = PatternFill(start_color='C5D9F1', end_color='C5D9F1',
                                                                     fill_type='solid')
                    ws2.cell(row=i + 1, column=2).font = Font(name='Arial', size=13, bold=True)

                else:
                    ws2.merge_cells(start_column=3, start_row=i + 1, end_column=11, end_row=i + 1)
                    ws2.cell(row=i + 1, column=3).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                        vertical='center')

            for col in range(13):
                ws2.column_dimensions[get_column_letter(col + 1)].width = colWidth[col]

            ws2.print_area = f'B1:L{self.table_widget.rowCount() + 45}'
            ws2.page_setup.fitToPage = True
            ws2.page_setup.fitToHeight = False
            ws2.page_setup.fitToWidth = True
            ws2.print_options.horizontalCentered = True
            # зададим размер листа
            ws2.page_setup.paperSize = ws2.PAPERSIZE_A4
            # содержимое по ширине страницы
            ws2.sheet_properties.pageSetUpPr.fitToPage = True
            ws2.page_setup.fitToHeight = False

        for row_ind, row in enumerate(ws2.iter_rows(values_only=True)):
            for col, value in enumerate(row):
                if 'А.Р. Хасаншин' in str(value):
                    coordinate = f'{get_column_letter(col + 1)}{row_ind - 1}'
                    self.insert_image(ws2, '_internal/imageFiles/Хасаншин.png', coordinate)
                elif 'Д.Д. Шамигулов' in str(value):
                    coordinate = f'{get_column_letter(col + 1)}{row_ind - 2}'
                    self.insert_image(ws2, '_internal/imageFiles/Шамигулов.png', coordinate)
                elif 'Зуфаров' in str(value):
                    coordinate = f'{get_column_letter(col - 2)}{row_ind}'
                    self.insert_image(ws2, '_internal/imageFiles/Зуфаров.png', coordinate)
                elif 'М.К.Алиев' in str(value):
                    coordinate = f'{get_column_letter(col - 1)}{row_ind - 2}'
                    self.insert_image(ws2, '_internal/imageFiles/Алиев махир.png', coordinate)
                elif 'З.К. Алиев' in str(value):
                    coordinate = f'{get_column_letter(col - 1)}{row_ind - 2}'
                    self.insert_image(ws2, '_internal/imageFiles/Алиев Заур.png', coordinate)
                    break
        print(f'{sheet_name} - вставлена')

    def save_to_gnkt(self):

        sheets = ["Титульник", 'Схема', 'Ход работ']
        tables = [self.table_title, self.table_schema, self.table_widget]

        for i, sheet_name in enumerate(sheets):
            worksheet = Work_with_gnkt.wb_gnkt_frez[sheet_name]
            table = tables[i]

            work_list = []
            for row in range(table.rowCount()):
                row_lst = []
                # self.ins_ind_border += 1
                for column in range(table.columnCount()):

                    item = table.item(row, column)
                    if not item is None:

                        row_lst.append(item.text())
                        # print(item.text())
                    else:
                        row_lst.append("")
                work_list.append(row_lst)
            Work_with_gnkt.count_row_height(self, worksheet, work_list, sheet_name)

        ws6 = Work_with_gnkt.wb_gnkt_frez.create_sheet(title="СХЕМЫ КНК_44,45")
        main.MyWindow.insert_image(self, ws6, '_internal/imageFiles/schema_well/СХЕМЫ КНК_44,45.png', 'A1', 550, 900)
        ws7 = Work_with_gnkt.wb_gnkt_frez.create_sheet(title="СХЕМЫ КНК_38,1")
        main.MyWindow.insert_image(self, ws7, '_internal/imageFiles/schema_well/СХЕМЫ КНК_38,1.png', 'A1', 550, 900)

        # path = 'workiii'
        if 'Зуфаров' in well_data.user:
            path = 'D:\Documents\Desktop\ГТМ'
        else:
            path = ""

        filenames = f"{well_data.well_number._value} {well_data.well_area._value} кат " \
                    f"{well_data.cat_P_1} {self.work_plan}.xlsx"
        full_path = path + '/' + filenames


        if well_data.bvo is True:
            ws5 = Work_with_gnkt.wb_gnkt_frez.create_sheet('Sheet1')
            ws5.title = "Схемы ПВО"
            ws5 = Work_with_gnkt.wb_gnkt_frez["Схемы ПВО"]
            Work_with_gnkt.wb_gnkt_frez.move_sheet(ws5, offset=-1)
            # schema_list = self.check_pvo_schema(ws5, ins_ind + 2)

        if Work_with_gnkt.wb_gnkt_frez:
            Work_with_gnkt.wb_gnkt_frez.remove(Work_with_gnkt.wb_gnkt_frez['Sheet'])

            main.MyWindow.saveFileDialog(self, Work_with_gnkt.wb_gnkt_frez, full_path)

            Work_with_gnkt.wb_gnkt_frez.close()

        if self.wb:
            self.wb.close()

    def create_title_list(self, ws2):

        well_data.region = block_name.region(well_data.cdng._value)
        self.region = well_data.region

        title_list = [
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ЗАКАЗЧИК:', None, None, None, None, None, None, None, None, None, None],
            [None, 'ООО «Башнефть-Добыча»', None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ПЛАН РАБОТ НА СКВАЖИНЕ С ПОМОЩЬЮ УСТАНОВКИ С ГИБКОЙ ТРУБОЙ', None, None, None,
             None, None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None],
            [None, None, '№ скважины:', f'{well_data.well_number._value}', 'куст:', None, 'Месторождение:', None, None,
             well_data.well_oilfield._value, None, None],
            [None, None, 'инв. №:', well_data.inv_number._value, None, None, None, None, 'Площадь: ',
             well_data.well_area._value, None,
             1],
            [None, None, None, None, None, None, None, None, 'цех', f'{well_data.cdng._value}', None, None]]

        razdel = razdel_1(self, well_data.region)
        for row in razdel:  # Добавлением работ
            title_list.append(row)

        for row in [
            [None, None, None, None, None, "дата", datetime.now().strftime('%d.%m.%Y'), None, None, None, None]]:
            title_list.insert(-1, row)
        # print(title_list)
        index_insert = 11
        ws2.cell(row=1, column=2).alignment = Alignment(wrap_text=False, horizontal='left',
                                                        vertical='center')
        ws2.cell(row=1, column=4).alignment = Alignment(wrap_text=False, horizontal='left',
                                                        vertical='center')
        # ws2.column_dimensions[get_column_letter(1)].width = 15
        # ws2.column_dimensions[get_column_letter(2)].width = 20
        # ws2.column_dimensions[get_column_letter(3)].width = 20
        a = None
        for row in range(len(title_list)):  # Добавлением работ
            if row not in range(8, 13):
                ws2.row_dimensions[row].height = 35
            for col in range(1, 12):
                ws2.column_dimensions[get_column_letter(col)].width = 15
                cell = ws2.cell(row=row + index_insert, column=col)
                # print(f' Х {title_list[i ][col - 1]}')
                if title_list[row - 1][col - 1] != None:
                    ws2.cell(row=row + index_insert, column=col).value = str(title_list[row - 1][col - 1])
                ws2.cell(row=row + index_insert, column=col).font = Font(name='Arial', size=11, bold=False)
                ws2.cell(row=row + index_insert, column=col).alignment = Alignment(wrap_text=False, horizontal='left',
                                                                                   vertical='center')
                if 'ПЛАН РАБОТ' in str(title_list[row - 1][col - 1]):
                    ws2.merge_cells(start_row=row + index_insert, start_column=2, end_row=row + index_insert,
                                    end_column=11)
                    ws2.merge_cells(start_row=row - 4 + index_insert, start_column=2, end_row=row - 4 + index_insert,
                                    end_column=4)
                    ws2.cell(row=row + index_insert, column=col).font = Font(name='Arial', size=14, bold=True)
                    ws2.cell(row=row + index_insert, column=col).alignment = Alignment(wrap_text=False,
                                                                                       horizontal='center',
                                                                                       vertical='center')
                if 'СОГЛАСОВАНО:' in str(title_list[row - 1][col - 1]):
                    a = row + index_insert

            # for row in range(len(title_list)):  # Добавлением работ
            if a:
                if a > row:
                    # print(f'сссооссоссо {row + index_insert}')
                    ws2.merge_cells(start_row=row + index_insert, start_column=2, end_row=row + index_insert,
                                    end_column=6)
                    ws2.merge_cells(start_row=row + index_insert, start_column=8, end_row=row + index_insert,
                                    end_column=11)

        ws2.print_area = f'B1:K{44}'
        ws2.page_setup.fitToPage = True
        ws2.page_setup.fitToHeight = False
        ws2.page_setup.fitToWidth = True
        ws2.print_options.horizontalCentered = True
        # зададим размер листа
        ws2.page_setup.paperSize = ws2.PAPERSIZE_A4

    def schema_well(self, ws3):

        from work_py.alone_oreration import volume_vn_nkt, well_volume

        boundaries_dict = {0: (13, 13, 14, 14), 1: (43, 12, 45, 12), 2: (40, 16, 42, 16), 3: (7, 19, 12, 19),
                           4: (17, 21, 18, 21), 5: (19, 21, 20, 21), 6: (13, 10, 30, 10), 7: (15, 15, 16, 15),
                           8: (1, 1, 49, 2), 9: (46, 19, 48, 19), 10: (27, 15, 28, 15), 11: (29, 15, 30, 15),
                           12: (40, 11, 42, 11), 13: (33, 5, 48, 5), 14: (23, 15, 26, 15), 15: (13, 16, 14, 16),
                           16: (15, 16, 16, 16), 17: (9, 34, 48, 34), 18: (19, 19, 20, 19), 19: (27, 16, 28, 16),
                           20: (29, 19, 30, 19), 21: (13, 18, 14, 18), 22: (40, 13, 42, 13), 23: (23, 16, 26, 16),
                           25: (27, 18, 28, 18), 26: (13, 17, 14, 17), 27: (15, 17, 16, 17),
                           28: (17, 17, 18, 17), 29: (32, 8, 39, 8), 30: (40, 15, 42, 15), 31: (19, 20, 20, 20),
                           32: (21, 20, 22, 20), 33: (21, 14, 22, 14), 34: (13, 19, 14, 19),
                           36: (11, 13, 12, 13), 37: (43, 7, 48, 7), 38: (37, 3, 48, 3), 39: (7, 6, 12, 11), 40:
                               (27, 22, 28, 22), 41: (19, 22, 20, 22), 42: (32, 18, 39, 18), 43: (21, 22, 22, 22),
                           44: (46, 23, 48, 23), 45: (7, 18, 12, 18), 47: (43, 8, 48, 8),
                           48: (13, 7, 30, 7), 49: (46, 15, 48, 15), 50: (43, 14, 45, 14), 51: (40, 18, 42, 18),
                           52: (17, 23, 18, 23), 53: (43, 10, 48, 10), 54: (29, 14, 30, 14),
                           56: (40, 8, 42, 8), 57: (29, 23, 30, 23), 58: (43, 13, 45, 13), 59: (7, 20, 12, 20),
                           60: (15, 13, 16, 14), 61: (40, 17, 42, 17), 62: (17, 13, 18, 14), 63: (43, 9, 48, 9),
                           64: (13, 8, 30, 8), 65: (27, 13, 30, 13), 67: (13, 21, 14, 21),
                           68: (15, 21, 16, 21), 69: (40, 10, 42, 10), 70: (17, 15, 18, 15), 71: (43, 15, 45, 15),
                           72: (32, 12, 39, 12), 73: (7, 22, 12, 22), 74: (40, 19, 42, 19), 75: (32, 21, 39, 21),
                           76: (46, 17, 48, 17), 77: (21, 18, 22, 18), 78: (46, 22, 48, 22), 79: (7, 12, 12, 12),
                           80: (40, 9, 42, 9), 81: (7, 21, 12, 21), 82: (10, 5, 30, 5), 83: (32, 7, 39, 7),
                           85: (7, 23, 12, 23), 86: (13, 9, 30, 9), 87: (46, 12, 48, 12),
                           88: (21, 19, 22, 19), 89: (43, 16, 45, 16), 90: (2, 34, 8, 34), 91: (32, 22, 45, 22),
                           92: (27, 17, 28, 17), 94: (22, 3, 26, 3), 95: (29, 17, 30, 17), 93: (2, 36, 6, 36),
                           96: (43, 21, 48, 21), 97: (23, 18, 26, 18), 98: (13, 11, 30, 11), 99: (46, 20, 48, 20),
                           100: (15, 19, 16, 19), 101: (10, 38, 11, 38), 102: (46, 14, 48, 14), 103: (43, 18, 45, 18),
                           104: (27, 19, 28, 19), 105: (23, 17, 26, 17), 106: (43, 17, 45, 17), 107: (40, 21, 42, 21),
                           108: (23, 19, 26, 19), 109: (13, 12, 30, 12), 110: (15, 20, 16, 20), 111: (14, 3, 21, 3),
                           112: (17, 20, 18, 20), 113: (43, 19, 45, 19), 114: (32, 16, 39, 16), 115: (19, 14, 20, 14),
                           116: (15, 22, 16, 22), 117: (40, 7, 42, 7), 118: (32, 9, 39, 9), 119: (13, 15, 14, 15),
                           120: (21, 21, 22, 21), 121: (32, 15, 39, 15), 122: (32, 11, 39, 11), 123: (46, 16, 48, 16),
                           124: (7, 13, 10, 13), 125: (27, 21, 28, 21), 126: (32, 23, 45, 23), 127: (29, 21, 30, 21),
                           128: (43, 11, 48, 11), 129: (23, 13, 26, 14), 130: (13, 23, 14, 23), 131: (40, 6, 48, 6),
                           132: (19, 13, 22, 13), 133: (15, 23, 16, 23), 134: (46, 18, 48, 18), 135: (27, 23, 28, 23),
                           137: (23, 21, 26, 21), 138: (19, 16, 22, 16), 139: (32, 13, 39, 13),
                           140: (40, 20, 42, 20), 141: (19, 15, 22, 15), 142: (17, 16, 18, 16), 143: (29, 16, 30, 16),
                           144: (15, 18, 16, 18), 145: (10, 37, 11, 37), 146: (17, 18, 18, 18), 147: (19, 18, 20, 18),
                           148: (29, 18, 30, 18), 149: (7, 15, 12, 15), 150: (40, 12, 42, 12), 151: (13, 6, 30, 6),
                           152: (13, 20, 14, 20), 153: (19, 16, 22, 16), 154: (30, 3, 36, 3), 155: (32, 10, 39, 10),
                           156: (40, 14, 42, 14), 157: (17, 19, 18, 19), 158: (19, 23, 26, 23), 159: (7, 24, 48, 24),
                           160: (32, 6, 39, 6), 161: (13, 22, 14, 22), 162: (27, 20, 28, 20), 163: (7, 16, 12, 16),
                           164: (29, 20, 30, 20), 165: (19, 17, 22, 17), 166: (17, 22, 18, 22), 167: (32, 20, 39, 20),
                           168: (32, 14, 39, 14), 169: (11, 14, 12, 14), 170: (23, 20, 26, 20), 171: (29, 22, 30, 22),
                           172: (46, 13, 48, 13), 173: (43, 20, 45, 20), 174: (32, 17, 39, 17), 175: (23, 22, 26, 22),
                           176: (7, 17, 12, 17), 177: (32, 19, 39, 19), 178: (27, 14, 28, 14), 179: (9, 36, 43, 36),
                           180: (7, 14, 10, 14)}

        rowHeights1 = [None, None, 27.75, 20.25, 20.25, 20.25, 20.25, 18.0, 22.5, 22.5, 22.5, 18.0, 18.0, 20.25,
                       20.25, 20.25, 20.25, 20.25, 20.25, 20.25, 20.25, 18.0, 20.25, 35.25, 17.25, 17.25, 79.5, 60.0,
                       13.5, 13.5, 43.5, 13.5, None, 45.75, None, 74.25, None, None, None, None, None, None, 13.5,
                       12.75, 12.75, 12.75, None, None, None, None, None, None, None, 15.75, 12.75, 12.75, 12.75,
                       None, None, None, None, None, None, None, 15.75, 15.75, 12.75, 12.75, 12.75, None, None, None,
                       None, None, None, 13.5, 12.75, None, None, None, None, None, None, None, None, None, None, 13.5,
                       None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                       None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
                       None, None, 134.25, None, None, None, None, None, None, None]

        colWidth = [2.28515625, 13.0, 4.5703125, 13.0, 13.0, 13.0, 5.7109375, 13.0, 13.0, 13.0, 4.7109375,
                    13.0, 5.140625, 13.0, 13.0, 13.0, 13.0, 13.0, 4.7109375, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
                    13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0, 13.0,
                    13.0, 13.0, 13.0, 5.42578125, 13.0, 4.5703125, 2.28515625, 10.28515625]

        self.plast_work = well_data.plast_all[0]
        plast_work = self.plast_work
        # print(self.plast_work, list(well_data.dict_perforation[plast_work]))
        self.pressuar = list(well_data.dict_perforation[plast_work]["давление"])[0]

        zamer = list(well_data.dict_perforation[plast_work]['замер'])[0]
        vertikal = min(map(float, list(well_data.dict_perforation[plast_work]["вертикаль"])))
        self.fluid = self.calc_fluid()
        self.zhgs = f'{self.fluid}г/см3'
        koef_anomal = round(float(self.pressuar) * 101325 / (float(vertikal) * 9.81 * 1000), 1)
        nkt = int(list(well_data.dict_nkt.keys())[0])
        if nkt == 73:
            nkt_widht = 5.5
        elif nkt == 89:
            nkt_widht = 6.5
        elif nkt == 60:
            nkt_widht = 5
        lenght_nkt = sum(list(map(int, well_data.dict_nkt.values())))

        bottom_first_port = self.ports_data['№1']['кровля']

        gnkt_lenght = int(well_data.gnkt_length)
        gnkt_lenght_str = f'длина ГНКТ - {well_data.gnkt_length}м; износ -{well_data.iznos}%; ' \
                      f'пробег '
        volume_vn_gnkt = round(30.2 ** 2 * 3.14 / (4 * 1000), 2)
        volume_gnkt = round(gnkt_lenght * volume_vn_gnkt / 1000, 1)

        well_volume_ek = well_volume(self, well_data.head_column_additional._value)
        well_volume_dp = well_volume(self, well_data.current_bottom) - well_volume_ek

        volume_pm_ek = round(
            3.14 * (well_data.column_diametr._value - 2 * well_data.column_wall_thickness._value) ** 2 / 4 / 1000, 2)
        volume_pm_dp = round(3.14 * (well_data.column_additional_diametr._value - 2 *
                                     well_data.column_additional_wall_thickness._value) ** 2 / 4 / 1000, 2)

        schema_well_list = [
            ['СХЕМА СКВАЖИНЫ', None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None, None,
             '№ скважины:', None, None, None, None, None, None, None, well_data.well_number._value, None, None, None, None,
             None, None,
             None, 'Месторождение:', None, None, None, None, None, None, well_data.well_oilfield._value, None, None, None,
             None,
             None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None,
             'Данные о размерности НКТ о конструкции скважины',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None,
             'Дополнительная информация', None, None, None, None, None, None, None, None, None, None, None, None, None,
             None],
            [None, None, None, None, None, None,
             'Оборудование \n устья скважины', None, None, None, None, None,
             'Лубрикатор + герметизатор', None, None, None, None, None, None,
             None, None, None, None, None, None, None,
             None, None, None, None, None,
             'Категория скважины по опасности', None, None, None, None, None, None, None,
             'первая [после бурения)', None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             'БП 80х70', None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             'Пластовое давление', None, None, None, None, None, None, None, f'{self.pressuar}атм',
             None, None, zamer, None, None,
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             'тройник 80х70-80х70 В60-В60',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             'Содержание H2S', None, None, None, None, None, None, None, round(well_data.h2s_mg[0], 5), None, None,
             None,
             None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None,
             f'ФА  ГРП ГУ 180х35-89 К1ХЛ № ',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             'Газовый фактор', None, None, None, None, None, None, None, well_data.gaz_f_pr[0], None, None, None, None,
             None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, 'Переходная катушка 180х21-89-3"',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             f'Глубина пл.{plast_work} по вертикали', None, None, None, None, None, None, None, vertikal, None, None,
             None, None, None, None, None],

            [' ', None, None, None, None, None, None, None, None, None, None, None,
             'Устьевая крестовина АУЭЦН-50х14-168 ОТТМ К1 ЛЗ ХЛ №181', None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None,
             'Коэффициент  аномальности', None, None, None, None, None, None, None, koef_anomal, None, None,
             None, None, None, None, None],

            [' ', None, None, None, None, None, 'Тип КГ', None, None, None, None, None,
             well_data.column_head_m, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None,
             'Макс.угол в горизонт. участке', None, None, None, None, None, None, None, well_data.max_angle._value, None, None,
             'на глубине', None, None, well_data.max_angle_H._value, None],

            [None, None, None, None, None, None,
             'Диаметр канавки', None, None, None, None, None, 'Наруж.\nдиаметр',
             None, 'Толщина стенки', None, 'Внутр.\nдиаметр', None, 'Глубина', None, None, None, 'ВЦП.\nДлина ПО', None,
             None, None, 'Объем', None, None, None, None, 'Макс. интенс. набора кривизны', None, None, None, None, None,
             None, None, f'{6.21}/10м', None, None, 'на глубине', None, None, '1310', None],

            [None, None, None, None, None, None,
             'Стол ротора', None, None, None, f'{well_data.stol_rotora._value}м', None, None, None, None, None,
             None, None, 'от', None, 'до', None, None, None, None, None, 'п.м', None, 'м3', None, None,
             'Жидкость глушения', None, None, None, None, None, None, None, self.zhgs, None, None, 'в объеме', None, None,
             f'{28.9}м3', None],
            [None, None, None, None, None, None, 'Направление', None, None, None, None, None,
             well_data.column_direction_diametr._value, None, well_data.column_direction_wall_thickness._value, None,
             well_data.column_direction_diametr._value - 2 * well_data.column_direction_wall_thickness._value, None,
             well_data.column_direction_lenght._value, None, None, None, well_data.level_cement_direction._value, None, None,
             None, None, None, None, None, None, 'Ожидаемый дебит',
             None, None, None, None, None, None, None, f'{well_data.Qwater}м3/сут', None, None,
             f'{well_data.Qoil}м3', None, None,
             f'{well_data.proc_water}%', None],
            [None, None, None, None, None, None, 'Кондуктор', None, None, None, None, None,
             well_data.column_conductor_diametr._value, None, well_data.column_conductor_wall_thickness._value, None,
             well_data.column_conductor_diametr._value - 2 * well_data.column_conductor_wall_thickness._value,
             None, well_data.column_conductor_lenght._value, None, None, None, well_data.level_cement_conductor._value,
             None, None, None, None, None, None, None, None,
             'Начало / окончание бурения', None, None, None, None, None, None, None,
             self.date_dmy(well_data.date_drilling_run), None, None,
             self.date_dmy(well_data.date_drilling_cancel), None, None, None,
             None],
            [None, None, None, None, None, None, 'Экспл. колонна', None, None, None, None, None,
             well_data.column_diametr._value, None, well_data.column_wall_thickness._value, None,
             well_data.column_diametr._value - 2 * well_data.column_wall_thickness._value, None,
             f'0-{well_data.shoe_column._value}м', None,
             None,
             None, well_data.level_cement_column._value, None, None, None, volume_pm_ek, None, well_volume_ek,
             None, None, 'Р в межколонном пространстве', None, None, None, None, None, None, None,
             f'{0}атм', None, None, None, self.date_dmy(well_data.date_drilling_cancel), None, None, None],
            [None, None, None, None, None, None, "Хвостовик  ''НТЦ ''ЗЭРС''", None, None, None, None,
             None, well_data.column_additional_diametr._value, None,
             well_data.column_additional_wall_thickness._value, None,
             well_data.column_additional_diametr._value - 2 * well_data.column_additional_wall_thickness._value, None,
             well_data.head_column_additional._value, None, well_data.shoe_column_additional._value, None,
             'не цементиров.', None,
             None, None, volume_pm_dp,
             None, well_volume_dp, None, None, 'Давление опрессовки МКП', None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, f'Подвеска НКТ {nkt}мм', None, None, None, None, None, nkt, None,
             nkt_widht,
             None, nkt - 2 * nkt_widht, None, f'{0}', None, f'{lenght_nkt - 0.5 - 2.6 - 3}м', None,
             f'{lenght_nkt - 0.5 - 2.6 - 3}м', None, None, None,
             volume_vn_nkt(well_data.dict_nkt), None, volume_vn_nkt(well_data.dict_nkt) + 0.47, None, None,
             'Давление опрессовки ЭК ', None, None, None, None,
             None, None, None, f'{well_data.max_admissible_pressure._value}атм', None, None,
             None,
             None, None, 'гермет.', None],
            [None, None, None, None, None, None, 'Гидроякорь ', None, None, None, None, None, 122, None, None, None, 71,
             None, f'{lenght_nkt}', None, f'{lenght_nkt + 1}м', None, f'{0.5}м3', None, None, None, None, None, None,
             None,
             None, 'Макс. допустимое Р опр-ки ЭК', None, None, None, None, None, None, None,
             well_data.max_admissible_pressure._value, None, None, None,
             None, None, None, None],
            [None, None, None, None, None, None, 'Патрубок 1 шт.', None, None, None, None, None, nkt, None, 6.5, None,
             74.2, None, lenght_nkt - 5.6, None, f'{lenght_nkt - 2.6}м', None, f'{3}м', None, None, None, None,
             None, None, None, None, 'Макс. ожидаемое Р на устье скв.', None, None, None, None, None, None, None, 92.8,
             None, None, None, None, None, None, None],
            [None, None, None, None, None, None, f'стингер {well_data.paker_do["do"]}', None, None, None, None, None,
             None, None,
             None, None, 71, None, lenght_nkt - 2.6, None, lenght_nkt, None, f'{2.6}м', None, None, None, None,
             None, None, None, None, 'Текущий забой до ГРП ', None, None, None, None, None, None, None, None, None,
             None, None, None, None, f'{well_data.current_bottom}м', None],
            [None, None, None, None, None, None, 'ГНКТ', None, None, None, None, None, 38.1, None, 3.96, None, 30.18,
             None, gnkt_lenght_str, None, None, None, None, None, None, None, volume_vn_gnkt, None,
             volume_gnkt, None,
             None, 'Искусственный забой  (МГРП №1)', None, None, None, None, None, None, None, None,
             None, None, None, None, None, f'{bottom_first_port}м', None]]

        # self.ports_data = self.work_with_port(self.plast_work, well_data.dict_perforation)
        self.ports_list, merge_port = self.insert_ports_data(self.ports_data)

        # print(ports_list)
        for row in self.ports_list:
            schema_well_list.append(row)

        border = Border(left=Side(border_style='dashed', color='FF000000'),
                        top=Side(border_style='dashed', color='FF000000'),
                        right=Side(border_style='dashed', color='FF000000'),
                        bottom=Side(border_style='dashed', color='FF000000'),
                        )
        border_left_top = Border(left=Side(border_style='thick', color='FF000000'),
                                 top=Side(border_style='thick', color='FF000000'),
                                 right=Side(border_style='dashed', color='FF000000'),
                                 bottom=Side(border_style='dashed', color='FF000000'),
                                 )
        border_left_bottom = Border(left=Side(border_style='thick', color='FF000000'),
                                    top=Side(border_style='dashed', color='FF000000'),
                                    right=Side(border_style='dashed', color='FF000000'),
                                    bottom=Side(border_style='thick', color='FF000000'),
                                    )
        border_right_bottom = Border(left=Side(border_style='dashed', color='FF000000'),
                                     top=Side(border_style='dashed', color='FF000000'),
                                     right=Side(border_style='thick', color='FF000000'),
                                     bottom=Side(border_style='thick', color='FF000000'),
                                     )
        border_right_top = Border(top=Side(border_style='thick', color='FF000000'),
                                  right=Side(border_style='thick', color='FF000000'))

        border_right = Border(left=Side(border_style='dashed', color='FF000000'),
                              top=Side(border_style='dashed', color='FF000000'),
                              right=Side(border_style='thick', color='FF000000'),
                              bottom=Side(border_style='dashed', color='FF000000'),
                              )
        border_left = Border(left=Side(border_style='thick', color='FF000000'),
                             top=Side(border_style='dashed', color='FF000000'),
                             right=Side(border_style='dashed', color='FF000000'),
                             bottom=Side(border_style='dashed', color='FF000000'),
                             )
        border_top = Border(left=Side(border_style='dashed', color='FF000000'),
                            top=Side(border_style='thick', color='FF000000'),
                            right=Side(border_style='dashed', color='FF000000'),
                            bottom=Side(border_style='dashed', color='FF000000'),
                            )
        border_bottom = Border(left=Side(border_style='dashed', color='FF000000'),
                               top=Side(border_style='dashed', color='FF000000'),
                               right=Side(border_style='dashed', color='FF000000'),
                               bottom=Side(border_style='thick', color='FF000000'),
                               )

        for row in range(1, len(schema_well_list) + 1):  # Добавлением работ
            # print(row, len(schema_well_list[row-1]), schema_well_list[row-1][15])
            for col in range(1, 48):
                cell = ws3.cell(row=row, column=col)
                # print(row, col)
                cell.value = schema_well_list[row - 1][col - 1]
                ws3.cell(row=row, column=col).font = Font(name='Arial', size=11, bold=False)
                ws3.cell(row=row, column=col).alignment = Alignment(wrap_text=True, horizontal='center',
                                                                    vertical='center')
                if cell.value != None and row > 24:
                    cell.border = border

        for row in range(6, 24):
            for col in range(7, 32):
                cell = ws3.cell(row=row, column=col)

                cell.border = border
                if col == 31:
                    cell.border = Border(left=Side(border_style='thick', color='FF000000'),
                                         right=Side(border_style='thick', color='FF000000'))
                if row == 6 and col != 31:
                    cell.border = border_top
                elif (row == 22) and col != 31:
                    cell.border = border_bottom
                elif (row == 23) and col != 31:
                    cell.border = border_bottom
                elif col == 7:
                    cell.border = border_left
                elif col == 30:
                    cell.border = border_right

                elif (row == 13 or row == 14) and col > 12 and col != 31:
                    cell.border = Border(left=Side(border_style='thin', color='FF000000'),
                                         top=Side(border_style='thin', color='FF000000'),
                                         right=Side(border_style='thin', color='FF000000'),
                                         bottom=Side(border_style='thin', color='FF000000'),
                                         )
            if row < 23:
                ws3.cell(row=row, column=7).font = Font(name='Arial', size=11, bold=True, color='002060')
                ws3.cell(row=row, column=32).font = Font(name='Arial', size=11, bold=True, color='002060')

            for col in range(32, 49):
                cell = ws3.cell(row=row, column=col)
                cell.border = border
                if row == 6:
                    cell.border = border_top
                elif (row == 22):
                    cell.border = border_bottom
                elif (row == 23):
                    cell.border = border_bottom
                elif (row == row and col == 32):
                    cell.border = border_left
                elif (row == row and col == 48):
                    cell.border = border_right

        ws3.cell(row=1, column=1).font = Font(name='Arial', size=18, bold=True)
        ws3.cell(row=3, column=14).font = Font(name='Arial', size=18, bold=True)
        ws3.cell(row=3, column=30).font = Font(name='Arial', size=18, bold=True)
        ws3.cell(row=3, column=22).font = Font(name='Arial', size=18, bold=True)
        ws3.cell(row=3, column=37).font = Font(name='Arial', size=18, bold=True)
        ws3.cell(row=24, column=7).font = Font(name='Arial', size=14, bold=True, color='002060')
        ws3.cell(row=34, column=1).font = Font(name='Arial', size=14, bold=True, color='002060')
        ws3.cell(row=34, column=9).font = Font(name='Arial', size=14, bold=True, color='002060')
        ws3.cell(row=5, column=10).font = Font(name='Arial', size=14, bold=False, color='002060', underline='single')
        ws3.cell(row=5, column=33).font = Font(name='Arial', size=14, bold=False, color='002060', underline='single')

        ws3.cell(6, 7).border = border_left_top
        ws3.cell(6, 32).border = border_left_top
        ws3.cell(22, 7).border = border_left_bottom
        ws3.cell(23, 7).border = border_left_bottom
        ws3.cell(22, 32).border = border_left_bottom
        ws3.cell(23, 32).border = border_left_bottom

        ws3.cell(6, 30).border = border_right_top
        ws3.cell(6, 48).border = border_right_top
        ws3.cell(22, 30).border = border_left_bottom
        ws3.cell(23, 30).border = border_left_bottom
        ws3.cell(22, 48).border = border_left_bottom
        ws3.cell(23, 48).border = border_right_bottom
        ws3.cell(23, 30).border = border_right_bottom

        for key, value in merge_port.items():
            boundaries_dict[key] = value

            if key % 2 == 0:
                coordinate = f'{get_column_letter(value[0] - 1)}{value[1] + 4}'
                # print(f'вставка1 ')
                column_img = f'H{value[1] + 6}'

                main.MyWindow.insert_image(self, ws3, '_internal/imageFiles/schema_well/port.png', coordinate, 200, 200)

            for i in range(3):
                cell = ws3.cell(row=27, column=value[0] + i)
                cell2 = ws3.cell(row=28, column=value[0] + i)
                font = Font(bold=True, italic=True)
                cell.font = font
                cell2.font = font
                cell.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')
                cell2.alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

        # print(boundaries_dict)
        for key, value in boundaries_dict.items():
            ws3.merge_cells(start_column=value[0], start_row=value[1],
                            end_column=value[2], end_row=value[3])

        for index_row, row in enumerate(ws3.iter_rows()):  # Копирование высоты строки
            ws3.row_dimensions[index_row].height = rowHeights1[index_row - 1]

        for col_ind in range(50):  # копирование ширины столба
            ws3.column_dimensions[get_column_letter(col_ind + 1)].width = colWidth[col_ind] / 1.9

        coordinate = f'B3'

        main.MyWindow.insert_image(self, ws3, '_internal/imageFiles/schema_well/gorizont_1.png', coordinate, 237, 1023)
        # print(Column_img)
        main.MyWindow.insert_image(self, ws3, '_internal/imageFiles/schema_well/gorizont_12.png', column_img, 1800, 120)

        ws3.print_area = f'A1:AW{37}'
        ws3.page_setup.fitToPage = True
        ws3.page_setup.fitToHeight = False
        ws3.page_setup.fitToWidth = True
        # Измените формат листа на альбомный
        ws3.page_setup.orientation = ws3.ORIENTATION_LANDSCAPE
        ws3.print_options.horizontalCentered = True
        # зададим размер листа
        ws3.page_setup.paperSize = ws3.PAPERSIZE_A4

    def work_gnkt_frez(self, ports_data, plast_work):
        from krs import GnoWindow
        from cdng import events_gnvp_frez
        if well_data.column_additional:
            if sum(list(well_data.dict_nkt.values())) != 0 and \
                    abs(well_data.depth_fond_paker_do['do'] - well_data.head_column_additional._value) > 30:
                ntk_true = True
                paker_true = False
                nkt_lenght = round(sum(list(well_data.dict_nkt.values())), 0)
            elif sum(list(well_data.dict_nkt.values())) != 0 and \
                        abs(well_data.depth_fond_paker_do['do'] - well_data.head_column_additional._value) < 30:
                ntk_true = True
                paker_true = True
                nkt_lenght = round(sum(list(well_data.dict_nkt.values())), 0)

            else:
                ntk_true = False
                paker_true = False
                nkt_lenght = 0
        else:
            if sum(list(well_data.dict_nkt.values())) != 0 and \
                    well_data.depth_fond_paker_do['do'] != 0:
                ntk_true = True
                paker_true = True
                nkt_lenght = round(sum(list(well_data.dict_nkt.values())), 0)
            elif sum(list(well_data.dict_nkt.values())) == 0:
                ntk_true = False
                paker_true = False
                nkt_lenght = round(sum(list(well_data.dict_nkt.values())), 0)
            else:
                ntk_true = True
                paker_true = False
                nkt_lenght = 0



        distance, _ = QInputDialog.getInt(None, 'Расстояние НПТЖ', 'Введите Расстояние до ПНТЖ')

        fluid_work_insert = GnktOsvWindow2.fluid_edit

        fluid_work, well_data.fluid_work_short = GnoWindow.calc_work_fluid(self, fluid_work_insert)

        block_gnvp_list = events_gnvp_frez(self, distance, fluid_work_insert)
        gnkt_work_firts = [
            [None, 'ЦЕЛЬ ПРОГРАММЫ', None, None, None, None, None, None, None, None, None, None, None],
            [None, 1,
             f'СПО промывочной КНК-1 с промывкой до МГРП {self.top_muft}. СПО фрезеровочной КНК-2: фрезерование '
             f'МГРП {self.top_muft}-{2}. '
             'Тех.отстой , замер Ризб. По доп.согласованию с Заказчиком, СПО промывочной КНК-1 до '
             'текущего забоя (МГРП №1).', None, None, None, None, None, None, None, None, None, None, None],
            [None, 2,
             'Внимание: Для проведения технологических операций завоз жидкости производить с ПНТЖ, '
             'согласованного с Заказчиком. Перед началом работ согласовать с Заказчиком пункт утилизации'
             ' жидкости.',
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'ПОРЯДОК ПРОВЕДЕНИЯ РАБОТ', None, None, None, None, None, None, None, None, None, None, None],
            [None, '№', 'Порядок работ', None, None, None, None, None, None, None, None,
             'Ответственный', None],
            [None, 1, 'Ознакомить бригаду с планом работ и режимными параметрами дизайна по промывке и СПО. '
                      'Провести инструктаж по промышленной безопасности',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 2, 'Принять скважину у Заказчика по акту (состояние ф/арматуры и кустовой площадки.)', None, None,
             None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 3,
             'Расставить оборудование и технику согласно «Типовой схемы расстановки оборудования и '
             'спецтехники при проведении капитального ремонта скважин с использованием '
             'установки «Койлтюбинг».',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 4,
             f'Произвести завоз технологической жидкости в объеме не менее 10м3 плотностью не более {fluid_work}'
             ' При интенсивном самоизливе скважины в процессе работ или при отрицательной температуре '
             'окружающего воздуха, только по доп.согласованию с Заказчиком, перейти на технологическую '
             'жидкость с удельным весом до 1,18г/см3.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ. Заказчик', None],
            [None, 6,
             'Внимание: при проведении работ по ОПЗ с кислотными составами, весь состав вахты обязан '
             'применять СИЗ (Инструкция П1-01.03 И-0128 ЮЛ-305 ООО"Башнефть-Добыча")',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 7,
             'Примечание: на месте проведения работ по ОПЗ кислотами и их смесями должен быть '
             'аварийный запас спецодежды, спецобуви и других средств индивидуальной защиты, запас '
             'чистой пресной воды и средств нейтрализации кислоты (мел, известь, хлорамин).',
             None, None, None, None, None, None, None, None, None, None],
            [None, 'Ограничения веса и скоростей при СПО', None, None, None, None, None, None, None, None, None,
             None, None],
            [None, 7,
             'Максимальный расчётный вес ГНКТ при подъёме с забоя – 4,2т; при спуске – 0,4т; в неподвижном состоянии '
             '- 2,4т. Максимальный допустимая нагрузка на ГНКТ - 18т.',
             None, None, None, None, None, None, None, None, None, None],
            [None, 8,
             f'Скорость спуска по интервалам:\nв устьевом оборудовании не более 0.5м/мин; \nв интервале 2- '
             f'{round(nkt_lenght - 20, 0)}м '
             f'не более 10-15м/мин - (первичный-последующий спуск); \n в интервале '
             f'{round(nkt_lenght - 20, 0)} - '
             f'{round(nkt_lenght + 20, 0)}м не более '
             f'2 м/мин;\n в интервале {round(nkt_lenght + 20, 0)}м - '
             f'{ports_data[self.top_muft]["кровля"] - 20}м не более 5-10 м/мин (фрез.КНК / промыв.КНК); в '
             f'интервале установки МГРП (± 20м) не более 2 м/мин; \nв интервале '
             f'{ports_data[self.bottom_muft]["кровля"] - 20}-{ports_data[self.bottom_muft]["кровля"]}м не более 2 м/мин;',
             None, None, None, None, None, None, None, None,
             'Мастер, бурильщик ГНКТ', None],
            [None, 9,
             f'Скорость подъёма по интервалам: \nв интервале {ports_data[self.bottom_muft]["кровля"]}-'
             f'{round(nkt_lenght + 20, 0)}м не более 10 м/мин; \n в интервале '
             f'установки МГРП (± 20м) не более 2 м/мин; \nв интервале '
             f'{round(nkt_lenght + 20, 0)}'
             f'-{round(nkt_lenght - 20, 0)}м не более 2 м/мин; \n в  '
             f'интервале {round(nkt_lenght - 20, 0)}-2м не более 12-15м/мин '
             f'(первичный-последующий подъем);\n в устьевом оборудовании '
             f'не более 0.5 м/мин.',
             None, None, None, None, None, None, None, None, 'Мастер, бурильщик ГНКТ', None],
            [None, 10,
             'При спуске производить приподъёмы для проверки веса на высоту не менее 20м со скоростью не более 5м/мин '
             'через каждые 300-500м (первичный-последующий спуск) в НКТ и 50-100м в ЭК.',
             None, None, None, None, None, None, None, None, 'Мастер, бурильщик ГНКТ', None],
            [None, 11, 'Перед каждой промывкой и после проверять веса ГТ (вверх, вниз, собств.)', None, None, None,
             None, None, None, None, None, 'Мастер, бурильщик ГНКТ', None],
            [None, 12,
             'При проведении технологического отстоя - не оставлять ГНКТ без движения - производить '
             'расхаживания г/трубы на 20м вверх и на 20м вниз со скоростью СПО не более 3м/мин. При '
             'отрицательной температуре окружающей среды, во избежании получения ледяной пробки в '
             'г/трубе при проведении тех.отстоя ни в коем случае не прекращать минимальную циркуляцию '
             'жидкости по г/трубе.',
             None, None, None, None, None, None, None, None, 'Мастер, бурильщик ГНКТ', None],
            [None, 13,
             'Не допускать увеличение нагрузки на г/трубу в процессе спуска. РАЗГРУЗКА Г/ТРУБЫ НЕ '
             'БОЛЕЕ 500 кг от собственного веса на этой глубине.',
             None, None, None, None, None, None, None, None,
             'Мастер, бурильщик ГНКТ', None],
            [None, 'Монтаж и опрессовка', None, None, None, None, None, None, None, None, None,
             None, None],
            [None, 14,
             'Собрать Компоновку Низа Колонны-1 далее КНК-1: коннектор + сдвоенный обратный клапан + '
             'насадка-промывочная Ø 38,1мм',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 15,
             f'Произвести монтаж 4-х секционного превентора БП 80-70.00.00.000 (700атм) и инжектора на устье '
             f'скважины согласно «Схемы обвязки №5 устья противовыбросовым оборудованием при производстве работ по '
             f'промывке скважины с установкой «ГНКТ» утвержденная главным инженером от '
             f'{well_data.dict_contractor[well_data.contractor]["Дата ПВО"]}г. Произвести обвязку установки ГНКТ, '
             f'насосно-компрессорного агрегата, желобной циркуляционной системы.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 16,
             'Внимание: Все требования ПБ и ОТ должны быть доведены до сведения работников, персонал должен быть '
             'проинформирован о начале проведения опрессовок. Все опрессовки производить согласно инструкции '
             'опрессовки ПВО и инструкции опрессовки нагнетательной и выкидной линии перед производством работ '
             'на скважине с Колтюбинговыми установками.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 17,
             'При отрицательной температуре окружающей среды, нагреть до 50ºC и прокачать по ГНКТ солевой раствор '
             'в объеме ГНКТ для предотвращения замерзания раствора внутри г/трубы (получения ледяной пробки).',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 18,
             f'При закрытой центральной задвижке фонтанной арматуры опрессовать ГНКТ и все нагнетательные '
             f'линии на 250атм. Опрессовать ПВО, обратные клапана и выкидную линию от устья скважины '
             f'до желобной ёмкости (надёжно закрепить, оборудовать дроссельными задвижками) опрессовать '
             f'на {well_data.max_admissible_pressure._value}атм с выдержкой 30мин. Результат опрессовки ПВО зафиксировать'
             f' в вахтовом журнале и '
             f'составить акт опрессовки ПВО. Установить на малом и большом затрубе технологический манометр. '
             f'Провести УТЗ и инструктаж. Опрессовку проводить в присутствии представителя ПФС, мастера, '
             f'бурильщика, машиниста подъемника и представителя супервайзерской службы. \nЗаявку на '
             f'представителя ПФЧ подавать за 24 часа телефонограммой. По окончании опрессовоки ПВО, '
             f'получить разрешения от представителя ПФС.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 'СПО промывочной КНК-1', None, None, None, None, None, None, None, None, None,
             None, None],
            [None, 19,
             f'Открыв скважину и записав число оборотов задвижки – зафиксировать дату и время.'
             f' Спустить КНК-1 в скважину с периодическими прокачками рабочей жидкостью (тех.вода 1,02г/см3)  '
             f'с проверкой веса на подъём через каждые 300м спуска до глубины '
             f'{round(nkt_lenght - 20, 0)}м.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 20,
             f'ВНИМАНИЕ: при получении посадки в НКТ в процессе спуска и наличии разгрузки на промывочный '
             f'инструмент более 500кг (уведомить Заказчика – составить АКТ на посадку). Приподнять КНК-1 на 20м '
             f'выше этой глубины.Произвести вывод НКА на рабочий режим, восстановить устойчивую циркуляцию промывочной'
             f' жидкости (тех.вода 1,02г/см3) , продолжить спуск до гл.'
             f'{round(nkt_lenght - 20, 0)}м с постоянным контролем промывочной '
             f'жидкости в обратной ёмкости на наличие мех. примесей. Скорость спуска при промывке НКТ не более '
             f'5м/мин. Контрольная проверка веса через каждые 100м промывки.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 21,
             f'На гл.{round(nkt_lenght + 20, 0)}м '
             f'произвести вывод НКА на рабочий режим, восстановить устойчивую циркуляцию '
             f'промывочной жидкости (тех.вода {fluid_work}), при необходимости произвести запуск и'
             ' вывод на режим МАК, получить стабильную круговую циркуляцию азотированной смеси. '
             'Промывка в течении 60мин с контролем на мех.примеси в обратной ёмкости.',
             None, None, None, None, None, None, None, None,
             None, None],
            [None, 22,
             f'ВНИМАНИЕ: В процессе промывки скважины - параметры азотированной промывочной смеси могут изменяться (от '
             f'80 до 200л/мин по жидкости (тех.вода {fluid_work}) и от 8 до 20м3/мин по азоту) в зависимости от качества '
             f'выноса посторонних частиц с забоя - данный процесс находиться под постоянным контролем у мастера по'
             f' сложным работам ГНКТ.',
             None, None, None, None,
             None, None, None, None,
             None, None],
            [None, 22,
             f'Произвести допуск КНК-1 с промывкой до МУФТЫ {self.top_muft} на гл.'
             f'{ports_data[self.top_muft]["кровля"]}-{ports_data[self.top_muft]["подошва"]}м.\nСкорость спуска при '
             f'промывке не более 5м/мин, проверка веса на подъём через каждые 30м.',
             None, None, None,
             None, None, None,
             None, None,
             'Мастер ГНКТ', None],
            [None, 23,
             f'При промывке, в случае выноса большого объёма проппанта из пласта (или в случае поглощения промывочной '
             f'жидкости) поинтервально через каждые 10м (или через каждые 2м) производить прокачку и  сопровождение '
             f'гелевой пачки объёмом 0,5-3м3 со скоростью 10 м/мин до гл.'
             f'{round(nkt_lenght - 20, 0)}м',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 24,
             f'При слабой циркуляции или аномальном поглощении (более 5м3/ч) промывочной жидкости '
             f'(тех.вода {fluid_work}) '
             f'в процессе промывки, уведомить Заказчика, приподнять КНК-1 до гл.'
             f'{round(nkt_lenght - 20, 0)}м восстановить стабильную круговой'
             f' циркуляции жидкости (тех.вода {fluid_work}). Допустить КНК-1 с циркуляцией (с контролем выхода на '
             'мех.примесей в смеси в обратной ёмкости) и продолжить промывку.',
             None, None, None, None, None, None, None,
             None, 'Мастер ГНКТ', None],
            [None, 25,
             'ВНИМАНИЕ: в процессе всего периода проведения работ на скважине при отсутствии проходки и получении жёс'
             'ткой посадки с разгрузкой более 500кг сверх собственного веса на данной глубине, по согласованию с '
             'Заказчиком произвести ОБСЛЕДОВАНИЕ ТЕКУЩЕГО ЗАБОЯ спуском торцевой печати на ГНКТ (Dпечати -согласовать '
             'с Заказчиком). Получить отпечаток разгрузкой на ГНКТ в 1000кг. Поднять печать из скважины. Дальнейшие '
             'работы по результатам обследования печати.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 26,
             'При отсутствии проходки и получения жесткой посадки, дальнейшие работы по согласованию с Заказчиком.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ представитель Заказчика', None],
            [None, 27,
             f'При достижении гл.{ports_data[self.top_muft]["кровля"]}м произвести промывку в следующем '
             f'порядке:\n- прокачать гелевую пачку в объеме '
             f'2-3м3;\n- промыть скважину в течении 120 минут до выхода чистой, без посторонних примесей, промывочной '
             f'жидкости (тех.вода {fluid_work}). Составить акт.',
             None, None, None,
             None, None, None,
             None, None,
             'Мастер ГНКТ', None],
            [None, 28,
             'Поднять КНК-1 на ГНКТ из скважины, закрыв скважину и записав число оборотов задвижки – '
             'зафиксировать дату и время. Демонтировать превентор, лубрикатор, КНК-1.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ представитель Заказчика', None],
            [None, 'Спуск фрезеровочной КНК-2. Фрезерование муфт ГРП (фрак-портов)', None, None, None, None, None, None,
             None, None, None, None, None],
            [None, 29,
             'Собрать фрезеровочную Компоновку Низа Колонны-2, далее КНК-2: наружный коннектор Ø 54мм + '
             'обратный клапан створчатого типа 54мм + гидравлический разъединитель 57мм + ВЗД Ø 54-55мм +'
             ' торцевой фрез Ø 68мм. Постоянно после сборки компоновки, проверять работоспособность ВЗД перед спуском '
             'в скважину на устье. Произвести замеры составных частей КНК с записью в журнале. Произвести монтаж '
             'лубрикатора и инжектора на устье скважины. Произвести необходимые опрессовки.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 30,
             'Открыв скважину и записав число оборотов задвижки – зафиксировать дату и время. Спустить КНК-2 в '
             f'скважину с периодическими прокачками рабочей жидкостью (тех.вода {fluid_work})  с проверкой веса на '
             f'подъём через каждые 300м спуска до глубины {ports_data[self.top_muft]["кровля"] - 20}м. '
             f'Убедиться в наличии свободного прохода по лифту НКТ.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 31,
             f'При получении посадки в НКТ и отсутствии прохода КНК-2 до гл.{ports_data[self.top_muft]["кровля"] - 20}м,'
             f' приподнять КНК-2 на 20м выше '
             f'глубины посадки. Вывести НКА на рабочий режим в соответствии с рабочими параметрами ВЗД. Произвести '
             'проработку (проходного сечения НКТ) места посадки до получения свободного прохода в НКТ с составлением '
             'АКТа.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 32,
             f'При свободном и беспрепятственном прохождении КНК-2 на г/трубе в НКТ до гл.'
             f'{ports_data[self.top_muft]["кровля"] - 20}м, продолжить '
             f'доспуск КНК-2 с минимальной подачей  ВЗД до {self.top_muft} до получения посадки на гл.'
             f'{ports_data[self.top_muft]["кровля"] - 20}м. '
             'Установить метку на г/трубе.',
             None, None, None, None, None, None, None,
             None, 'Мастер ГНКТ', None]]

        frez_mufts = []
        count_muft = -2
        for muft, muft_data in sorted(ports_data.items(), reverse=True)[:-1]:
            frez_muft = [
                [None,
                 f'ФРЕЗЕРОВАНИЕ МУФТЫ {muft}',
                 None, None, None, None, None, None, None, None, None,
                 'Мастер ГНКТ', None],
                [None, 33,
                 f'После соприкосновения с МУФТОЙ {muft} приподнять КНК-2 на 10м выше. Проверить вес ГНКТ и '
                 'давление циркуляции - эти значения будут ориентиром во время работы в случае заклинивания ВЗД '
                 'и закупорки насадки. Вывести НКА на рабочий режим в соответствии с рабочими параметрами ВЗД.',
                 None, None, None,
                 None, None, None,
                 None, None,
                 'Мастер ГНКТ', None],
                [None, 34,
                 'Внимание: рабочее давление на устье в процессе разбуривания не должно превышать 100атм. '
                 'Если циркуляционное давление выше 250атм, произвести закачку понизителя трения в концентрации 3-5л/1м3.',
                 None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
                [None, 35,
                 f'Допустить КНК-2 с циркуляцией, и с гл.{muft_data["кровля"]}м произвести фрезерование посадочного седла '
                 f'МУФТЫ {muft} до  гл.{muft_data["подошва"]} до снижения рабочего давления и получения провала. '
                 f'Следить за устьевым давлением и '
                 f'постоянно контролировать выходящую из скважины жидкость на наличие мех.примесей.',
                 None, None, None, None, None, None, None,
                 None, 'Мастер ГНКТ', None],
                [None, 36,
                 f'ВНИМАНИЕ: при слабой циркуляции или аномальном поглощении (более 5м3/ч) промывочной жидкости (тех.вода '
                 f'{fluid_work})  в процессе фрезерования, уведомить Заказчика, приподнять КНК-2 до гл.'
                 f'{round(nkt_lenght - 20, 0)}м восстановить '
                 f'стабильную циркуляцию и допустить КНК-2 до МГРП продолжить работы по фрезерованию.',
                 None, None, None,
                 None, None, None,
                 None, None,
                 'Мастер ГНКТ', None],
                [None, 37,
                 f'После окончания фрезерования МУФТЫ {muft} ({muft_data["кровля"]}-{muft_data["подошва"]}м) и получения '
                 f'прохода КНК-2 ниже глубины '
                 f'{muft_data["подошва"]}м и возвращение веса к нормальным значениям (снижения рабочего давления и '
                 f'получения прохода ГНКТ),'
                 f' при необходимости прокачать на циркуляцию по г/трубе вязкую пачку в объеме 1м3. Проработать интервал'
                 f' МУФТУ {muft} три раза с выходом 5 метров ниже и выше. Минимизировать нахождение фрезы за интервалом '
                 f'разбуривания.',
                 None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
                [None, 38,
                 f'После окончания проработки интервала МУФТЫ {muft}, произвести допуск КНК-2 на г/трубе с циркуляцией до '
                 f'следующей {list(ports_data.keys())[count_muft]}; уведомить Заказчика, составить АКТ на посадку. ',
                 None, None, None, None, None, None, None,
                 None, 'Мастер ГНКТ', None],
                [None, 39,
                 'Внимание: при отсутствии проходки вследствии предполагаемого износа фреза, произвести смену '
                 'вооружения: поднять фрез.КНК, заменить фрез, спустить фрез.КНК, продолжить работы по'
                 ' фрезерованию седел муфт ГРП.',
                 None, None, None,
                 None, None, None,
                 None, None,
                 'Мастер ГНКТ', None],
                [None, 40,
                 'Внимание: при отсутствии свободного и беспрепятственного прохода КНК-2 до следующей муфты'
                 ' по согласованию с Заказчиком, произвести  промежутучную промывку на '
                 'промывочной КНК-1 до следующей муфты',
                 None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None]]
            count_muft -= 1
            frez_mufts.extend(frez_muft)

        flushing_list = [
            [None,
             'По согласованию с Заказчиком, проведение процедуры промежуточной промывки:\n СПО '
             'промывочной КНК-1, промывка до МГРП ',
             None, None, None, None, None, None, None,
             None, None, None, None],
            [None, 41,
             f'Поднять КНК-2 на г/трубе из скважины. Закрыть коренную задвижку. Демонтировать инжектор,'
             f' лубрикатор, КНК-2 (ВЗД с т/ф). Собрать КНК-1 (насадка промывочная Ø38.1мм + сдвоенный '
             f'обратный клапан). Произвести монтаж лубрикатора и инжектора на устье скважины.Произвести необходимые '
             f'опрессовки. Открыть скважину. Спустить КНК-1 в скважину с периодическими прокачками рабочей жидкостью '
             f'(тех.вода {fluid_work}) с проверкой веса на подъём через каждые 500м спуска до гл.'
             f'{round(nkt_lenght - 20, 0)}м. Вывести НКА на'
             f' рабочий режим промывки и получить стабильную круговую циркуляцию промывочной жидкости (тех.вода'
             f' {fluid_work}) произвести запуск азотного комплекса, вывести его на рабочий режим.Дождаться выхода пузыря '
             f'азота. Получить стабильную круговуюциркуляцию азотированной смеси. Доспустить КНК-1 с циркуляцией на '
             f'азотированной смеси до глубины непрохода КНК-2 и произвести промывку скважины до '
             f'{self.top_muft} - {self.bottom_muft}'
             f' до получения жесткой посадки.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 42,
             f'При достижении {self.top_muft} - {self.bottom_muft} произвести промывку:\n- прокачать на циркуляцию по '
             f'г/трубе вязкую пачку в V=2-3м3;\n- произвести промывку в течении не менее 2 часов, до чистой, '
             f'без посторонних мех. примесей промывочной жидкости (тех.вода {fluid_work}).\nСоставить акт на '
             'нормализацию в присутствии представителя Заказчика.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ представитель заказчика', None],
            [None, 43,
             f'Поднять КНК-1 на ГНКТ из скважины. Закрыть коренную задвижку. Сменить промывочную КНК-1 на '
             f'фрезеровочную КНК-2. Продолжить работы по фрезерованию МГРП.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 'Подъем фрезеровочной КНК-2', None, None, None, None, None, None, None, None, None, None, None],
            [None, 1,
             'ВНИМАНИЕ БУРИЛЬЩИК! ПОСТОЯННО!!! При подъеме ВЗД после фрезерования седел и шаров МГРП,'
             ' во избежание заклинивания и получения прихвата ГНКТ (от возможного попадания остатков частиц шара'
             ' или седла после разбуривания ) остановить г/трубу не доходя 50м до воронки и прокачать малый затруб '
             f'тех.жидкостью (тех.вода {fluid_work}) в объеме не менее 2х объемов НКТ.',
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 44,
             f'После окончания проработки {list(ports_data.keys())[-2]} от забоя" поднять КНК-2 до гл.'
             f'{nkt_lenght - 20}м.\nПроизвести тех.отстой в '
             'течении 2-х часов для замера Ризб на тех.воде. Пересчитать забойное давление и необходимый удельный '
             'вес жидкости глушения. По доп.согласованию с Заказчиком, произвести СПО пром.КНК-1 с целью глушения '
             'скважины -  выполнение ',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 45,
             'После тех.отстоя произвести подъем КНК-2 на г/трубе из скважины соблюдая скорости безопасного СПО.'
             ' Закрыть коренную задвижку.',
             None, None, None, None, None, None, None,
             None, 'Мастер ГНКТ', None],
            [None, 46,
             'Демонтировать превентор, лубрикатор, КНК-2 (ВЗД с т/ф). Обрезать 1 метр ГНКТ после СПО '
             'фрезеровочной КНК.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 'Промывку произвести по доп. согласованию с Заказчиком.', None, None, None, None, None, None, None,
             None, None, None, None],
            [None, 'Спуск промывочной КНК-1', None, None, None, None, None, None, None, None, None, None, None],
            [None, 47,
             'Собрать промывочную КНК-1: коннектор + сдвоенный обратный клапан + насадка промывочная Ø 38,1мм. '
             'Произвести монтаж лубрикатора и инжектора на устье скважины. Произвести необходимые опрессовки.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 48,
             'Открыв скважину и записав число оборотов задвижки – зафиксировать дату и время. Спустить КНК-1 в '
             f'скважину до гл.{nkt_lenght - 20}м с ПЕРИОДИЧЕСКОЙ прокачкой рабочей жидкостью (тех.вода '
             f'{fluid_work}) и проверкой'
             ' веса на подъём. Убедится в наличии свободного прохода КНК-1 по НКТ.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 49,
             'Произвести запуск и вывести Азотный комплекс и НКА на рабочий режим. Получить стабильную круговую '
             'циркуляцию азотированной смеси, промывка в течении 60мин с контролем на мех.примеси в обратной ёмкости.',
             None, None, None, None, None, None, None,
             None, 'Мастер ГНКТ', None],
            [None, 49,
             f'Расчетные параметры циркуляции: по жидкости (тех.вода {fluid_work}) 120л/мин; 10м3/мин по азоту.\nВ '
             'процессе промывки скважины, параметры азотированной промывочной смеси могут изменяться (от 80 до '
             '200л/мин по жидкости и от 8 до 20м3/мин по азоту) в зависимости от качества выноса посторонних '
             'частиц с забоя. данный процесс находится под постоянным контролем у мастера ГНКТ.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None],
            [None, 50,
             f'Произвести допуск КНК-1 с промывкой на азотированной смеси до текущего забоя на гл. муфты №'
             f'{self.bottom_muft} (при '
             'отсутствии проходки согласовать достигнутый забой с Заказчиком)',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 51,
             'При необходимости, при промывке производить сопровождение вымытой пачки со скоростью 2-3м/мин до '
             f'глубины {nkt_lenght - 20}м. Промывку производить до выхода чистой тех. жидкости (тех.вода '
             f'{fluid_work}) и только '
             'после этого продолжать промывку.',
             None, None, None, None, None, None, None,
             None, 'Мастер ГНКТ', None],
            [None, 52,
             f'При достижении глубины МУФТЫ {self.bottom_muft} (или согласованного забоя) произвести промывку в следующем '
             'порядке:\n- прокачать гелевую пачку в объеме 2-3м3;\n- промыть скважину в течении 2 часов до выхода '
             f'чистой, без посторонних примесей, промывочной жидкости (тех.вода {fluid_work}).Составить Акт на промывку '
             'в присутствии представителя Заказчика.',
             None, None, None,
             None, None, None,
             None, None,
             'Мастер ГНКТ', None],
            [None, 'По согласованию с Заказчиком, подтверждение нормализованного забоя', None, None, None, None, None,
             None, None, None, None, None, None],
            [None, 53,
             f'Приподнять КНК-1 на ГНКТ не прекращая циркуляции до гл.{nkt_lenght - 20}м. Убедиться в отсутствии мех. '
             f'примесей в '
             f'промывочной жидкости (тех.вода {fluid_work}) , остановить подачу жидкости НКА и ПАУ.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 54, 'Произвести тех.отстой скважины для оседания твёрдых частиц в течении 2х часов.', None, None,
             None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 55,
             f'После технологического отстоя допустить КНК-1 на г/трубе в скважину «без циркуляции» до гл.'
             f'{self.bottom_muft}м, '
             'забой должен соответствовать ранее нормализованному. Составить АКТ с представителем Заказчика. При '
             'отсутствии ранее нормализованного забоя по согл. с Заказчиком, провести работы по нормализации забоя.',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ представитель Заказчика', None],
            [None, 'Подъем промывочной КНК-1', None, None, None, None, None, None, None, None, None, None, None],
            [None, 56,
             self.jamming_well_str(ntk_true, paker_true),
             None, None,
             None, None,
             None, None,
             None, None,
             'Мастер ГНКТ представитель Заказчика', None],
            [None, 57,
             'Извлечь КНК-1 на ГНКТ из скважины. Закрыть скважину записав и сверив число оборотов задвижки – '
             'зафиксировать дату и время.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ представитель Заказчика', None],
            [None, 'ДЕМОНТАЖ И ОСВОБОЖДЕНИЕ ТЕРРИТОРИИ', None, None, None, None, None, None, None, None, None, None,
             None],
            [None, 58, 'После закрытия задвижки - отдуть г/трубу азотом.', None, None, None, None, None, None, None,
             None, 'Мастер ГНКТ', None],
            [None, 59,
             'Произвести демонтаж превентора и инжектора, установки ГНКТ. Очистить желобные ёмкости от проппанта в '
             'мешки – приготовить к вывозу. Составить Акт на количество вымытого проппанта. Произвести демонтаж '
             'рабочих линий, рабочей площадки.\nВнимание: произвести вывоз отработанной технологической жидкости'
             ' и мешки с вымытым проппантом на пункт(ы) утилизации, согласованный с Заказчиком.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 60, 'Сдать скважину представителю Заказчика Составить АКТ.', None, None, None, None, None, None,
             None, None, 'Мастер ГНКТ', None],
            [None, 'Контроль выхода малого затруба', None, None, None, None, None, None, None, None, None, None, None],
            [None, 61,
             'Во время промывки - выход малого затруба постоянно должен находиться под контролем. На желобной ёмкости п'
             'остоянно осуществляется наблюдение за наличием проппанта и мех. примесей на выходной линии. \nПеред'
             ' началом промывки – необходимо отрегулировать штуцерный монифольд так, как это необходимо – уровень'
             ' промывочной жидкости в циркуляционной ёмкости не должен уменьшаться. Уровень жидкости должен находиться'
             ' под постоянным наблюдением, чтобы избежать потери жидкости в пласт. Во время промывки уровень жидкости '
             'должен немного увеличиваться или оставаться неизменным.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 'Действия при приватах ГНКТ.', None, None, None, None, None, None, None, None, None, None, None],
            [None, 62,
             f'ВНИМАНИЕ: При наличии посадок КНК - спуск производить с остановками для промежуточных промывок. В случае '
             f'прихвата ГНКТ в скважине - проинформировсть ответственного представителя Заказчика и руководство ГНКТ'
             f'{well_data.contractor}. Дальнейшие действия производить в присутствии представителя Заказчика с '
             f'составлением АКТа '
             f'согласно "Плана-Схемы действий при прихватах ГНКТ" ТЕХНОЛОГИЧЕСКОЙ ИНСТРУКЦИИ {well_data.costumer}',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ, предст.Заказчика Мастер по сложным работам ГНКТ', None],
            [None, 'Использование хим. реагентов в процессе работ', None, None, None, None, None, None, None, None,
             None, None, None, None],
            [None, 63,
             'а) Во время промывки возможен резкий вынос большого объёма проппанта из пласта, что может привести к'
             ' потере циркуляции и последующему прихвату ГНКТ, данную ситуацию можно проследить, при этом вес ГНКТ '
             'резко понизится, а циркуляционное давление начнёт повышаться, в данном случае необходимо приостановить '
             'спуск ГНКТ, произвести промывку с добавлением понизителя трения гидравлического давления '
             '(дозировка до 3-5л /1м3 в зависимости от применяемого вида) до стабилизации рабочего давления, '
             f'после чего продолжить промывку.\nб) В случае поглащения промывочной жидкости (тех.вода {fluid_work}) '
             'в процессе промывки, после взятия каждой пачки проппанта производить прокачку загеленной жидкости'
             ' (вязких пачек) в объеме 2-4м3 с сопровождением пачек в НКТ до гл.стингера с последующей промывкой '
             'до полного выноса проппанта на желобную ёмкость.\nв) При наличии посадок КНК, спуск производить с '
             'остановками для промежуточных промывок.\nг) В случае использования мембранной азотной установки, для '
             'уменьшения коррозионного влияния кислорода на ГНКТ, приготовить промывочную жидкость с добавлением'
             f' ингибитора коррозии в расчете 120 л на 25м3 жидкости (тех.вода {fluid_work}).\nд) При выходе густого '
             'высоковязкого геля (во избежании закупорки циркуляционной системы) использовать диструктор - '
             'лимонную кислоту в жидком виде.',
             None, None, None, None, None, None, None, None, 'Мастер ГНКТ', None],
            [None, 64,
             'После закрытия задвижки, приготовить и прокачать по г/трубе по циркуляции на желобную ёмкость пачку – '
             'ингибитора коррозии в объёме 40л, с целью предотвращнения коррозийных отложений в г/трубе. '
             'Предположительный расход хим.реагентов на скважину: 1) Понизитель трения Лубритал - 30л '
             '(концентрация 1л/м3); 2) Загуститель ВГ-4 - 20л (для загеливания тех.жидкости и прокачки '
             'вязких пачек концентрация 5кг/м3)',
             None, None, None, None, None, None, None, None,
             'Мастер ГНКТ', None]]

        gnkt_work_list = []

        for row in block_gnvp_list:
            gnkt_work_list.append(row)

        for row in gnkt_work_firts:
            gnkt_work_list.append(row)

        for row in frez_mufts:
            gnkt_work_list.append(row)

        for row in flushing_list:
            gnkt_work_list.append(row)

        # for row in range(1, len(krs_begin_gnkt) + 1):  # Добавлением работ
        #     # print(row, len(schema_well_list[row-1]), schema_well_list[row-1][15])
        #     for col in range(1, 48):
        #         cell = ws4.cell(row=row, column=col)
        #
        #         cell.value = krs_begin_gnkt[row - 1][col - 1]
        #         ws4.cell(row=row, column=col).font = Font(name='Arial', size=11, bold=False)
        #         ws4.cell(row=row, column=col).alignment = Alignment(wrap_text=True, horizontal = 'center',
        #                                                                            vertical = 'center')
        #         if cell.value != None and row > 24:
        #             cell.border = border
        return gnkt_work_list
    def jamming_well_str(self, ntk_true, paker_true):
        if ntk_true is True and paker_true is True:

            jamming_well = f'Произвести подъем с замещением скважинной жидкости на раствор глушения, удельного веса по согласованию ' \
                             f'с Заказчиком, рассчитанного по замеру Ризб после 2-х часов отстоя и удел.веса рабочей жидкости в скважин, '\
                 f'но не менее удельного веса расчитанного для пластового давления указанного в настоящем плане работ '\
                 f'{self.zhgs} (при Рпл={self.pressuar}атм).  До завоза раствора, '\
                 f'скважину разряжать. Перед замещением КНК установить '\
                 f'в интервале нижнего фрак-порта.\nПрокачать на циркуляцию жидкость глушения в объеме не менее '\
                 f'{self.volume_dumping(ntk_true, paker_true, self.bottom_muft)}м3 '\
                 f'(трубного пространства) с одновременным подъемом ГНКТ (с протяжкой ГНКТ перевести хвостовик). '\
                 f'В процессе перевода соблюдать равенство объемов закаченной и отобранной из скважины жидкости, '\
                 f'т.е. не допускать режима фонтанирования (поглощения).'
        elif ntk_true is True and paker_true is False:
            # print(f'объем скважины {well_volume(self, well_data.perforation_sole)}')
            # print(f'объем металла {volume_nkt_metal(well_data.dict_nkt)}')
            # print(f'объем внутренний НКТ {volume_nkt(well_data.dict_nkt)}')
            volume_first = round((well_volume(self, well_data.perforation_sole) -volume_nkt_metal(well_data.dict_nkt) -volume_nkt(well_data.dict_nkt)) *1.2, 1)
            # print(volume_first)

            jamming_well = f'Произвести замер избыточного давления в течении 2ч при условии заполнения ствола ствола ' \
                         f'жидкостью уд.весом 1.01г/см3. Произвести перерасчет забойного давления, Согласовать с ' \
                         f'заказчиком глушение скважин и необходимый удельный вес жидкости глушения,  но не менее ' \
                           f'удельного веса расчитанного для пластового ' \
                         f'давления указанного в настоящем плане работ {self.zhgs} (при Рпл={self.pressuar}атм). ' \
                           f'Допустить КНК до нижнего фрак-порта.' \
                         f'До завоза раствора, скважину разряжать. При достаточной вязкости раствора  предусмотреть ' \
                           f'работу без обратного клапана'\
                           f'Произвести перевод на тех жидкость расчетного удельного ' \
                         f'веса в объеме {volume_first}м3 (объем подпакерного пространства + затруб + 20% запас), ' \
                         f'вывести циркуляцию с большого затруба  с ПРОТЯЖКОЙ ГНКТ СНИЗУ ВВЕРХ  с выходом ' \
                         f'циркуляции по большому затрубу до башмака НКТ до гл. {well_data.depth_fond_paker_do["do"]}м ' \
                         f'В башмаке НКТ промыть до выхода жидкости глушения по малому затрубу в объеме ' \
                         f'{round(volume_nkt(well_data.dict_nkt),1)}м3 с одновременным подъемом ГНКТ. Тех отстой 2ч.' \
                         f' В случае отрицательного результата по глушению скважины произвести перерасчет ЖГС и ' \
                         f'повторить операцию. В процессе перевода соблюдать равенство объемов закаченной и ' \
                           f'отобранной из скважины жидкости, '\
                        f'т.е. не допускать режима фонтанирования (поглощения).'

        return jamming_well
    def volume_dumping(self, ntk_true, first_muft):
        from work_py.alone_oreration import volume_pod_NKT, volume_jamming_well

        if ntk_true is True:
            volume = volume_pod_NKT(self) * 1.2
        else:
            volume = volume_jamming_well(self, first_muft) * 1.1
        return round(volume, 1)

    def date_dmy(self, date_str):
        date_obj = date_str
        # print(date_obj)
        # print(date_str)

        if isinstance(date_obj, datetime):
            return date_obj.strftime('%d.%m.%Y')
        else:
            date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M:%S")
            # print(f' даь {date_obj}')
        return date_obj.strftime('%d.%m.%Y')

    def insert_ports_data(self, ports_data):

        ports_list = [
            [None, None, None, None, None, None, 'Интервалы установки фрак-портов  (муфт ГРП)', None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None,
             None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],

            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None],
            [None, 'Цель работ', None, None, None, None, None, None,
             f'СПО промывочной КНК-1 с промывкой до МГРП №{self.top_muft}. СПО фрезеровочной КНК-2: фрезерование '
             f'МГРП №{self.top_muft}-№2.'
             f' Тех.отстой , замер Ризб. По доп.согласованию с Заказчиком, СПО промывочной КНК-1 до '
             f'текущего забоя (МГРП №1).',
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None],
            [None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None, None,
             None, None, None, None, None, None, None, None, None, None, None]
        ]
        port_len = len(ports_data)
        col_port = int(36 / port_len)
        for index_row, row in enumerate(ports_list):
            # dict_ports[f'Муфта №{index + 1}'] = {'кровля': port[0], 'подошва': port[1], 'шар': ball, 'седло': saddle,
            #                                      'тип': type_saddles}
            # print(f'порты {ports_data}')
            col = 45
            n = 3
            m = 0
            merge_port = {}

            for index, port in enumerate(ports_data):
                if index_row == 1:
                    ports_list[index_row][col - n - 2] = port
                elif index_row == 2:
                    ports_list[index_row][col - n - 2] = ports_data[port]['тип']
                elif index_row == 3:
                    ports_list[index_row][col - n - 2] = f'{ports_data[port]["подошва"]}м'
                    ports_list[index_row][col - n - 1] = 'Ø седла'
                    ports_list[index_row][col - n] = 'Ø шара'
                elif index_row == 4:
                    ports_list[index_row][col - n - 2] = f'{ports_data[port]["кровля"]}-'
                    ports_list[index_row][col - n - 1] = f'{ports_data[port]["седло"]}'
                    ports_list[index_row][col - n] = f'{ports_data[port]["шар"]}мм'
                # print(col - col_port - n)
                merge_port[182 + m] = (col - n - 1, 25, col - n + 1, 25)
                merge_port[182 + m + 1] = (col - n - 1, 26, col - n + 1, 26)
                n += col_port
                m += 2


        return ports_list, merge_port

    def calc_fluid(self):

        fluid_list = []
        try:

            fluid_p = 0.83
            for plast in well_data.plast_work:
                if float(list(well_data.dict_perforation[plast]['рабочая жидкость'])[0]) > fluid_p:
                    fluid_p = list(well_data.dict_perforation[plast]['рабочая жидкость'])[0]
            fluid_list.append(fluid_p)

            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите расчетный удельный вес жидкости глушения в '
                                                           'конце ремонта',
                                                           max(fluid_list), 0.87, 2, 2)
        except:
            fluid_work_insert, ok = QInputDialog.getDouble(self, 'Рабочая жидкость',
                                                           'Введите удельный вес рабочей жидкости',
                                                           0, 0.87, 2, 2)
        return fluid_work_insert


if __name__ == '__main__':
    app = QApplication([])
    window = Work_with_gnkt()
    window.show()
    app.exec_()
