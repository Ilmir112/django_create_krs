import well_data
from datetime import datetime
from PyQt5.QtWidgets import QInputDialog, QMessageBox, QMainWindow
from openpyxl_image_loader import SheetImageLoader
from openpyxl.utils.cell import get_column_letter
from openpyxl.styles import Font, Alignment
from main import ExcelWorker

from cdng import events_gnvp, itog_1, events_gnvp_gnkt
from find import ProtectedIsNonNone
from plan import delete_rows_pz
from block_name import region, razdel_1, curator_sel, pop_down
from work_py.dop_plan_py import DopPlanWindow


class CreatePZ(QMainWindow):
    def __init__(self, wb, ws, data_window, perforation_correct_window2, parent=None):
        super(CreatePZ, self).__init__(parent)

        self.wb = wb
        self.ws = ws
        self.data_window = data_window
        self.perforation_correct_window2 = perforation_correct_window2

    def open_excel_file(self, ws, work_plan):
        from find import FindIndexPZ
        from category_correct import CategoryWindow
        from main import MyWindow
        from find import WellNkt, Well_perforation, WellCondition, WellHistory_data, Well_data, Well_Category, \
            WellFond_data, WellSucker_rod, Well_expected_pick_up

        well_data.work_plan = work_plan

        well_data.dict_category = CategoryWindow.dict_category

        # Запуск основного класса и всех дочерних классов в одной строке
        well_pz = FindIndexPZ(ws)
        # well_pz.read_pz(ws)

        well_data.region = region(well_data.cdng._value)
        WellNkt.read_well(self, ws, well_data.pipes_ind._value, well_data.condition_of_wells._value)
        if well_data.work_plan not in ['application_pvr', 'application_gis']:
            WellSucker_rod.read_well(self, ws, well_data.sucker_rod_ind._value, well_data.pipes_ind._value)
            WellFond_data.read_well(self, ws, well_data.data_fond_min._value, well_data.condition_of_wells._value)
        WellHistory_data.read_well(self, ws, well_data.data_pvr_max._value, well_data.data_fond_min._value)
        WellCondition.read_well(self, ws, well_data.condition_of_wells._value, well_data.data_well_max._value)

        Well_expected_pick_up.read_well(self, ws, well_data.data_x_min._value, well_data.data_x_max._value)
        Well_data.read_well(self, ws, well_data.cat_well_max._value, well_data.data_pvr_min._value)

        well_data.region = region(well_data.cdng._value)
        if work_plan == 'dop_plan':
            number_list = list(map(str, range(1, 50)))

            well_data.number_dp, ok = QInputDialog.getItem(self, 'Номер дополнительного плана работ',
                                                           'Введите номер дополнительного плана работ',
                                                           number_list, 0, False)

        Well_perforation.read_well(self, ws, well_data.data_pvr_min._value, well_data.data_pvr_max._value + 1)
        Well_Category.read_well(self, ws, well_data.cat_well_min._value, well_data.data_well_min._value)

        if work_plan == 'plan_change':
            DopPlanWindow.extraction_data(self)
            ws.delete_rows(well_data.plan_correct_index._value, ws.max_row)
            return ws

        elif work_plan not in ['application_pvr', 'application_gis']:
            if work_plan != 'plan_change':
                for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
                    ws.row_dimensions[row_ind].hidden = False

                    if any(['ПЛАН РАБОТ' in str(col).upper() for col in row]) \
                            and work_plan == 'dop_plan':
                        ws.cell(row=row_ind + 1, column=2).value = f'ДОПОЛНИТЕЛЬНЫЙ ПЛАН РАБОТ № {well_data.number_dp}'
                        print(f'номер доп плана {well_data.number_dp}')

                    if 'План-заказ' in row:
                        # print(row)

                        ws.cell(row=row_ind + 1, column=2).value = 'ПЛАН РАБОТ'

                    for col, value in enumerate(row):
                        if not value is None and col <= 12:
                            if 'гипс' in str(value).lower() or 'гидратн' in str(value).lower():
                                well_data.gipsInWell = True

                if well_data.emergency_well is True:
                    emergency_quest = QMessageBox.question(self, 'Аварийные работы ',
                                                           'Программа определела что в скважине'
                                                           f'авария - {well_data.emergency_count}, верно ли?')
                    if emergency_quest == QMessageBox.StandardButton.Yes:
                        well_data.emergency_well = True
                        well_data.emergency_bottom, ok = QInputDialog.getInt(self, 'Аварийный забой',
                                                                         'Введите глубину аварийного забоя',
                                                                         0, 0, int(well_data.bottomhole_artificial._value))
                    else:
                        well_data.emergency_well = False
                if well_data.problemWithEk is True:
                    problemWithEk_quest = QMessageBox.question(self, 'ВНИМАНИЕ НЕПРОХОД ',
                                                               f'Программа определела что в скважине '
                                                               f'ссужение в ЭК -, верно ли?')
                    if problemWithEk_quest == QMessageBox.StandardButton.Yes:
                        well_data.problemWithEk = True
                        well_data.problemWithEk_depth, ok = QInputDialog.getInt(self, 'Глубина сужения',
                                                                                "ВВедите глубину cсужения", 0, 0,
                                                                                int(well_data.current_bottom))
                        well_data.problemWithEk_diametr = QInputDialog.getInt(self, 'диаметр внутренний cсужения',
                                                                              "ВВедите внутренний диаметр cсужения", 0, 0,
                                                                              int(well_data.current_bottom))[0]
                    else:
                        well_data.problemWithEk = ProtectedIsNonNone(False)

                if well_data.gipsInWell is True:
                    gips_true_quest = QMessageBox.question(self, 'Гипсовые отложения',
                                                           'Программа определела что скважина осложнена гипсовыми отложениями '
                                                           'и требуется предварительно определить забой на НКТ, верно ли это?')

                    if gips_true_quest == QMessageBox.StandardButton.Yes:
                        well_data.gipsInWell = True
                    else:
                        well_data.gipsInWell = False

            try:
                # Копирование изображения
                image_loader = SheetImageLoader(ws)
            except:
                mes = QMessageBox.warning(None, 'Ошибка', 'Ошибка в копировании изображений')

            for row in range(1, well_data.data_well_max._value):
                for col in range(1, 12):
                    try:
                        image = image_loader.get(f'{get_column_letter(col)}{row}')
                        image.save(f'_internal/imageFiles/image_work/image{get_column_letter(col)}{row}.png')
                        image_size = image.size
                        image_path = f'_internal/imageFiles/image_work/image{get_column_letter(col)}{row}.png'

                        coord = f'{get_column_letter(col)}{row + 17 - well_data.cat_well_min._value}'

                        well_data.image_list.append((image_path, coord, image_size))

                    except:
                        pass
            if work_plan != 'plan_change':
                for j in range(well_data.data_x_min._value,
                               well_data.data_x_max._value):  # Ожидаемые показатели после ремонта
                    lst = []
                    for i in range(0, 12):
                        lst.append(ws.cell(row=j + 1, column=i + 1).value)
                    well_data.row_expected.append(lst)

            if well_data.work_plan not in ['gnkt_frez', 'application_pvr',
                                           'application_gis', 'gnkt_after_grp', 'gnkt_opz', 'plan_change']:
                # print(f'план работ {well_data.work_plan}')
                delete_rows_pz(self, ws)
                razdel = razdel_1(self, well_data.region)

                for i in range(1, len(razdel)):  # Добавлением подписантов на вверху
                    for j in range(1, 13):
                        ws.cell(row=i, column=j).value = razdel[i - 1][j - 1]
                        ws.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=False)
                    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=7)
                    ws.merge_cells(start_row=i, start_column=8, end_row=i, end_column=13)
                well_data.ins_ind = 0

                well_data.ins_ind += well_data.data_well_max._value - well_data.cat_well_min._value + 19
                # print(f' индекс вставки ГНВП{well_data.ins_ind}')
                dict_events_gnvp = {}
                dict_events_gnvp['krs'] = events_gnvp()
                dict_events_gnvp['gnkt_opz'] = events_gnvp_gnkt()
                dict_events_gnvp['dop_plan'] = events_gnvp()
                # if work_plan != 'dop_plan':
                text_width_dict = {20: (0, 100), 30: (101, 200), 40: (201, 300), 60: (301, 400), 70: (401, 500),
                                   90: (501, 600), 110: (601, 700), 120: (701, 800), 130: (801, 900),
                                   150: (901, 1500), 270: (1500, 2300)}

                for i in range(well_data.ins_ind, well_data.ins_ind + len(dict_events_gnvp[work_plan]) - 1):
                    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=12)
                    data = ws.cell(row=i, column=2)
                    data.value = dict_events_gnvp[work_plan][i - well_data.ins_ind][1]

                    if 'Мероприятия' in str(data.value) or \
                            'Меры по предупреждению' in str(data.value) or \
                            "о недопустимости нецелевого расхода" in str(data.value):
                        data.alignment = Alignment(wrap_text=True, horizontal='center',
                                                   vertical='center')
                        data.font = Font(name='Arial', size=13, bold=True)

                    else:
                        data.alignment = Alignment(wrap_text=True, horizontal='left',
                                                   vertical='top')
                        data.font = Font(name='Arial', size=12)

                    if not data.value is None:
                        text = data.value
                        for key, value in text_width_dict.items():
                            if value[0] <= len(text) <= value[1]:
                                ws.row_dimensions[i].height = int(key)

                well_data.ins_ind += len(dict_events_gnvp[work_plan]) - 1

                ws.row_dimensions[2].height = 30

                if len(well_data.row_expected) != 0:
                    for i in range(1, len(well_data.row_expected) + 1):  # Добавление  показатели после ремонта
                        ws.row_dimensions[well_data.ins_ind + i - 1].height = None
                        for j in range(1, 12):
                            if i == 1:
                                ws.cell(row=i + well_data.ins_ind, column=j).font = Font(name='Arial', size=13,
                                                                                         bold=True)
                                ws.cell(row=i + well_data.ins_ind, column=j).alignment = Alignment(wrap_text=False,
                                                                                                   horizontal='center',
                                                                                                   vertical='center')
                                ws.cell(row=i + well_data.ins_ind, column=j).value = well_data.row_expected[i - 1][
                                    j - 1]
                            else:
                                ws.cell(row=i + well_data.ins_ind, column=j).font = Font(name='Arial', size=13,
                                                                                         bold=True)
                                ws.cell(row=i + well_data.ins_ind, column=j).alignment = Alignment(wrap_text=False,
                                                                                                   horizontal='left',
                                                                                                   vertical='center')
                                ws.cell(row=i + well_data.ins_ind, column=j).value = well_data.row_expected[i - 1][
                                    j - 1]
                    ws.merge_cells(start_column=2, start_row=well_data.ins_ind + 1, end_column=12,
                                   end_row=well_data.ins_ind + 1)
                    well_data.ins_ind += len(well_data.row_expected)

                self.ins_ind_border = well_data.ins_ind
                MyWindow.create_database_well(self, work_plan)

            return ws

        elif work_plan in ['application_pvr']:
            for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
                for col_ind, col in enumerate(row):
                    if col_ind in [3, 2]:
                        if 'кровля' in str(col).lower():
                            type_pvr = ws.cell(row=row_ind, column=3).value
                            index_row_pvr_begin = row_ind + 1
                        if 'произвести контрольную' in str(col).lower():
                            index_row_pvr_cancel = row_ind
                            if index_row_pvr_begin < index_row_pvr_cancel:
                                well_data.index_row_pvr_list.append(
                                    (index_row_pvr_begin, index_row_pvr_cancel, type_pvr))
                                index_row_pvr_begin, index_row_pvr_cancel = 0, 0
            for pvr in well_data.index_row_pvr_list:
                for row in range(pvr[0], pvr[1]):
                    row_list = []
                    for col in range(2, 9):
                        row_list.append(str(ws.cell(row=row + 1, column=col + 1).value))
                    well_data.pvr_row.append(row_list)

            # print(f'Индексы ПВР {well_data.pvr_row}')

        elif work_plan in ['application_gis']:
            for row_ind, row in enumerate(ws.iter_rows(values_only=True)):
                for col_ind, col in enumerate(row):
                    if col_ind in [3, 2]:
                        if ('задача ' in str(col).lower() or 'техкарт' in str(col).lower() or
                            'задаче №' in str(col).lower()) and \
                                'перфорация' not in str(col).lower() and 'привязка' not in str(col).lower() and \
                                'отбивка' not in str(col).lower():
                            type_pvr = ws.cell(row=row_ind + 1, column=3).value
                            well_data.gis_list.append(type_pvr)

    def add_itog(self, ws, ins_ind, work_plan):

        ws.delete_rows(ins_ind, self.table_widget.rowCount() - ins_ind + 1)
        if work_plan not in ['gnkt_frez', 'application_pvr', 'gnkt_after_grp', 'gnkt_opz']:
            for i in range(ins_ind, len(itog_1(self)) + ins_ind):  # Добавлением итогов
                if i < ins_ind + 6:
                    for j in range(1, 13):
                        ws.cell(row=i, column=j).value = itog_1(self)[i - ins_ind][j - 1]
                        if j != 1:
                            ws.cell(row=i, column=j).border = well_data.thin_border
                            ws.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=False)
                    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=11)
                    ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                   vertical='center')
                else:
                    for j in range(1, 13):
                        ws.row_dimensions[i].height = 50

                        ws.cell(row=i, column=j).value = itog_1(self)[i - ins_ind][j - 1]
                        ws.cell(row=i, column=j).border = well_data.thin_border
                        ws.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=False)
                        ws.cell(row=i, column=2).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                       vertical='center')

                    ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=12)
                    ws.cell(row=i + ins_ind, column=2).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                             vertical='center')

            ins_ind += len(itog_1(self)) + 2

        curator_s = curator_sel(self, well_data.curator, well_data.region)
        # print(f'куратор {curator_sel, well_data.curator}')
        podp_down = pop_down(self, well_data.region, curator_s)

        for i in range(1 + ins_ind, 1 + ins_ind + len(podp_down)):  # Добавлением подписантов внизу
            for j in range(1, 13):
                ws.cell(row=i, column=j).value = podp_down[i - 1 - ins_ind][j - 1]
                ws.cell(row=i, column=j).font = Font(name='Arial', size=13, bold=False)
            if i in [1 + ins_ind + 7, 1 + ins_ind + 8, 1 + ins_ind + 9,
                     1 + ins_ind + 10, 1 + ins_ind + 11,
                     1 + ins_ind + 12, 1 + ins_ind + 13, 1 + ins_ind + 14]:
                ws.merge_cells(start_row=i, start_column=2, end_row=i, end_column=6)
                ws.cell(row=i, column=2).alignment = Alignment(wrap_text=False, vertical='bottom', horizontal='left')
                ws.row_dimensions[i - 1].height = 30

                if i == 1 + ins_ind + 11:
                    ws.row_dimensions[i].height = 55
        ins_ind += len(podp_down)

    def is_valid_date(date):
        try:
            datetime.strptime(date, '%Y-%m-%d')
            return True
        except ValueError:
            return False

#
