import json
import re
import sqlite3
import psycopg2
from PyQt5.QtWidgets import QInputDialog
from collections import namedtuple

from datetime import datetime

from PyQt5 import QtWidgets
from PyQt5.QtSql import QSqlDatabase, QSqlQuery
from PyQt5.QtWidgets import QMessageBox, QTableWidgetItem, QLineEdit, QHeaderView, QVBoxLayout, QMainWindow, QWidget, \
    QTableWidget

from openpyxl import load_workbook

import well_data
from main import MyWindow


class Classifier_well(QMainWindow):
    number_well = None

    def __init__(self, costumer, region, classifier_well, parent=None):

        super(Classifier_well, self).__init__()
        self.centralWidget = QWidget()
        self.setCentralWidget(self.centralWidget)
        self.table_class = QTableWidget()
        self.region = region
        self.costumer = costumer
        self.number_well = None
        if well_data.well_number:
            self.number_well = well_data.well_number._value

        self.setCentralWidget(self.table_class)
        self.model = self.table_class.model()
        if classifier_well == 'classifier_well':
            self.open_to_sqlite_class_well(costumer, region)
        elif classifier_well == 'damping':
            self.open_to_sqlite_without_juming(costumer, region)

    def open_to_sqlite_without_juming(self, costumer, region):
        layout = QVBoxLayout()
        self.edit_well_number = QLineEdit()

        self.edit_well_number.setPlaceholderText("Ввести номер скважины для фильтрации")

        self.edit_well_number.textChanged.connect(self.filter)
        self.edit_well_number.setText(self.number_well)
        layout.addWidget(self.edit_well_number)

        data = self.get_data_from_db(region)

        self.table_class.setColumnCount(len(data[0]))
        self.table_class.setRowCount(len(data))
        self.table_class.setCellWidget(0, 0, self.edit_well_number)
        for row in range(len(data)):
            for col in range(len(data[row])):
                item = QTableWidgetItem(str(data[row][col]))
                self.table_class.setItem(row + 1, col, item)

        self.table_class.setHorizontalHeaderLabels(['номер скважины', 'площадь', 'Текущий квартал'])
        self.table_class.horizontalHeader().setStretchLastSection(True)
        self.table_class.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # layout.addWidget(table)
        self.setLayout(layout)

    def open_to_sqlite_class_well(self, costumer, region):
        layout = QVBoxLayout()
        self.edit_well_number = QLineEdit()
        self.edit_well_number.setPlaceholderText("Ввести номер скважины для фильтрации")

        self.edit_well_number.textChanged.connect(self.filter_class)
        self.edit_well_number.setText(self.number_well)
        layout.addWidget(self.edit_well_number)

        self.edit_well_area = QLineEdit()
        self.edit_well_area.setPlaceholderText("Ввести площадь для фильтрации")
        self.edit_well_area.textChanged.connect(self.filter_class_area)
        layout.addWidget(self.edit_well_area)
        region = f'{region}_классификатор'
        # print(region)
        data = self.get_data_from_class_well_db(region)
        # print(data)

        self.table_class.setColumnCount(len(data[0]))
        self.table_class.setRowCount(len(data))
        self.table_class.setCellWidget(0, 1, self.edit_well_number)
        self.table_class.setCellWidget(0, 2, self.edit_well_area)
        for row in range(len(data)):
            for col in range(len(data[row])):
                if col in [5, 6, 10, 11, 13]:
                    if str(data[row][col]).replace('.', '').isdigit() and str(data[row][col]).count('.') < 2:
                        item = QTableWidgetItem(str(round(float(data[row][col]), 1)))
                    else:
                        item = QTableWidgetItem(str(data[row][col]))
                elif col in [9]:
                    if str(data[row][col]).replace('.', '').isdigit() and str(data[row][col]).count('.') < 2:
                        item = QTableWidgetItem(str(round(float(data[row][col]), 7)))
                    else:
                        item = QTableWidgetItem(str(data[row][col]))
                # elif col in [7]:
                #     date_string = data[row][col]
                #     date_format = "%Y-%m-%d %H:%M"
                #     date = datetime.strptime(date_string, date_format)
                #     print(date)
                #     item = QTableWidgetItem(str(date))
                else:
                    item = QTableWidgetItem(str(data[row][col]))
                self.table_class.setItem(row + 1, col, item)

        self.table_class.setHorizontalHeaderLabels(
            ['ЦДНГ', 'номер скважины', 'площадь', 'Месторождение', 'Категория \n по Рпл',
             'Ргд', 'Рпл', 'Дата замера', 'категория \nH2S', 'H2S-%', "H2S-мг/л",
             "H2S-мг/м3", 'Категория по газу', "Газовый фактор", "версия от"])
        self.table_class.horizontalHeader().setStretchLastSection(True)
        self.table_class.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        # layout.addWidget(table)
        self.setLayout(layout)

    def get_data_from_db(self, region):

        # Параметры подключения к PostgreSQL
        try:
            # Создание подключения к базе данных PostgreSQL
            conn = psycopg2.connect(**well_data.postgres_params_classif)

            # Выполнение SQL-запроса для получения данных
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT well_number, deposit_area, today
                    FROM {region};
                """)
                data = cur.fetchall()

        except psycopg2.Error as e:
            mes = QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')

        finally:
            if conn:
                conn.close()

        return data

    def get_data_from_class_well_db(self, region):
        # Параметры подключения к PostgreSQL

        try:
            # Создание подключения к базе данных PostgreSQL
            conn = psycopg2.connect(**well_data.postgres_params_classif)

            # Выполнение SQL-запроса для получения данных
            with conn.cursor() as cur:
                cur.execute(f"""
                    SELECT cdng, well_number, deposit_area, oilfield, categoty_pressure,
                           pressure_Ppl, pressure_Gst, date_measurement, categoty_h2s,
                           h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor, today
                    FROM {region};
                """)
                data = cur.fetchall()

        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
            mes = QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')

            return []
        finally:
            if conn:
                conn.close()

        return data

    def export_to_sqlite_without_juming(self, fname, costumer, region):
        try:
            # Подключение к базе данных
            conn = psycopg2.connect(**well_data.postgres_params_classif)
            cursor = conn.cursor()
            region_list = ['ЧГМ', 'АГМ', 'ТГМ', 'ИГМ', 'КГМ', ]

            for region_name in region_list:
                if region_name == region:
                    # # Удаление всех данных из таблицы
                    # cursor.execute("DROP TABLE my_table")

                    # Удаление всех данных из таблицы
                    # cursor.execute(f"DELETE FROM {region_name}")

                    # Создание таблицы в базе данных
                    cursor.execute(f'CREATE TABLE IF NOT EXISTS {region_name}'
                                   f'(well_number TEXT,'
                                   f'deposit_area TEXT, '
                                   f'today TEXT,'
                                   f'region TEXT,'
                                   f'costumer TEXT)')

                    # Загрузка файла Excel
                    wb = load_workbook(fname)
                    ws = wb.active

                    # Получение данных из Excel и запись их в базу данных
                    for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                        for col, value in enumerate(row):
                            if not value is None and col <= 18:
                                # print(value)
                                if 'туймазин' in str(value).lower():
                                    check_param = 'ТГМ'
                                if 'ишимбай' in str(value).lower():
                                    check_param = 'ИГМ'
                                if 'чекмагуш' in str(value).lower():
                                    check_param = 'ЧГМ'
                                if 'красно' in str(value).lower():
                                    check_param = 'КГМ'
                                if 'арлан' in str(value).lower():
                                    check_param = 'АГМ'
                                if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                        value) or '01.10.' in str(value):

                                    version_year = re.findall(r'[0-9.]', str(value))
                                    version_year = ''.join(version_year)
                                    if version_year[-1] == '.':
                                        version_year = version_year[:-1]
                        if index_row > 18:
                            break
                    # print(region_name, version_year)
                    # print(check_param)
                    if check_param == region_name:
                        mes = QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                  f'регион выбрано корректно  {region_name}')
                        try:
                            # Получение данных из Excel и запись их в базу данных
                            for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                if 'ПЕРЕЧЕНЬ' in row:
                                    check_file = True
                                if 'Скважина' in row:
                                    area_row = index_row + 2
                                    for col, value in enumerate(row):
                                        if not value is None and col <= 20:
                                            if 'Скважина' == value:
                                                well_column = col
                                            elif 'Площадь' == value:
                                                area_column = col

                            for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                if index_row > area_row:

                                    well_number = row[well_column]
                                    area_well = row[area_column]

                                    if well_number:
                                        cursor.execute(
                                            f"INSERT INTO {region_name} (well_number, deposit_area, today, region, costumer) "
                                            f"VALUES (%s, %s, %s, %s,%s)",
                                            (well_number, area_well, version_year, region_name, costumer))

                            mes = QMessageBox.information(self, 'данные обновлены', 'Данные обновлены')
                        except:
                            mes = QMessageBox.warning(self, 'ОШИБКА', 'Выбран файл с не корректными данными')

                    else:
                        mes = QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                  f'в Данном перечне отсутствую скважины {region_name}')

            # Сохранение изменений
            conn.commit()

        except psycopg2.Error as e:
            # Выведите сообщение об ошибке
            mes = QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')
        finally:
            # Закройте курсор и соединение
            if cursor:
                cursor.close()
            if conn:
                conn.close()


    def filter(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(0, 1, 2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def filter_class(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(1, 2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break

    def filter_class_area(self, filter_text):
        for i in range(1, self.table_class.rowCount() + 1):
            for j in range(2):
                item = self.table_class.item(i, j)
                if item:
                    match = filter_text.lower() not in item.text().lower()
                    self.table_class.setRowHidden(i, match)
                    if not match:
                        break



    def export_to_sqlite_class_well(self, fname, costumer, region):
        # Параметры подключения к PostgreSQL

        region_list = ['ЧГМ_классификатор', 'АГМ_классификатор', 'ТГМ_классификатор', 'ИГМ_классификатор',
                       'КГМ_классификатор']

        try:
            # Создание подключения к базе данных PostgreSQL
            conn = psycopg2.connect(**well_data.postgres_params_classif)
            cursor = conn.cursor()

            for region_name in region_list:
                if region in region_name:
                    # Удаление всех данных из таблицы (опционально)
                    cursor.execute(f"DROP TABLE IF EXISTS {region_name};")

                    # Создание таблицы, если она не существует
                    cursor.execute(f"""
                        CREATE TABLE IF NOT EXISTS {region_name} (
                            ID SERIAL PRIMARY KEY NOT NULL,
                            cdng TEXT,
                            well_number TEXT,
                            deposit_area TEXT,
                            oilfield TEXT,
                            categoty_pressure TEXT,
                            pressure_Ppl TEXT,
                            pressure_Gst TEXT,
                            date_measurement TEXT,
                            categoty_h2s TEXT,
                            h2s_pr TEXT,
                            h2s_mg_l TEXT,
                            h2s_mg_m TEXT,
                            categoty_gf TEXT,
                            gas_factor TEXT,
                            today TEXT,
                            region TEXT,
                            costumer TEXT
                        );
                    """)

                    # Загрузка файла Excel
                    wb = load_workbook(fname)
                    ws = wb.active

                    # Определение столбцов
                    well_column, cdng, area_column, oilfield, categoty_pressure = None, None, None, None, None
                    pressure_Gst, date_measurement, pressure_Ppl, categoty_h2s = None, None, None, None
                    h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor = None, None, None, None, None
                    area_row = None
                    check_file = False

                    for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                        if 'Классификация' in row:
                            check_file = True
                        if 'Скважина' in row:
                            area_row = index_row + 2
                            for col, value in enumerate(row):
                                if not value is None and col <= 20:
                                    if 'Скважина' == value:
                                        well_column = col
                                    elif 'Цех' == value:
                                        cdng = col
                                    elif 'Площадь' == value:
                                        area_column = col
                                    elif 'Месторождение' == value:
                                        oilfield = col
                                    elif 'Пластовое давление' == value:
                                        categoty_pressure = col
                                        pressure_Gst = col + 1
                                        date_measurement = col + 2
                                        pressure_Ppl = col + 3
                                    elif 'содержание сероводорода' in str(value).lower():
                                        categoty_h2s = col
                                        h2s_pr = col + 1
                                        h2s_mg_l = col + 2
                                        h2s_mg_m = col + 3
                                    elif 'Газовый фактор' == value:
                                        categoty_gf = col
                                        gas_factor = col + 1
                                    if 'туймазин' in str(value).lower():
                                        check_param = 'ТГМ'
                                    if 'ишимбай' in str(value).lower():
                                        check_param = 'ИГМ'
                                    if 'чекмагуш' in str(value).lower():
                                        check_param = 'ЧГМ'
                                    if 'красно' in str(value).lower():
                                        check_param = 'КГМ'
                                    if 'арлан' in str(value).lower():
                                        check_param = 'АГМ'
                                    if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                            value) or '01.10.' in str(value):

                                        version_year = re.findall(r'[0-9.]', str(value))
                                        version_year = ''.join(version_year)
                                        if version_year[-1] == '.':
                                            version_year = version_year[:-1]

                    if check_param == region_name:
                        mes = QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                  f'регион выбрано корректно  {region_name}')

                        try:
                            # Вставка данных в таблицу
                            for index_row, row in enumerate(ws.iter_rows(min_row=2, values_only=True)):
                                if index_row > area_row and check_file:
                                    well_number = row[well_column]
                                    area_well = row[area_column]
                                    oilfield_str = row[oilfield]
                                    version_year = None

                                    for col, value in enumerate(row):
                                        if not value is None and col <= 18:
                                            if '01.01.' in str(value) or '01.04.' in str(value) or '01.07.' in str(
                                                    value) or '01.10.' in str(value):
                                                version_year = re.findall(r'[0-9.]', str(value))
                                                version_year = ''.join(version_year)
                                                if version_year[-1] == '.':
                                                    version_year = version_year[:-1]

                                    if well_number:
                                        cursor.execute(f"""
                                            INSERT INTO {region_name} (
                                                cdng, well_number, deposit_area, oilfield,
                                                categoty_pressure, pressure_Ppl, pressure_Gst, date_measurement,
                                                categoty_h2s, h2s_pr, h2s_mg_l, h2s_mg_m, categoty_gf, gas_factor,
                                                today, region, costumer
                                            )
                                            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s);
                                        """, (
                                            row[cdng], well_number, area_well, oilfield_str, row[categoty_pressure],
                                            row[pressure_Ppl], row[pressure_Gst], row[date_measurement], row[categoty_h2s],
                                            row[h2s_pr], row[h2s_mg_l], row[h2s_mg_m], row[categoty_gf], row[gas_factor],
                                            version_year, region, costumer
                                        ))
                        except:
                            mes = QMessageBox.warning(self, 'ОШИБКА', 'Выбран файл с не корректными данными')

                    else:
                        mes = QMessageBox.warning(self, 'ВНИМАНИЕ ОШИБКА',
                                                  f'в Данном перечне отсутствую скважины {region_name}')
                    conn.commit()

        except (psycopg2.Error, Exception) as e:
            # Выведите сообщение об ошибке
            mes = QMessageBox.warning(self, 'Ошибка', 'Ошибка подключения к базе данных')
        finally:
            # Закройте курсор и соединение
            if cursor:
                cursor.close()
            if conn:
                conn.close()



def read_database_gnkt(contractor, gnkt_number):
    try:
        # Подключение к базе данных
        conn = psycopg2.connect(**well_data.postgres_conn_gnkt)

        if 'ойл-сервис' in contractor.lower():
            contractor = 'oil_service'
        cursor = conn.cursor()
        cursor.execute(f"SELECT * FROM КГМ WHERE today =(%s)", (gnkt_number, '1963'))
        print(f' база данных открыта')
        result = cursor.fetchone()
    except psycopg2.Error as e:
        # Выведите сообщение об ошибке
        mes = QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных')
    finally:
        # Закройте курсор и соединение
        if cursor:
            cursor.close()
        if conn:
            conn.close()
    # print(result)


def create_database_well_db(work_plan, number_dp):
    # print(row, well_data.count_row_well)
    try:

        conn = psycopg2.connect(**well_data.postgres_conn_work_well)
        cursor = conn.cursor()
        if number_dp == 0:
            number_dp =''

        # Создаем таблицу для хранения данных
        number = json.dumps(str(well_data.well_number._value) + well_data.well_area._value + work_plan + str(number_dp),
                            ensure_ascii=False)

        # Попытка удалить таблицу, если она существует
        cursor.execute(f'DROP TABLE IF EXISTS {number}')

        cursor.execute(f'CREATE TABLE IF NOT EXISTS {number}'
                       f'(index_row INTEGER,'
                       f'current_bottom FLOAT,'
                       f'perforation TEXT, '
                       f'plast_all TEXT, '
                       f'plast_work TEXT, '
                       f'leakage TEXT,'
                       f'column_additional TEXT,'
                       f'fluid TEXT,'
                       f'category_pressuar TEXT,'
                       f'category_h2s TEXT,'
                       f'category_gf TEXT,'
                       f'template_depth FLOAT,'
                       f'skm_list TEXT,'
                       f'problemWithEk_depth FLOAT,'
                       f'problemWithEk_diametr FLOAT)')

        for index, data in enumerate(well_data.data_list):
            current_bottom = data[1]
            dict_perforation_json = data[2]
            plast_all = data[3]
            plast_work = data[4]
            dict_leakiness = data[5]
            column_additional = data[6]
            fluid_work = data[7]
            template_depth = int(data[11])
            skm_interval = data[12]
            problemWithEk_depth = data[13]
            problemWithEk_diametr = data[14]

            # Подготовленные данные для вставки (пример)
            data_values = (index, current_bottom, dict_perforation_json, plast_all, plast_work,
                           dict_leakiness, column_additional, fluid_work, well_data.category_pressuar,
                           well_data.category_h2s, well_data.category_gf, template_depth, skm_interval,
                           problemWithEk_depth, problemWithEk_diametr)

            # Подготовленный запрос для вставки данных с параметрами
            query = f"INSERT INTO {number} VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"

            # Выполнение запроса с использованием параметров
            cursor.execute(query, data_values)


        # Сохранить изменения и закрыть соединение
        conn.commit()
    except psycopg2.Error as e:
        # Выведите сообщение об ошибке
        mes = QMessageBox.warning(None, 'Ошибка', 'Ошибка подключения к базе данных, Скважине не добавлена в базу')
    finally:
        # Закройте курсор и соединение
        if cursor:
            cursor.close()
        if conn:
            conn.close()

        mes = QMessageBox.information(None, 'база данных', 'Скважина добавлена в базу данных')

if __name__ == "__main__":
    import sys

    app = QtWidgets.QApplication(sys.argv)
    # app.setStyleSheet()
    window = Classifier_well()
    window.show()
    sys.exit(app.exec_())