from copy import copy

from openpyxl.styles import Alignment
from openpyxl.utils.cell import range_boundaries, get_column_letter

import well_data


def delete_rows_pz(self, ws):


    boundaries_dict = {}

    for ind, _range in enumerate(ws.merged_cells.ranges):
        boundaries_dict[ind] = range_boundaries(str(_range))

    # rowHeights_top = [None, 18.0, 18, 18,None, 18.0, 18, 18,None, 18.0, 18, 18, 18.0, 18, 18, 18.0, 18, 18, 18.0, 18, 18]
    rowHeights1 = [ws.row_dimensions[i + 1].height for i in range(well_data.cat_well_min._value, ws.max_row)]
    for key, value in boundaries_dict.items():
        ws.unmerge_cells(start_column=value[0], start_row=value[1],
                         end_column=value[2], end_row=value[3])
    # print(f'индекс удаления {1, well_data.cat_well_min - 1} , {well_data.data_well_max + 2, ws.max_row - well_data.data_well_max}')

    ws.delete_rows(well_data.data_x_max._value, ws.max_row - well_data.data_x_max._value)

    ws.delete_rows(1, well_data.cat_well_min._value - 1)

    # print(sorted(boundaries_dict))
    well_data.rowHeights = rowHeights1
    # print(rowHeights1[well_data.cat_well_min:])
    # print(len(well_data.rowHeights))
    # print(f'251po {16}')
    for _ in range(16):
        ws.insert_rows(1, 1)
    for key, value in boundaries_dict.items():
        if value[1] <= well_data.data_well_max._value + 1 and value[1] >= well_data.cat_well_min._value:
            ws.merge_cells(start_column=value[0], start_row=value[1] + 16 - well_data.cat_well_min._value + 1,
                           end_column=value[2], end_row=value[3] + 16 - well_data.cat_well_min._value + 1)

    # print(f'{ws.max_row, len(well_data.rowHeights)}dd')
    for index_row, row in enumerate(ws.iter_rows()):  # Копирование высоты строки
        ws.row_dimensions[index_row + 17].height = well_data.rowHeights[index_row - 1]


def head_ind(start, finish):
    return f'A{start}:L{finish}'


def copy_row(ws, ws2, head):
    boundaries_dict = {}

    for ind, _range in enumerate(ws.merged_cells.ranges):
        boundaries_dict[ind] = range_boundaries(str(_range))

    rowHeights1 = [ws.row_dimensions[i + 1].height for i in range(ws.max_row)]
    colWidth = [ws.column_dimensions[get_column_letter(i + 1)].width for i in range(0, 13)] + [None]
    for key, value in boundaries_dict.items():
        ws.unmerge_cells(start_column=value[0], start_row=value[1],
                         end_column=value[2], end_row=value[3])
    copy_true_ws(ws, ws2, head)



    # print(f'Вставлены данные по скважине')
    for key, value in boundaries_dict.items():
       ws2.merge_cells(start_column=value[0], start_row=value[1],
                           end_column=value[2], end_row=value[3])

    for index_row, row in enumerate(ws.iter_rows()):  # Копирование высоты строки
        ws2.row_dimensions[index_row].height = rowHeights1[index_row]
        if index_row == 2:
            for col_ind, col in enumerate(row):
                if col_ind <= 12:
                    ws2.column_dimensions[get_column_letter(col_ind + 1)].width = colWidth[col_ind]

def copy_true_ws(ws, ws2, head):
    for row_number, row in enumerate(ws[head]):
        for col_number, cell in enumerate(row):
            # print(cell.value)
            if 'катег' in str(cell.value).lower():
                ws2.cell(row=row_number+1, column=col_number+1).alignment = Alignment(wrap_text=True, horizontal='left',
                                                                            vertical='center')

            if type(cell.value) == float:
                ws2.cell(row_number + 1, col_number + 1, round(cell.value, 5))
            else:
                ws2.cell(row_number + 1, col_number + 1, cell.value)


            if cell.has_style:
                ws2.cell(row_number + 1, col_number + 1).font = copy(cell.font)
                ws2.cell(row_number + 1, col_number + 1).fill = copy(cell.fill)
                ws2.cell(row_number + 1, col_number + 1).border = copy(cell.border)
                ws2.cell(row_number + 1, col_number + 1).number_format = copy(
                    cell.number_format)
                ws2.cell(row_number + 1, col_number + 1).protection = copy(
                    cell.protection)
                ws2.cell(row_number + 1, col_number + 1).alignment = copy(
                    cell.alignment)
                ws2.cell(row_number + 1, col_number + 1).quotePrefix = copy(
                    cell.quotePrefix)
                ws2.cell(row_number + 1, col_number + 1).pivotButton = copy(
                    cell.pivotButton)

